"""
Fichier qui regroupe des tâches de gestion des intervenants au sein
d'une UV : fichier Excel du nombre d'heures théoriques, décompte des
heures remplacées.
"""

import os
import re
import pandas as pd
from pandas.api.types import CategoricalDtype

from .utc import UtcUvListToCsv
from ..utils_config import Output, selected_uv
from ..utils import lib_list, rel_to_dir
from .base import UVTask, TaskBase
from ..scripts.excel_hours import create_excel_file

import openpyxl
from ..openpyxl_patched import fixit
fixit(openpyxl)

from ..openpyxl_utils import get_range_from_cells, fill_row


def create_insts_list(df):
    "Agrège les données d'affectation des Cours/TD/TP"

    def course_list(e):
        "Return course list like C1, D2, T1A"
        return ', '.join(sorted(e, key=lib_list))

    def score(libs):
        "Renvoie un tuple comptant les types de cours Cours/TD/TP"
        sc = [0, 0, 0]
        mapping = {'C': 0, 'D': 1, 'T': 2}
        for lib in libs:
            m = re.search('([CDT])[0-9]*([AB]?)', lib)
            if m:
                ix = mapping[m.group(1)]
                if m.group(2):
                    sc[ix] += .5
                else:
                    sc[ix] += 1
            else:
                raise Exception(f"L'identifiant {lib} n'est pas matché")
        return tuple(sc)

    def myapply(df):
        e = df['Lib. créneau'] + df['Semaine'].fillna('')
        resp = int(df['Responsable'].sum())
        s = score(e)
        return pd.Series({
            'CourseList': course_list(e),
            'SortCourseList': s,
            'Cours': s[0],
            'TD': s[1],
            'TP': s[2],
            'Responsable': resp
        })

    df = df.groupby('Intervenants')
    df = df.apply(myapply)
    df = df.reset_index()

    return(df)


class XlsInstructors(TaskBase):
    """Fichier de détails global des intervenants toutes UV confondues

    Il sert à la tâche html_inst pour générer un descriptif des
    intervenants d'une UV et à la tâche xls_utp pour le calcul des
    UTP effectuées et des remplacements.
    """

    target_dir = "documents"
    target_name = "intervenants.xlsx"

    def setup(self):
        super().setup()
        self.target = self.build_target()

    def run(self):
        if not os.path.exists(self.target):
            print(
                "Le fichier '{}' n'existe pas, création d'un fichier vide".format(
                    rel_to_dir(self.target, self.settings.SEMESTER_DIR)
                )
            )
            columns = ["Intervenants", "Statut", "Email", "Website"]
            with Output(self.target) as target:
                pd.DataFrame(columns=columns).to_excel(target())


class AddInstructors(TaskBase):
    """Ajoute les intervenants dans la liste csv des créneaux"""

    target_dir = "generated"
    target_name = "UTC_UV_list_instructors.csv"

    def setup(self):
        super().setup()
        self.uv_list = UtcUvListToCsv.target_from()
        self.target = self.build_target()
        self.affectations = [
            (uv, XlsAffectation.target_from(**info))
            for planning, uv, info in selected_uv()
        ]
        files = [f for _, f in self.affectations]
        self.file_dep = files + [self.uv_list]

    def run(self):
        df_csv = pd.read_csv(self.uv_list)
        df_csv.Semaine = df_csv.Semaine.astype(object)

        df_affs = [
            pd.read_excel(xls_aff, engine="openpyxl").assign(**{"Code enseig.": uv})
            for uv, xls_aff in self.affectations
        ]

        df_aff = pd.concat(df_affs, ignore_index=True)
        df_aff.Semaine = df_aff.Semaine.astype(object)

        df_merge = pd.merge(
            df_csv,
            df_aff,
            how="left",
            on=[
                "Code enseig.",
                "Jour",
                "Heure début",
                "Heure fin",
                "Semaine",
                "Lib. créneau",
                "Locaux",
            ],
        )

        with Output(self.target) as target:
            df_merge.to_csv(target(), index=False)


def read_xls_details(fn):
    """Lit un fichier Excel avec un ordre sur la colonne 'Statut'."""

    sts = ["MCF", "PR", "PRAG", "PRCE", "PAST", "ECC", "Doct", "ATER", "Vacataire"]
    status_type = CategoricalDtype(categories=sts, ordered=True)

    return pd.read_excel(fn, engine="openpyxl", dtype={
        'Statut': status_type
    })


class XlsInstDetails(UVTask):
    """Fichier Excel des intervenants par UV avec détails

    Les détails sont pris dans le fichier de détails global. Les
    affectations sont prises pour chaque UV.
    """

    target_dir = "generated"
    target_name = "intervenants_details.xlsx"

    def setup(self):
        super().setup()
        self.insts_details = XlsInstructors.target_from()
        self.inst_uv = XlsAffectation.target_from(**self.info)
        self.target = self.build_target()
        self.file_dep = [self.inst_uv, self.insts_details]

    def run(self):
        inst_uv = pd.read_excel(self.inst_uv, engine="openpyxl")
        insts_details = pd.read_excel(self.insts_details, engine="openpyxl")

        # Add details from inst_details
        df = inst_uv.merge(
            insts_details, how="left", left_on="Intervenants", right_on="Intervenants"
        )

        with Output(self.target) as target:
            df.to_excel(target(), index=False)


class XlsUTP(UVTask):
    """Crée un document Excel pour calcul des heures et remplacements."""

    target_dir = "generated"
    target_name = "remplacement.xlsx"

    def setup(self):
        super().setup()
        self.xls = XlsAffectation.target_from(**self.info)
        self.insts = XlsInstructors.target_from()
        self.target = self.build_target()
        self.file_dep = [self.xls, self.insts]

    def run(self):
        df = pd.read_excel(self.xls, engine="openpyxl")

        # Add details
        df_details = read_xls_details(self.insts)

        if df["Intervenants"].isnull().all():
            raise Exception(
                "Pas d'intervenants renseignés dans le fichier %s" % self.xls
            )
        else:
            df_insts = create_insts_list(df)

        # Add details from df_details
        df = df_insts.merge(
            df_details, how="left", left_on="Intervenants", right_on="Intervenants"
        )

        dfs = df.sort_values(
            ["Responsable", "Statut", "SortCourseList"], ascending=False
        )
        dfs = dfs.reset_index()

        with Output(self.target, protected=True) as target:
            create_excel_file(target(), dfs)


class XlsAffectation(UVTask):
    """Fichier Excel des créneaux de toutes les UV configurées."""

    unique_uv = False
    target_name = "intervenants.xlsx"
    target_dir = "documents"

    def setup(self):
        super().setup()
        self.uvlist_csv = UtcUvListToCsv.target_from()
        self.target = self.build_target()
        self.file_dep = [self.uvlist_csv]

    def run(self):
        df = pd.read_csv(self.uvlist_csv)
        df_uv = df.loc[df["Code enseig."] == self.uv, :]

        selected_columns = [
            "Jour",
            "Heure début",
            "Heure fin",
            "Locaux",
            "Semaine",
            "Lib. créneau",
        ]
        df_uv_select = df_uv[selected_columns]
        df_uv_select = df_uv_select.sort_values(["Lib. créneau", "Semaine"])

        df_uv_select["Intervenants"] = ""
        df_uv_select["Responsable"] = ""

        # Copy for modifications
        with Output(self.target, protected=True) as target:
            df_uv_select.to_excel(target(), sheet_name="Intervenants", index=False)

        N = 10
        workbook = openpyxl.load_workbook(target())
        worksheet = workbook.create_sheet("Décompte des heures")

        ref_cell = worksheet.cell(2, 2)

        keywords = [
            "Intervenants",
            "Statut",
            "Cours",
            "TD",
            "TP",
            "Heures Cours prév",
            "Heures TD prév",
            "Heures TP prév",
            "UTP",
            "Heure équivalent TD",
            "Heure brute"
        ]

        fill_row(ref_cell, *keywords)

        def get_cell_in_row(ref_cell, *keywords):
            def func(cell, keyword):
                if keyword in keywords:
                    return cell.parent.cell(
                        row=cell.row,
                        column=ref_cell.col_idx + keywords.index(keyword)
                    )
                else:
                    raise Exception

            return func

        cell_in_row = get_cell_in_row(ref_cell, *keywords)

        for i in range(N):
            cell = ref_cell.below(i + 1)
            elts = [
                None,
                None,
                None,
                None,
                None,
                lambda cell: "=2*16*{}".format(cell_in_row(cell, "Cours").coordinate),
                lambda cell: "=2*16*{}".format(cell_in_row(cell, "TD").coordinate),
                lambda cell: "=2*16*{}".format(cell_in_row(cell, "TP").coordinate),
                lambda cell: "=2*16*2.25*{}+2*16*1.5*{}+2*16*{}*{}".format(
                    cell_in_row(cell, "Cours").coordinate,
                    cell_in_row(cell, "TD").coordinate,
                    cell_in_row(cell, "TP").coordinate,
                    cell_in_row(cell, "Statut").coordinate,
                ),
                lambda cell: "=2/3*{}".format(
                    cell_in_row(cell, "UTP").coordinate,
                ),
                lambda cell: "=2*16*{}+2*16*{}+2*16*{}".format(
                    cell_in_row(cell, "Cours").coordinate,
                    cell_in_row(cell, "TD").coordinate,
                    cell_in_row(cell, "TP").coordinate,
                )
            ]
            fill_row(cell, *elts)

        total_cell = ref_cell.below(N+1)
        expected_cell = ref_cell.below(N+2)

        n_cours = len(df_uv.loc[df["Activité"] == "Cours"])
        n_TD = len(df_uv.loc[df["Activité"] == "TD"])
        n_TP = len(df_uv.loc[df["Activité"] == "TP"])

        first_cours = ref_cell.below().right(2)
        first_TD = ref_cell.below().right(3)
        first_TP = ref_cell.below().right(4)

        last_cours = ref_cell.below(N).right(2)
        last_TD = ref_cell.below(N).right(3)
        last_TP = ref_cell.below(N).right(4)

        range_ = get_range_from_cells(first_cours, last_cours)
        total_cell.right(2).text(f"=SUM({range_})")

        range_ = get_range_from_cells(first_TD, last_TD)
        total_cell.right(3).text(f"=SUM({range_})")

        range_ = get_range_from_cells(first_TP, last_TP)
        total_cell.right(4).text(f"=SUM({range_})")

        expected_cell.right().text(n_cours)
        expected_cell.right(2).text(n_TD)
        expected_cell.right(3).text(n_TP)

        workbook.save(target())
