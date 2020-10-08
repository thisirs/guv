#!/usr/bin/env python3

__doc__ = """
Create an Excel file to compute grades
"""

import os
import sys
import math
import argparse
from collections import OrderedDict

from schema import Schema, And, Use, Or, Optional

import openpyxl
from .openpyxl_patched import fixit
fixit(openpyxl)

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

from .openpyxl_utils import (
    frame_range,
    get_address_of_cell,
    get_range_from_cells,
    get_segment,
    row_and_col,
    fit_cells_at_col,
    generate_ranges
)

import pandas as pd
import oyaml as yaml  # Ordered yaml

from .utils import sort_values


def walk_tree(tree, depth=None):
    "Generate coordinates of cells to represent a tree"

    def compute_depth(tree):
        if isinstance(tree, (OrderedDict, dict)):
            return 1 + max(compute_depth(child) for child in tree.values())
        else:
            return 1

    if depth is None:
        depth = compute_depth(tree)

    class RunOut(Exception):
        pass

    def walk_tree0(name, tree, current_depth, nleaves):
        if isinstance(tree, (OrderedDict, dict)):  # Inner node
            current_leaves = 0
            sdepth = None
            x_span = None
            for key, child in tree.items():
                try:
                    yield from walk_tree0(
                        key, child, current_depth + 1, nleaves + current_leaves
                    )
                except RunOut as e:
                    nleaf, sdepth = e.args
                    current_leaves += nleaf
                    x_span = math.ceil(sdepth // (current_depth + 1))
                    sdepth = sdepth - x_span
            if current_depth != 0:
                yield name, nleaves, sdepth, current_leaves, x_span
                raise RunOut(current_leaves, sdepth)
        else:  # Leaf
            x_span = math.ceil((depth + 1) // (current_depth + 1))
            sdepth = depth - x_span
            yield name, nleaves, sdepth, 1, x_span
            raise RunOut(1, sdepth)

    yield from walk_tree0("root", tree, 0, 0)


class CliArgsMixin:
    def add_arguments(self):
        self._parser = argparse.ArgumentParser(add_help=False)

    def add_argument(self, *args, **kwargs):
        self._parser.add_argument(*args, **kwargs)

    def parse_arguments(self):
        self.args = self._parser.parse_args(self.argv)


class GroupCliOpt(CliArgsMixin):
    """Classe pour la spécification d'une colonne dénotant des sous-groupes"""

    def add_arguments(self):
        super().add_arguments()
        self.add_argument("--worksheets", dest="group_by")

    def get_columns(self, **kwargs):
        columns = super().get_columns(**kwargs)

        if self.args.group_by is not None:
            columns.append((self.args.group_by, "raw", 5))

        return columns


class GroupCliOpt(CliArgsMixin):
    def add_arguments(self):
        super().add_arguments()
        self.add_argument("--group", dest="group_by")


class ConfigCliOpt(CliArgsMixin):
    """Classe pour la spécification d'un fichier de configuration"""

    config_argname = "--config"

    def add_arguments(self):
        super().add_arguments()
        self.add_argument(self.config_argname, dest="config_file", required=True)

    @property
    def config(self):
        if not hasattr(self, "_config") or self._config is None:
            self._config = self.parse_config()
        return self._config

    def validate_config(self, config):
        return config

    def parse_config(self):
        config_file = self.args.config_file
        if not os.path.exists(config_file):
            raise Exception(f"Configuration file '{config_file}' not found")

        with open(config_file, "r") as stream:
            config = list(yaml.load_all(stream, Loader=yaml.SafeLoader))[0]
            config = self.validate_config(config)
            return config


class FirstGradeSheet(CliArgsMixin):
    """Classe de base pour la création d'une feuille de notes"""

    def __init__(self, argv=sys.argv[1:]):
        self.argv = argv

    def add_arguments(self):
        super().add_arguments()
        self.add_argument("--name", required=True, help="Nom de la feuille de notes")
        self.add_argument("-d", "--data", dest="data_file", help=""),
        self.add_argument("-o", "--output-file")
        self.add_argument(
            "--order-by",
            required=False,
            help="Colonne utilisée pour ordonner la liste des étudiants "
            "dans les feuilles de notations",
        )

    def setup(self):
        # Setting source of grades
        if self.args.data_file:
            if os.path.isdir(self.args.data_file):
                fn = "student_data_merge.xlsx"
                self.data_file = os.path.join(self.args.data_file, fn)
            else:
                self.data_file = self.args.data_file
        else:
            self.data_file = "student_data_merge.xlsx"

        # Reading source of information into a Pandas dataframe
        if not os.path.exists(self.data_file):
            raise Exception(f"Data file `{self.data_file}` does not exist")
        self.data_df = pd.read_excel(self.data_file)

        # Setting path of gradebook file to be written
        if self.args.output_file:
            if os.path.isdir(self.args.output_file):
                fn = f"{self.args.name}_gradebook.xlsx"
                self.output_file = os.path.join(self.args.output_file, fn)
            else:
                self.output_file = self.args.output_file
        else:
            self.output_file = f"{self.args.name}_gradebook.xlsx"

    def get_columns(self, **kwargs):
        "Renvoie les colonnes à utiliser dans la première feuille de calculs"

        # Default columns
        columns = [("Nom", "raw", 0), ("Prénom", "raw", 0), ("Courriel", "raw", 0)]

        # Add order_by columns if specified in command line and don't
        # include it in worksheet
        if self.args.order_by is not None:
            columns.append((self.args.order_by, "hide", 5))

        # Add column for grades
        columns.append((self.args.name, "cell", 100))

        return columns

    def create_first_worksheet(self):
        # Create workbook and first worksheet named "data"
        self.workbook = Workbook()
        self.first_ws = self.workbook.active
        self.first_ws.title = "data"

        # Pandas dataframe that mirrors the first worksheet
        self.first_df = pd.DataFrame()

        # Get columns to be copied from source of information DATA_DF
        # to first worksheet
        columns = self.get_columns()

        # Sort columns according to priority and index in columns
        columns = [(i, *rest) for i, rest in enumerate(columns)]

        def sort_func(e):
            i, name, type, priority = e
            return priority, i

        columns = sorted(columns, key=sort_func)
        columns = [(name, type) for i, name, type, priority in columns]

        # Number of students
        N = len(self.data_df.index)

        # Write `first_ws` and `first_df` column by column from
        # data in `data_df`
        idx = 1
        for name, type in columns:
            # Update first_ws
            if type != "hide":
                # Write header of column with Pandas style
                self.first_ws.cell(1, idx).value = name
                self.first_ws.cell(1, idx).style = "Pandas"

                # Copy data from `data_df` if existing into first
                # worksheet
                if name in self.data_df.columns:
                    for i, value in enumerate(self.data_df[name]):
                        self.first_ws.cell(i + 2, idx).value = value

                # Get cells
                cells = self.first_ws[get_column_letter(idx)][1 : (N + 1)]

                # Fit width of column to actual content
                fit_cells_at_col(self.first_ws.cell(1, idx), *cells)

                # Next column index
                idx += 1

            else:
                # Non-existent column
                cells = None

            # Update first_df
            if name in self.data_df.columns:  # If column exists
                if type in ["hide", "raw"]:
                    self.first_df[name] = self.data_df[name]
                elif type in ["grade", "cell"]:
                    if cells is None:
                        raise Exception("Type is `{type}` but ")
                    self.first_df[name] = cells
                else:
                    raise Exception("Unkown type of column ", type)
            else:
                if type in ["grade", "cell"]:
                    self.first_df[name] = cells
                elif type in ["hide", "raw"]:
                    raise Exception(
                        "No data to copy from and type is 'raw' or 'hide'",
                        name,
                        self.data_df.columns,
                    )
                else:
                    raise Exception("Unkown type of column ", type)

        # Sort all columns
        self.first_ws.auto_filter.ref = "A1:{}{}".format(
            get_column_letter(idx - 1), N + 1
        )

        if N >= 10:
            # Freeze header row and first two columns if more than 10 columns
            self.first_ws.freeze_panes = "C2"
        else:
            # Freeze header row
            self.first_ws.freeze_panes = "A2"

    def write_workbook(self):
        self.workbook.save(self.output_file)

    def create_worksheets(self):
        pass

    def write(self):
        self.add_arguments()
        self.parse_arguments()
        self.setup()
        self.create_first_worksheet()
        self.create_worksheets()
        self.write_workbook()


class GradeSheetMultiple(MultipleCliOpt, BaseGradeSheet):
    """Classe abstraite pour la création de plusieurs feuilles de notes"""

    def create_worksheets(self):
        """Create one or more worksheets based on `worksheets` argument."""

        if self.args.group_by is not None:
            # On groupe avec group_by issu de l'option --worksheets
            gb = self.first_df.sort_values(self.args.group_by).groupby(
                self.args.group_by
            )
            for name, group in gb:
                if self.args.order_by is not None:
                    group = sort_values(group, [self.args.order_by])
                self.create_worksheet(name, group)
        else:
            group = self.first_df
            if self.args.order_by is not None:
                group = sort_values(group, [self.args.order_by])
            self.create_worksheet("grade", group)

    def create_worksheet(self, name, group):
        pass


class MarkingScheme:
    def __init__(self, tree):
        self.tree = tree
        self.width = None
        self.height = None

    @property
    def points(self):
        "Return list of points"

        def get_points(tree):
            if isinstance(tree, dict):
                return [pts for node in tree.values() for pts in get_points(node)]
            else:
                # Should be a list of dict, concatenating
                a = {}
                for s in tree:
                    a.update(s)
                return [a["points"]]

        return get_points(self.tree)

    @property
    def coeffs(self):
        "Return list of coeffs"

        def get_coeffs(tree):
            if isinstance(tree, dict):
                return [pts for node in tree.values() for pts in get_coeffs(node)]
            else:
                # Should be a list of dict, concatenating
                a = {}
                for s in tree:
                    a.update(s)
                if "coeffs" in a:
                    return [a["coeffs"]]
                else:
                    return [1]

        return get_coeffs(self.tree)

    def write(self, worksheet, ref):
        row = ref.row
        col = ref.col_idx
        maxi = maxj = -1

        for name, i, j, di, dj in walk_tree(self.tree):
            maxi = max(maxi, i + di)
            maxj = max(maxj, j + dj)
            worksheet.merge_cells(
                start_row=row + i,
                start_column=col + j,
                end_row=row + i + di - 1,
                end_column=col + j + dj - 1,
            )
            worksheet.cell(row=row + i, column=col + j).value = name
            al = Alignment(horizontal="center", vertical="center")
            worksheet.cell(row=row + i, column=col + j).alignment = al

        tree_height = maxi
        tree_width = maxj

        # Columns of coeffs
        ref_coeffs = ref.right(tree_width)
        ref_coeffs.above().text("Coeffs")

        for i, coeff in enumerate(self.coeffs):
            ref_coeffs.below(i).text(coeff)
        self.sum_coeffs = sum(self.coeffs)
        self.coeff_cells = [ref_coeffs.below(i) for i in range(len(self.coeffs))]

        # Columns of points
        ref_points = ref_coeffs.right()
        ref_points.above().text("Points")

        for i, points in enumerate(self.points):
            ref_points.below(i).text(points)
        self.points_cells = [ref_points.below(i) for i in range(len(self.points))]

        ref_points_last = ref_points.below(len(self.points) - 1)

        self.global_total = ref_points_last.below().text(
            "="
            + "+".join(
                "{0} * {1}".format(cell_coeff.coordinate, cell_point.coordinate)
                for cell_point, cell_coeff in zip(self.points_cells, self.coeff_cells)
            )
        )
        ref_points_last.below().left().text("Grade")

        ref_points_last.below(2).text(20)
        ref_points_last.below(2).left().text("Grade /20")

        self.width = tree_width + 2
        self.height = tree_height
        self.bottom_right = worksheet.cell(row + self.height, col + self.width)


class GradeSheetNoGroup(ConfigCliOpt, FirstGradeSheet):
    name = "simple"
    config_argname = "--marking-scheme"

    def create_worksheet(self, name, group):
        gradesheet = self.workbook.create_sheet(title=name)

        # Make current worksheet the default one, useful for get_address_of_cell
        self.workbook.active = gradesheet

        ref = gradesheet.cell(3, 1)
        ms = MarkingScheme(self.config)
        ms.write(gradesheet, ref)

        ref = row_and_col(ref, ms.bottom_right)

        # Freeze the structure
        gradesheet.freeze_panes = ref.top()

        def insert_record(ref_cell, record):
            last_name = ref_cell.text(record["Nom"])
            first_name = last_name.below().text(record["Prénom"])
            first_grade = first_name.below()
            last_grade = first_grade.below(ms.height - 1)
            total = last_grade.below()
            total_20 = total.below()

            range = get_range_from_cells(first_grade, last_grade)
            formula = f'=IF(COUNTBLANK({range})>0, "", SUM({range}))'
            total.value = formula

            total_20.text(
                '=IF(ISTEXT(%s),"",%s/%s*20)'
                % (
                    get_address_of_cell(total),
                    get_address_of_cell(total),
                    get_address_of_cell(ms.global_total, absolute=True),
                )
            )

            # Cell in first worksheet
            cell = record[self.args.name]
            cell.value = "=" + get_address_of_cell(
                total_20, add_worksheet_name=True, absolute=True
            )

            fit_cells_at_col(last_name, first_name)

        for j, (index, record) in enumerate(group.iterrows()):
            ref_cell = ref.right(j).above(2)
            insert_record(ref_cell, record)


class GradeSheetGroup(GroupCliOpt, ConfigCliOpt, FirstGradeSheet):
    """Base class for workbook with grades by group of students"""

    name = "group"
    config_argname = "--marking-scheme"

    def add_arguments(self):
        super().add_arguments()

        # Make groups inside each worksheet
        self.add_argument("--groups", dest="subgroup_by", default=lambda x: 0)

        # FIXME: Use it or delete it
        self.add_argument(
            "-t",
            "--grade-type",
            default="num",
            required=False,
            choices=["num", "sym"],
            dest="grade_type",
        )

    # FIXME: Check order_by and group_by
    def setup(self):
        super().setup()

    def get_columns(self):
        columns = super().get_columns()

        if not callable(self.args.subgroup_by):
            columns.append((self.args.subgroup_by, "raw", 6))

        if self.args.grade_type == "num":
            columns.append((self.args.name + " brut", "cell", 9))

        return columns

    def create_worksheet(self, name, group):
        gradesheet = self.workbook.create_sheet(title=name)

        # Write marking scheme
        ref = gradesheet.cell(3, 1)
        ms = MarkingScheme(self.config)
        ms.write(gradesheet, ref)

        # Add room for name and total
        height = ms.height + 4

        # Write each block with write_group_block
        ref = ref.right(ms.width).above(2)

        # FIXME: Sort groups according to order_by
        for subname, subgroup in group.groupby(self.args.subgroup_by):
            bottom_right = self.write_group_block(
                gradesheet, ref, subname, subgroup, height, ms
            )
            frame_range(ref, bottom_right)
            frame_range(ref.below(2), bottom_right)
            ref = row_and_col(ref, bottom_right.right())

    def write_group_block(self, gradesheet, ref_cell, name, group, height, ms):
        # Make current worksheet the default one, useful for get_address_of_cell
        self.workbook.active = gradesheet

        # First column is group column
        group_range = list(get_segment(ref_cell, ref_cell.below(height - 1)))

        # Group name
        group_range[0].text(name).merge(group_range[1]).center()

        # Total of column group
        group_total = group_range[-2]

        formula = '=IFERROR({0},"")'.format(
            "+".join(
                "{1} * IF(ISBLANK({0}), 1/0, {0})".format(
                    get_address_of_cell(group_cell, absolute=True),
                    get_address_of_cell(coeff_cell, absolute=True),
                )
                for group_cell, coeff_cell in zip(group_range[2:-2], ms.coeff_cells)
            )
        )
        group_total.value = formula

        group_total_20 = group_range[-1].text(
            '=IF(ISTEXT({0}),"",{0}/{1}*20)'.format(
                get_address_of_cell(group_total),
                get_address_of_cell(ms.global_total, absolute=True),
            )
        )

        # Next columns are per-student columns
        N = len(group)  # Number of students
        gen = generate_ranges(ref_cell.right(), by="col", length=height, nranges=N)

        # Group student dataframe record and corresponding column range
        for stu_range, (index, record) in zip(gen, group.iterrows()):
            stu_range[0].value = record["Nom"]
            stu_range[1].value = record["Prénom"]

            # Total of student
            stu_total = stu_range[-2]
            formula = '=IFERROR({0},"")'.format(
                "+".join(
                    "{2} * IF(ISBLANK({0}), IF(ISBLANK({1}), 1/0, {1}), {0})".format(
                        get_address_of_cell(stu_cell),
                        get_address_of_cell(group_cell),
                        get_address_of_cell(coeff_cell, absolute=True),
                    )
                    for group_cell, stu_cell, coeff_cell in zip(
                        group_range[2:-2], stu_range[2:-2], ms.coeff_cells
                    )
                )
            )
            stu_total.value = formula

            # Total of student over 20
            stu_total_20 = stu_range[-1]
            stu_total_20.text(
                '=IF(ISTEXT(%s),"",%s/%s*20)'
                % (
                    get_address_of_cell(stu_total),
                    get_address_of_cell(stu_total),
                    get_address_of_cell(ms.global_total, absolute=True),
                )
            )

            record[self.args.name].value = "=" + get_address_of_cell(
                stu_total_20, add_worksheet_name=True
            )
            record[self.args.name + " brut"].value = "=" + get_address_of_cell(
                stu_total, add_worksheet_name=True
            )

        return_cell = ref_cell.below(height - 1).right(N)
        return return_cell


class GradeSheetJury(ConfigCliOpt, FirstGradeSheet):
    name = "jury"
    config_argname = "--config"

    def validate_config(self, config):
        """Validation du fichier de configuration"""

        def validate_grade2(data):
            if data.get("type", "grade") == "grade":
                schema = Schema(
                    {
                        Optional("type", default="grade"): "grade",
                        Optional("Note éliminatoire", default=-1): -1,
                        Optional(str): object,
                    }
                )
            else:
                schema = Schema(
                    {"type": Or("raw", "hide", "cell"), Optional(str): object,}
                )
            return schema.validate(data)

        validate_grade = Schema(
            Or(
                And(
                    Or(None, {}, ""),
                    Use(lambda dummy: {"type": "grade", "Note éliminatoire": -1}),
                ),
                And(dict, Use(validate_grade2)),
            )
        )

        DEFAULT_COLUMNS = {"Note agrégée": {"type": "grade", "Note éliminatoire": -1}}

        validate_columns = Or(
            And(Or(None, {}, ""), Use(lambda dummy: DEFAULT_COLUMNS)),
            {
                Optional("Note agrégée", default=DEFAULT_COLUMNS["Note agrégée"]): {},
                Optional(str): validate_grade,
            },
        )

        DEFAULT_OPTIONS = {
            "Percentile note A": 0.9,
            "Percentile note B": 0.65,
            "Percentile note C": 0.35,
            "Percentile note D": 0.1,
        }

        validate_options = Or(
            And(Or(None, {}, ""), Use(lambda dummy: DEFAULT_OPTIONS)),
            {
                "Percentile note A": float,
                "Percentile note B": float,
                "Percentile note C": float,
                "Percentile note D": float,
                Optional(str): object,
            },
            {Optional(str): object},
        )

        sc = Schema(
            Or(
                And(
                    None,
                    Use(
                        lambda dummy: {
                            "columns": validate_columns.validate(None),
                            "options": validate_options.validate(None),
                        }
                    ),
                ),
                {
                    Optional(
                        "columns", default=validate_columns.validate(None)
                    ): validate_columns,
                    Optional(
                        "options", default=validate_options.validate(None)
                    ): validate_options,
                },
            )
        )

        return sc.validate(config)

    def get_columns(self, **kwargs):
        "Les colonnes à utiliser dans la feuille Excel"

        # Les colonnes classiques
        columns = [("Nom", "raw", 0), ("Prénom", "raw", 0), ("Courriel", "raw", 0)]

        # Les colonnes nécessaires pour les différents calculs
        columns.append(("Admis", "cell", 0))
        columns.append(("Note admis", "cell", 0))

        # Les colonnes spécifiées dans le fichier de configuration
        for name, props in self.config["columns"].items():
            col_type = props.get("type") if props is not None else None
            columns.append((name, col_type, 0))

        # La colonne de la note finale : A, B, C, D, E, F
        columns.append(("Note ECTS", "cell", 0))

        return columns

    def write(self):
        self.add_arguments()
        self.parse_arguments()
        self.setup()
        super().create_first_worksheet()
        self.create_second_worksheet()
        self.update_first_worksheet()
        self.write_workbook()

    def get_column_range(self, colname):
        "Renvoie la plage de cellule de la colonne COLNAME sans l'en-tête."

        if colname not in self.first_df.columns:
            raise Exception("Unknown column name: {}".format(colname))

        cells = self.first_df[colname]
        first, last = cells.iloc[0], cells.iloc[-1]

        return get_range_from_cells(first, last)

    @property
    def grade_columns(self):
        """Colonnes associées à des notes"""

        for name, props in self.config["columns"].items():
            if props["type"] == "grade":
                props2 = props.copy()
                del props2["type"]
                yield name, props2

    def write_key_value_props(self, ref_cell, title, props):
        "Helper function to write key-value table"

        keytocell = {}
        ref_cell.text(title).merge(ref_cell.right()).style = "Pandas"
        for key, value in props.items():
            ref_cell = ref_cell.below()
            ref_cell.value = key
            ref_cell.right().value = value
            keytocell[key] = ref_cell.right()
        return ref_cell.right(), keytocell

    def create_second_worksheet(self):
        # Write new gradesheet
        self.gradesheet = self.workbook.create_sheet(title=self.name)
        self.workbook.active = self.gradesheet
        current_cell = self.gradesheet.cell(row=1, column=1)

        # Write option blocks for grade columns
        self.grades_options = {}
        for name, props in self.grade_columns:
            # Write key-value table
            lower_right, keytocell = self.write_key_value_props(
                current_cell, name, props
            )
            self.grades_options[name] = keytocell
            current_cell = lower_right.below(2).left(1)

        # Add default global options block
        options = self.config.get("options", {})
        for ects, grade in zip("ABCD", [0.9, 0.65, 0.35, 0.1]):
            if ("Percentile note " + ects) not in options:
                options["Percentile note " + ects] = grade

        lower_right, self.global_options = self.write_key_value_props(
            current_cell, "Options globales", options
        )
        current_cell = lower_right.below(2).left(1)

        # On écrit les seuils des notes correspondants aux percentiles choisis
        props = {}
        for i, ects in enumerate("ABCD"):
            percentile_cell = self.global_options["Percentile note " + ects]
            props[
                ects + " si >="
            ] = "=IF(ISERROR(PERCENTILE({0}, {1})), NA(), PERCENTILE({0}, {1}))".format(
                self.get_column_range("Note admis"),
                get_address_of_cell(percentile_cell),
            )

        lower_right, percentiles_theo = self.write_key_value_props(
            current_cell, "Percentiles théoriques", props
        )
        current_cell = lower_right.below(2).left(1)

        # Percentiles effectifs
        props = {}
        for i, ects in enumerate("ABCD"):
            props[ects + " si >="] = "=" + get_address_of_cell(
                percentiles_theo[ects + " si >="]
            )

        lower_right, self.percentiles_used = self.write_key_value_props(
            current_cell, "Percentiles utilisés", props
        )
        current_cell = lower_right.below(2).left(1)

        # On écrit les proportions de notes ECTS en fonction de la
        # colonne `Admis`
        ref_cell = self.gradesheet.cell(row=1, column=4)
        props = {}
        for ects in "ABCDEF":
            props["Nombre de " + ects] = ('=COUNTIF({}, "{}")').format(
                self.get_column_range("Note ECTS"), ects
            )
        props["Nombre d'admis"] = '=SUMIF({}, "<>#N/A")'.format(
            self.get_column_range("Admis")
        )
        props["Effectif total"] = "=COUNTA({})".format(self.get_column_range("Admis"))
        props["Ratio"] = '=AVERAGEIF({}, "<>#N/A")'.format(
            self.get_column_range("Admis")
        )
        lower_right, statistiques = self.write_key_value_props(
            ref_cell, "Statistiques", props
        )

    def update_first_worksheet(self):
        # Pour que get_address_of_cell marche correctement
        self.workbook.active = self.first_ws

        # On écrit la colonne "Admis" des admis/refusés basée sur les
        # barres
        for i, (index, record) in enumerate(self.first_df.iterrows()):
            # Teste si une des notes est vide
            any_blank_cell = "OR({})".format(
                ", ".join(
                    "ISBLANK({})".format(get_address_of_cell(record[name]))
                    for name, props in self.grade_columns
                )
            )

            # Teste si toutes les barres sont atteintes
            above_all_threshold = "AND({})".format(
                ", ".join(
                    "IF(ISNUMBER({0}), {0}>={1}, 1)".format(
                        get_address_of_cell(record[name]),
                        get_address_of_cell(
                            self.grades_options[name]["Note éliminatoire"],
                            absolute=True,
                        ),
                    )
                    for name, props in self.grade_columns
                )
            )

            # NA() s'il y a un problème, 0 si recalé, 1 si reçu
            formula = f"=IFERROR(IF({any_blank_cell}, NA(), IF({above_all_threshold}, 1, 0)), NA())"
            record["Admis"].value = formula

        # On écrit la note agrégée des admis dans la colonne "Note
        # admis" pour faciliter le calcul des percentiles sur les
        # admis
        for i, (index, record) in enumerate(self.first_df.iterrows()):
            record["Note admis"].value = '=IFERROR(IF({}=1, {}, ""), "")'.format(
                get_address_of_cell(record["Admis"]),
                get_address_of_cell(record["Note agrégée"]),
            )

        # On écrit la note ECTS en fonction de la note agrégée et des
        # percentiles utilisés
        perc_A = get_address_of_cell(self.percentiles_used["A si >="], absolute=True)
        perc_B = get_address_of_cell(self.percentiles_used["B si >="], absolute=True)
        perc_C = get_address_of_cell(self.percentiles_used["C si >="], absolute=True)
        perc_D = get_address_of_cell(self.percentiles_used["D si >="], absolute=True)
        percs = dict(perc_A=perc_A, perc_B=perc_B, perc_C=perc_C, perc_D=perc_D)

        for i, (index, record) in enumerate(self.first_df.iterrows()):
            note_admis = get_address_of_cell(record["Admis"])
            note_agregee = get_address_of_cell(record["Note agrégée"])
            opts = dict(note_admis=note_admis, note_agregee=note_agregee)
            opts.update(percs)

            # Fonction pour la construction des IF imbriqués
            def switch(ifs, default):
                if ifs:
                    cond, then = ifs[0]
                    return "IF({cond}, {then}, {else_})".format(
                        cond=cond.format(**opts),
                        then=then,
                        else_=switch(ifs[1:], default),
                    )
                else:
                    return default

            # Formule de type switch/case
            formula = switch(
                (
                    ("ISNA({note_admis})", "NA()"),
                    ('{note_agregee}="RESERVE"', '"RESERVE"'),
                    ('{note_agregee}="ABS"', '"ABS"'),
                    ("{note_admis}=0", '"F"'),
                    ("{note_agregee}>={perc_A}", '"A"'),
                    ("{note_agregee}>={perc_B}", '"B"'),
                    ("{note_agregee}>={perc_C}", '"C"'),
                    ("{note_agregee}>={perc_D}", '"D"'),
                ),
                '"E"',
            )

            record["Note ECTS"].value = "=" + formula

        # On centre les notes ECTS
        for cell in self.first_df["Note ECTS"]:
            cell.alignment = Alignment(horizontal="center")

        # On colore la cellule en fonction de la note ECTS
        range = self.get_column_range("Note ECTS")
        for ects, color in zip(
            "ABCDEF", ["00FF00", "C2FF00", "F7FF00", "FFC100", "FF6900", "FF0000"]
        ):
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            rule = CellIsRule(operator="equal", formula=[f'"{ects}"'], fill=fill)
            self.first_ws.conditional_formatting.add(range, rule)

        # Formattage conditionnel pour les notes éliminatoires
        red_fill = PatternFill(
            start_color="EE1111", end_color="EE1111", fill_type="solid"
        )

        for name, opts in self.grades_options.items():
            threshold_cell = opts["Note éliminatoire"]
            threshold_addr = get_address_of_cell(threshold_cell, compat=True)
            self.first_ws.conditional_formatting.add(
                self.get_column_range(name),
                CellIsRule(
                    operator="lessThan", formula=[threshold_addr], fill=red_fill
                ),
            )

        # On offre la possibilité de trier
        max_column = self.first_ws.max_column
        max_row = self.first_ws.max_row
        self.first_ws.auto_filter.ref = "A1:{}{}".format(
            get_column_letter(max_column), max_row
        )
        range = self.get_column_range("Note ECTS")
        self.first_ws.auto_filter.add_sort_condition(range)


class GradeSheetSimpleWriter(GradeSheetMultiple):
    """Feuille de notes simple par étudiant et sans barème."""

    name = "simple"

    def get_columns(self, **kwargs):
        "Renvoie les colonnes à utiliser dans la première feuille de calculs"

        # Default columns
        columns = {"Nom": "raw", "Prénom": "raw", "Courriel": "raw"}

        # Add order_by columns if specified in command line
        if self.args.order_by is not None:
            columns[self.args.order_by] = "raw"

        if self.args.group_by is not None:
            columns[self.args.group_by] = "raw"

        # Add column for grades
        columns[self.args.name] = "cell"

        return columns

    def create_worksheet(self, name, group):
        gradesheet = self.workbook.create_sheet(title=name)
        ref = (1, 1)
        row, col = ref

        lastname = gradesheet.cell(*ref).text("Nom")
        name = lastname.below().text("Prénom")
        grade = name.below().text("Note")

        ref_list = lastname.right()
        max_len = 0
        for i, (index, record) in enumerate(group.iterrows()):
            max_len = max(max_len, len(record["Nom"]), len(record["Prénom"]))

            lastname = ref_list.right(i)
            name = lastname.below()
            grade = name.below()

            lastname.value = record["Nom"]
            name.value = record["Prénom"]

            # Get cell to be filled in first worksheet and make it
            # point to current cell in second worksheet.
            cell = record[self.args.name]

            cell.value = if_blank_formula(
                get_address_of_cell(grade, add_worksheet_name=True)
            )

        for i in range(len(group)):
            gradesheet.column_dimensions[
                get_column_letter(col + i + 1)
            ].width = max_len

        # On fige la première ligne
        gradesheet.freeze_panes = ref_list.top()


class GradeSheetExamWriter(ConfigCliOpt, GradeSheetMultiple):
    """Feuille de notes pour un examen type médian/final avec des
questions structurées."""

    # Name used to identify the class to use in the sub_command
    # parser.
    name = "exam"

    def get_columns(self, **kwargs):
        return {
            "Nom": "raw",
            "Prénom": "raw",
            "Courriel": "raw",
            self.args.group: "raw",
            self.args.name: "cell",
        }

    def create_worksheet(self, name, group):
        gradesheet = self.workbook.create_sheet(title=name)


WRITERS = [GradeSheetNoGroup, GradeSheetGroup, GradeSheetJury]


def run(argv=sys.argv[1:], prog=os.path.basename(__file__), description=None):
    parser = argparse.ArgumentParser(prog=prog, description=description)
    subparsers = parser.add_subparsers(dest="sub_command", required=True)
    name_to_writer = {}
    for type in WRITERS:
        name_to_writer[type.name] = type
        obj = type()
        obj.add_arguments()
        base_parser = obj._parser
        subparsers.add_parser(type.name, help=type.__doc__, parents=[base_parser])
    args = parser.parse_args(argv)

    writer_class = name_to_writer[args.sub_command]
    writer = writer_class(argv=argv[1:])
    writer.write()
