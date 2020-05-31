#!/usr/bin/env python3

__doc__ = """
Create an Excel file to compute grades
"""

import os
import sys
import math
import argparse

from collections import OrderedDict
from openpyxl import Workbook
from openpyxl import utils
from openpyxl.cell import Cell, MergedCell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, PatternFill, Side, Border
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils.cell import absolute_coordinate

import pandas as pd
import oyaml as yaml            # Ordered yaml

# Custom navigation functions
def left(self, step=1):
    return self.offset(0, -step)

def right(self, step=1):
    return self.offset(0, step)

def above(self, step=1):
    return self.offset(-step, 0)

def below(self, step=1):
    return self.offset(step, 0)

def top(self):
    return self.parent.cell(column=self.column, row=1)

def text(self, value):
    self.value = value
    return self

def center(self):
    self.alignment = Alignment(horizontal='center', vertical='center')
    return self

def set_border(self):
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    self.border += border
    return self

# Add offset function to MergedCell
def offset(self, row=0, column=0):
    offset_column = self.column + column
    offset_row = self.row + row
    return self.parent.cell(column=offset_column, row=offset_row)

def merge(self, cell):
    assert self.parent == cell.parent

    self.parent.merge_cells(
        start_row=self.row,
        start_column=self.column,
        end_row=cell.row,
        end_column=cell.column
    )
    return self


MergedCell.offset = offset

Cell.left = left
Cell.right = right
Cell.above = above
Cell.below = below
Cell.top = top
Cell.text = text
Cell.center = center
Cell.set_border = set_border
Cell.merge = merge
MergedCell.left = left
MergedCell.right = right
MergedCell.above = above
MergedCell.below = below
MergedCell.top = top
MergedCell.text = text
MergedCell.center = center
MergedCell.set_border = set_border


def merge_cells2(self, cell1, cell2):
    """Merge rectangle defined by upper left and lower right cells"""

    self.merge_cells(
        start_row=cell1.row,
        start_column=cell1.col_idx,
        end_row=cell2.row,
        end_column=cell2.col_idx
    )

    return self.cell(row=cell1.row, column=cell1.col_idx)

Worksheet.merge_cells2 = merge_cells2


def frame_range(cell1, cell2):
    cell_range = cell1.coordinate + ":" + cell2.coordinate
    thin = Side(border_style="thin", color="000000")
    top = Border(top=thin)
    left = Border(left=thin)
    right = Border(right=thin)
    bottom = Border(bottom=thin)

    rows = cell1.parent[cell_range]
    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right


def get_address_of_cell(cell, absolute=False, add_worksheet_name=None, compat=False):
    """Renvoie l'adresse d'un objet Cell sous forme "A1" en prenant en
    compte la feuille courante si la cellule se trouve sur une
    autre feuille.
    """

    cell_worksheet = cell.parent
    workbook = cell.parent.parent
    active_worksheet = workbook.active

    if absolute:
        coordinate = absolute_coordinate(cell.coordinate)
    else:
        coordinate = cell.coordinate

    if add_worksheet_name == False or (add_worksheet_name is None and cell_worksheet == active_worksheet):
        return coordinate
    else:
        if compat:          # GoogleSheet compatibility
            return "INDIRECT(\"'{}'!{}\")".format(cell_worksheet.title, coordinate)
        else:
            return "'{}'!{}".format(cell_worksheet.title, coordinate)


def get_range_from_cells(cell1, cell2, absolute=False, add_worksheet_name=None, compat=False):
    cell_worksheet = cell1.parent
    workbook = cell1.parent.parent
    active_worksheet = workbook.active

    # Upper-left and lower right cells
    cell1_row, cell2_row = (cell1.row, cell2.row) if cell1.row < cell2.row else (cell2.row, cell1.row)
    cell1_col, cell2_col = (cell1.column, cell2.column) if cell1.column < cell2.column else (cell2.column, cell1.column)

    # Excel-like coordinates
    cell1_addr = utils.get_column_letter(cell1_col) + str(cell1_row)
    cell2_addr = utils.get_column_letter(cell2_col) + str(cell2_row)

    if absolute:
        range = absolute_coordinate(cell1_addr) + ":" + absolute_coordinate(cell2_addr)
    else:
        range = cell1_addr + ":" + cell2_addr

    if add_worksheet_name == False or (add_worksheet_name is None and cell_worksheet == active_worksheet):
        return range
    else:
        if compat:          # GoogleSheet compatibility
            return "INDIRECT(\"'{}'!{}\")".format(cell_worksheet.title, range)
        else:
            return "'{}'!{}".format(cell_worksheet.title, range)


def get_segment(cell1, cell2):
    """Return a generator giving cells from cell1 to cell2"""
    if cell1.row == cell2.row:
        if cell1.column > cell2.column:
            cell1, cell2 = cell2, cell1

        for i in range(cell1.column, cell2.column+1):
            yield cell1.parent.cell(row=cell1.row, column=i)
    elif cell1.column == cell2.column:
        if cell1.row > cell2.row:
            cell1, cell2 = cell2, cell1

        for i in range(cell1.row, cell2.row+1):
            yield cell1.parent.cell(column=cell1.column, row=i)
    else:
        raise Exception('Must have same row or column')


def row_and_col(cell1, cell2):
    assert cell1.parent == cell2.parent
    parent = cell1.parent
    return parent.cell(row=cell1.row, column=cell2.column)


def col_and_row(cell1, cell2):
    return row_and_col(cell2, cell1)


def if_empty_formula(formula, blank_value=""):
    if formula.startswith("="):
        formula = formula[1:]

    return "=IF(ISBLANK(%s),\"%s\",%s)" % (
        formula,
        blank_value,
        formula
    )


from operator import attrgetter
from itertools import groupby

def fit_cells_at_col(*cells):
    worksheet = cells[0].parent
    assert all(c.parent == worksheet for c in cells)
    cells = list(cells)
    cells.sort(key=attrgetter("column"))
    for k, g in groupby(cells, key=attrgetter("column")):
        lengths = [len(l) for c in g if c.value is not None for l in str(c.value).splitlines()]
        if lengths:
            max_len = max(lengths)
            worksheet.column_dimensions[utils.get_column_letter(k)].width = 1.3*max_len


def walk_tree(tree, depth=None, start_at=0):
    def compute_depth(tree):
        if isinstance(tree, (OrderedDict, dict)):
            return 1 + max(compute_depth(child) for child in tree.values())
        else:
            return 1

    if depth is None:
        depth = compute_depth(tree)

    class RunOut(Exception):
        pass

    def walk_tree0(name, tree, current_depth, nleaves):
        if isinstance(tree, (OrderedDict, dict)):  # Inner node
            current_leaves = 0
            sdepth = None
            x_span = None
            for key, child in tree.items():
                try:
                    yield from walk_tree0(key, child, current_depth + 1, nleaves + current_leaves)
                except RunOut as e:
                    nleaf, sdepth = e.args
                    current_leaves += nleaf
                    x_span = math.ceil(sdepth // (current_depth + 1))
                    sdepth = sdepth - x_span
            if current_depth != 0:
                yield name, nleaves, sdepth, current_leaves, x_span
                raise RunOut(current_leaves, sdepth)
        else:                   # Leaf
            x_span = math.ceil((depth + 1) // (current_depth + 1))
            sdepth = depth - x_span
            yield name, nleaves, sdepth, 1, x_span
            raise RunOut(1, sdepth)

    yield from walk_tree0('root', tree, 0, 0)


class GradeSheetWriter:
    """Base class to build and write a workbook. It consists in only one
    worksheet."""

    def __init__(self, args):  # args argument comes from parser returned by get_parser

        # args always has a name, store it
        if args.name:
            self.name = args.name

        # Setting source of grades
        if args.data_file:
            if os.path.isdir(args.data_file):
                fn = 'student_data_merge.xlsx'
                self.data_file = os.path.join(args.data_file, fn)
            else:
                self.data_file = args.data_file
        else:
            self.data_file = 'student_data_merge.xlsx'

        # Reading source of information into a Pandas dataframe
        if not os.path.exists(self.data_file):
            raise Exception(f'Data file `{self.data_file}` does not exist')
        self.data_df = pd.read_excel(self.data_file)

        # Setting path of gradebook file to be written
        if args.output_file:
            if os.path.isdir(args.output_file):
                fn = f'{args.name}_gradebook.xlsx'
                self.output_file = os.path.join(args.output_file, fn)
            else:
                self.output_file = args.output_file
        else:
            self.output_file = f'{args.name}_gradebook.xlsx'

        # Create workbook and first worksheet named "data"
        self.wb = Workbook()
        self.ws_data = self.wb.active
        self.ws_data.title = "data"

        # Pandas dataframe that mirrors the first worksheet
        self.df = pd.DataFrame()

        # Get columns to be copied from source of information DATA_DF
        # to first worksheet
        columns = self.get_columns(**args.__dict__)
        N = len(self.data_df.index)
        for i, (name, type) in enumerate(columns.items()):
            idx = i + 1

            # Write header of column with Pandas style
            self.ws_data.cell(1, idx).value = name
            self.ws_data.cell(1, idx).style = "Pandas"

            if name in self.data_df.columns:
                # Copy data from DATA_DF if existing into first worksheet
                for i, value in enumerate(self.data_df[name]):
                    self.ws_data.cell(i + 2, idx).value = value
                cells = self.ws_data[utils.get_column_letter(idx)][1:(N+1)]

                # Fit width of column to actual content
                fit_cells_at_col(*cells)

                # Add values or cells to DF
                if type == "raw":
                    self.df[name] = self.data_df[name]
                elif type == "cell" or type is None:
                    self.df[name] = cells
                else:
                    raise Exception("Unkown type of column ", type)

            else:
                # Non-existent column
                if type == 'cell' or type is None:
                    cells = self.ws_data[utils.get_column_letter(idx)][1:(N+1)]
                    self.df[name] = cells
                elif type == 'raw':
                    raise Exception('No data to copy from and type is data', name, self.data_df.columns)
                else:
                    raise Exception("Unkown type of column ", type)

        # Sort all columns
        self.ws_data.auto_filter.ref = 'A1:{}{}'.format(
            utils.get_column_letter(idx),
            N + 1)

        if N >= 10:
            # Freeze header row and first two columns if more than 10 columns
            self.ws_data.freeze_panes = "C2"
        else:
            # Freeze header row
            self.ws_data.freeze_panes = "A2"

    # Create parser whose arguments will be available to the constructor
    @classmethod
    def get_parser(cls):
        parser = argparse.ArgumentParser(description=cls.__doc__, add_help=False)
        parser.add_argument('--name', dest='name', required=True)
        parser.add_argument('-d', '--data', dest='data_file')
        parser.add_argument('-o', '--output-file', dest='output_file')
        return parser

    # Return columns to create in first worksheet. `value' columns are
    # copied directly from source of information, `cell` columns are
    # created to be referenced later in the workbook.
    def get_columns(self, **kwargs):
        return {
            'Nom': 'raw',
            'Prénom': 'raw',
            'Courriel': 'raw',
            self.name: 'cell'
        }

    def write(self):
        raise NotImplementedError


class GradeSheetWriterConfig(GradeSheetWriter):
    def __init__(self, args):
        self.config = self.parse_config(args.config)
        super().__init__(args)

    def parse_config(self, config):
        if not os.path.exists(config):
            raise Exception("Configuration file not found")

        with open(config, "r") as stream:
            return list(yaml.load_all(stream, Loader=yaml.SafeLoader))[0]


class GradeSheetSimpleWriter(GradeSheetWriter):
    """Feuille de notes simple par étudiant et sans barème."""

    # Name used to identify the class to use in the sub_command
    # parser.
    name = 'simple'

    def write(self, ref=None):
        # Create a second worksheet
        self.gradesheet = self.wb.create_sheet(title=self.name)

        if ref is None:
            ref = (3, 1)
        row, col = ref

        lastname = self.gradesheet.cell(*ref).text("Nom")
        name = lastname.below().text("Prénom")
        grade = name.below().text("Note")

        ref_list = lastname.right()
        max_len = 0
        for i, (index, record) in enumerate(self.df.iterrows()):
            max_len = max(max_len, len(record['Nom']), len(record['Prénom']))

            lastname = ref_list.right(i)
            name = lastname.below()
            grade = name.below()

            lastname.value = record['Nom']
            name.value = record['Prénom']

            # Get cell to be filled in first worksheet and make it
            # point to current cell in second worksheet.
            cell = record[self.gradesheet.title]

            cell.value = if_empty_formula(
                get_address_of_cell(grade, add_worksheet_name=True)
            )

        for i in range(len(self.df)):
            self.gradesheet.column_dimensions[utils.get_column_letter(col+i+1)].width = max_len

        # On fige la première ligne
        self.gradesheet.freeze_panes = ref_list.top()

        # Write workbook
        self.wb.save(self.output_file)


class GradeSheetExamWriter(GradeSheetWriterConfig):
    """Feuille de notes pour un examen type médian/final avec des
questions structurées."""

    # Name used to identify the class to use in the sub_command
    # parser.
    name = 'exam'

    # Add a structure argument
    @classmethod
    def get_parser(cls):
        parser = super(GradeSheetExamWriter, GradeSheetExamWriter).get_parser()
        parser.add_argument('-s', '--struct', required=True, dest='config')
        return parser

    @property
    def tree(self):
        return self.config

    def write_structure(self, upper_left):
        "Write structure at UPPER_LEFT and return lower right cell."
        maxi = maxj = -1

        for name, i, j, di, dj in walk_tree(self.tree):
            maxi = max(maxi, i+di)
            maxj = max(maxj, j+dj)
            self.gradesheet.merge_cells(
                start_row=upper_left[0]+i,
                start_column=upper_left[1]+j,
                end_row=upper_left[0]+i+di-1,
                end_column=upper_left[1]+j+dj-1)
            self.gradesheet.cell(
                row=upper_left[0]+i,
                column=upper_left[1]+j).value = name
            al = Alignment(horizontal="center", vertical="center")
            self.gradesheet.cell(
                row=upper_left[0]+i,
                column=upper_left[1]+j).alignment = al

        # Return lower right cell of structure
        return maxi + upper_left[0] - 1, maxj + upper_left[1] - 1

    def write(self, ref=None):
        # Create new gradesheet
        self.gradesheet = self.wb.create_sheet(title=self.name)
        self.wb.active = self.gradesheet

        # Write structure at ref
        if ref is None:
            ref = (3, 1)
        row, col = self.write_structure(upper_left=ref)

        # Write header for grade row
        self.gradesheet.cell(row=row, column=col).below().value = "Grade"
        self.gradesheet.cell(row=row, column=col).below(2).value = "Grade /20"

        # Write total of points
        ref_points = self.gradesheet.cell(row=ref[0], column=col+1)

        def get_points(struct):
            if isinstance(struct, (OrderedDict, dict)):
                return [
                    j for i in struct.values() for j in get_points(i)
                ]
            else:
                a = {}
                for s in struct:
                    a.update(s)
                return [a["points"]]

        ref_points.above().text("Points")
        points_list = get_points(self.tree)
        for i, points in enumerate(points_list):
            ref_points.below(i).text(points)
        ref_points_last = ref_points.below(len(points_list)-1)

        global_total = ref_points_last.below().text(
            "=SUM(%s)" % get_range_from_cells(ref_points, ref_points_last)
        )

        ref_points_last.below(2).text(20)

        # Width of structure
        n_questions = row - ref[0] + 1

        ref_names = ref_points.right()

        # Freeze the structure
        self.gradesheet.freeze_panes = ref_names.top()

        def insert_record(ref_cell, record):
            last_name = ref_cell.text(record['Nom'])
            first_name = last_name.below().text(record['Prénom'])
            first_grade = first_name.below()
            last_grade = first_grade.below(n_questions - 1)
            total = last_grade.below()
            total_20 = total.below()

            range = get_range_from_cells(first_grade, last_grade)
            formula = f'=IF(COUNTBLANK({range})>0, "", SUM({range}))'
            total.value = formula

            total_20.text(
                "=IF(ISTEXT(%s),\"\",%s/%s*20)" % (
                    get_address_of_cell(total),
                    get_address_of_cell(total),
                    get_address_of_cell(global_total, absolute=True)
                )
            )

            # Cell in first worksheet
            cell = record[self.gradesheet.title]
            cell.value = "=" + get_address_of_cell(
                total_20,
                add_worksheet_name=True,
                absolute=True
            )

            fit_cells_at_col(last_name, first_name)

        for j, (index, record) in enumerate(self.df.iterrows()):
            ref_cell = ref_names.right(j).above(2)
            insert_record(ref_cell, record)

        self.wb.save(self.output_file)


class GradeSheetExamMultipleWriter(GradeSheetExamWriter):
    """Feuille de notes avec barème et liste des correcteurs."""

    name = 'exammult'

    def __init__(self, args):
        super().__init__(args)
        insts_file = args.insts
        df = pd.read_excel(insts_file)
        insts = df['Intervenants'].unique()

        def select_inst(i):
            while True:
                try:
                    choice = input(f'Sélectionner le correcteur {i} (y/n)? ')
                    if choice.upper() == "Y":
                        print(f"Correcteur {i} sélectionné")
                        return True
                    elif choice.upper() == "N":
                        print(f"Correcteur {i} non sélectionné")
                        return False
                    else:
                        raise ValueError
                except ValueError:
                    continue
                else:
                    break

        self.insts = [i for i in insts if select_inst(i)]

    @classmethod
    def get_parser(cls):
        parser = super(GradeSheetExamMultipleWriter, GradeSheetExamMultipleWriter).get_parser()
        parser.add_argument('-i', '--instructors', required=True, dest='insts')
        return parser

    def get_columns(self, **kwargs):
        return {
            'Nom': 'raw',
            'Prénom': 'raw',
            'Courriel': 'raw',
            'Note': 'cell',
            "Correcteur": 'cell',
        }

    def write(self, ref=None):
        if ref is None:
            ref = (3, 1)

        for inst in self.insts:
            self.gradesheet = self.wb.create_sheet(title=inst)

            row, col = self.write_structure(upper_left=ref)
            self.gradesheet.cell(row+1, col, "Note")

            # On fige les colonnes après le barème
            self.gradesheet.freeze_panes = self.gradesheet.cell(1, col+1).coordinate

            for i, (index, record) in enumerate(self.df.iterrows()):
                self.gradesheet.cell(ref[0]-2, col+i+1, record['Nom'])
                self.gradesheet.cell(ref[0]-1, col+i+1, record['Prénom'])

                # On écrit le total des points pour former la note globale
                range = (
                    utils.get_column_letter(col+i+1) + str(ref[0]) +
                    ':' +
                    utils.get_column_letter(col+i+1) + str(row)
                )
                formula = f'=IF(COUNTBLANK({range})>0, "", SUM({range}))'

                self.gradesheet.cell(row+1, col+i+1, formula)

        # Formula to return first non empty cell
        grade_format = '""'
        for inst in self.insts:
            cell = "'{}'!{{0}}{{1}}".format(inst)
            grade_format = f"IF({cell} = \"\", {grade_format}, {cell})"

        # Formula to return first instructor
        inst_format = '""'
        for inst in self.insts:
            cell = "'{}'!{{0}}{{1}}".format(inst)
            inst_format = f"IF({cell} = \"\", {inst_format}, \"{inst}\")"
        # inst_format = "CONCATENATE(\"Corrigé par \", {})".format(inst_format)

        # List of booleans of non empty cells from each sheet
        non_empty_grades = ["IF('{}'!{{0}}{{1}} <> \"\", 1, 0)".format(inst)
                            for inst in self.insts]
        non_empty_grades = ", ".join(non_empty_grades)

        for i, (index, record) in enumerate(self.df.iterrows()):
            formula = "=IF(SUM({0}) = 0, \"\", IF(SUM({0}) = 1, {1}, NA())".format(
                non_empty_grades.format(
                    utils.get_column_letter(col+i+1).upper(),
                    row+1
                ),
                grade_format.format(
                    utils.get_column_letter(col+i+1).upper(),
                    row+1
                )
            )
            record["Note"].value = formula

            formula = "=IF(SUM({0}) = 0, \"\", IF(SUM({0}) = 1, {1}, NA())".format(
                non_empty_grades.format(
                    utils.get_column_letter(col+i+1).upper(),
                    row+1
                ),
                inst_format.format(
                    utils.get_column_letter(col+i+1).upper(),
                    row+1
                )
            )
            record["Correcteur"].value = formula

        self.wb.save(self.output_file)


class GradeSheetSimpleGroup(GradeSheetSimpleWriter):
    """Simple gradesheet per groups, no subgrading"""

    name = 'group'

    def __init__(self, args):
        super(GradeSheetSimpleGroup, self).__init__(args)
        self.group = args.group
        self.grade_type = args.grade_type

    def get_columns(self, **kwargs):
        return {
            'Nom': 'raw',
            'Prénom': 'raw',
            'Courriel': 'raw',
            kwargs['group']: 'raw',
            self.name: 'cell'
        }

    @classmethod
    def get_parser(cls):
        parser = super(GradeSheetSimpleGroup, GradeSheetSimpleGroup).get_parser()
        parser.add_argument('-g', '--group', required=True, dest='group')
        parser.add_argument('-t', '--grade-type', default='num', required=False, choices=['num', 'sym'], dest='grade_type')
        return parser

    def write(self, ref=None):
        # Write new gradesheet and make it default
        self.gradesheet = self.wb.create_sheet(title=self.name)
        self.wb.active = self.gradesheet

        header_ref_cell = self.gradesheet['B1']
        header_ref_cell.below().left().text("Note").below().text("Correction")

        for i, (key, group) in enumerate(self.df.groupby(self.group)):
            # Group name
            self.gradesheet.merge_cells2(
                header_ref_cell.right(2*i),
                header_ref_cell.right(2*i+1))
            header_ref_cell.right(2*i).value = key

            # Note
            self.gradesheet.merge_cells2(
                header_ref_cell.below().right(2*i),
                header_ref_cell.below().right(2*i+1))

            name_cells = []
            for j, (index, record) in enumerate(group.iterrows()):
                name = record['Nom'] + ' ' + record['Prénom']
                name_cell = header_ref_cell.right(2*i).below(2+j)
                name_cell.value = name
                name_cells.append(name_cell)

                grade_addr = get_address_of_cell(
                    header_ref_cell.below().right(2*i),
                    absolute=True,
                    add_worksheet_name=True
                )
                grademod_addr = get_address_of_cell(
                    header_ref_cell.below(2+j).right(2*i+1),
                    absolute=True,
                    add_worksheet_name=True
                )
                if self.grade_type == "num":
                    formula = "=%s+%s" % (grade_addr, grademod_addr)
                    record[self.name].value = formula
                elif self.grade_type == "sym":
                    formula = "=IF(ISBLANK(%s),IF(ISBLANK(%s),\"\", %s),%s)" % (
                        grademod_addr,
                        grade_addr,
                        grade_addr,
                        grademod_addr
                    )
                    record[self.name].value = formula
                else:
                    raise Exception("Unknown grade type")

            fit_cells_at_col(*name_cells)

        self.wb.save(self.output_file)


class GradeSheetAssignmentWriter(GradeSheetExamWriter):
    """Feuille pour attribution d'une note par groupe.
    """

    name = 'assignment'

    def __init__(self, args):
        super().__init__(args)
        self.group = args.group

    def get_columns(self, **kwargs):
        return {
            'Nom': 'raw',
            'Prénom': 'raw',
            'Courriel': 'raw',
            kwargs['group']: 'raw',
            self.name: 'cell'
        }

    @classmethod
    def get_parser(cls):
        parser = super(GradeSheetAssignmentWriter, GradeSheetAssignmentWriter).get_parser()
        parser.add_argument('-g', '--group', required=True, dest='group')
        return parser

    def write(self, ref=None):
        # Write new gradesheet
        self.gradesheet = self.wb.create_sheet(title=self.name)

        if ref is None:
            ref = (3, 1)
        row, col = self.write_structure(upper_left=ref)

        for i, (key, group) in enumerate(self.df.groupby(self.group)):
            group_names = ', '.join(group['Nom'] + ' ' + group['Prénom'])
            self.gradesheet.cell(ref[0]-2, col+i+1, key)
            self.gradesheet.cell(ref[0]-1, col+i+1, group_names)

            first = self.gradesheet.cell(ref[0], col+i+1)
            last = self.gradesheet.cell(row, col+i+1)
            grade = self.gradesheet.cell(row+1, col+i+1)

            formula = "=SUM(%s)" % get_range_from_cells(
                first,
                last,
                add_worksheet_name=False
            )
            grade.value = formula

            # Write link to grade for each member of group
            for index, record in group.iterrows():
                record[self.name].value = if_empty_formula(
                    get_address_of_cell(
                        grade,
                        add_worksheet_name=True
                ))

        self.wb.save(self.output_file)


class GradeSheetJuryWriter(GradeSheetWriterConfig):
    """Feuille Excel pour jury avec les notes, une colonne des notes
    agrégées, des percentiles pour les notes ECTS.
    """
    name = "jury"

    def get_columns(self, **kwargs):
        columns = {
            'Nom': 'raw',
            'Prénom': 'raw',
            'Courriel': 'raw',
            'Admis': 'cell',
            'Note admis': 'cell'
        }

        for name, props in self.config['columns'].items():
            if props is not None:
                columns[name] = props.get('type')
            else:
                columns[name] = None

        columns['Note agrégée'] = 'cell'
        columns['Note ECTS'] = 'cell'
        return columns

    @classmethod
    def get_parser(cls):
        parser = super(GradeSheetJuryWriter, GradeSheetJuryWriter).get_parser()
        parser.add_argument('-c', '--config', required=True, dest='config')
        return parser

    def get_column_range(self, colname):
        "Renvoie la plage de cellule de la colonne COLNAME sans l'en-tête."

        if colname not in self.df.columns:
            raise Exception('Unknown column name: {}'.format(colname))

        cells = self.df[colname]
        first, last = cells.iloc[0], cells.iloc[-1]

        return get_range_from_cells(first, last)

    def write(self, ref=None):
        # Write new gradesheet
        self.gradesheet = self.wb.create_sheet(title=self.name)

        current_cell = self.gradesheet.cell(row=1, column=1)

        # Helper function to write key-value table
        def write_key_value_props(ref_cell, title, props):
            keytocell = {}
            ref_cell.text(title).merge(ref_cell.right()).style = "Pandas"
            for key, value in props.items():
                ref_cell = ref_cell.below()
                ref_cell.value = key
                ref_cell.right().value = value
                keytocell[key] = ref_cell.right()
            return ref_cell.right(), keytocell

        # Write options for all "columns" items in configuration file
        grades_options = {}
        if "Note agrégée" not in self.config['columns']:
            self.config['columns']['Note agrégée'] = {
                'note_éliminatoire': -1
            }
        for name, props in self.config['columns'].items():
            # Add default options
            if props is None:
                opts = {'note_éliminatoire': -1}
            elif 'options' not in props:
                opts = {'note_éliminatoire': -1}
            elif 'note_éliminatoire' not in props['options']:
                opts = props['options']
                opts['note_éliminatoire'] = -1

            # Write key-value table
            lower_right, keytocell = write_key_value_props(
                current_cell,
                name,
                opts
            )
            grades_options[name] = keytocell
            current_cell = lower_right.below(2).left(1)

        # Add default global options
        options = self.config.get('options', {})
        for ects, grade in zip('ABCD', [0.9, 0.65, 0.35, 0.1]):
            if ('Percentile note ' + ects) not in options:
                options['Percentile note ' + ects] = grade

        lower_right, global_options = write_key_value_props(
            current_cell,
            "Options globales",
            options
        )
        current_cell = lower_right.below(2).left(1)

        # On écrit les seuils des notes correspondants aux percentiles choisis
        props = {}
        self.wb.active = self.gradesheet
        for i, ects in enumerate('ABCD'):
            percentile_cell = global_options['Percentile note ' + ects]
            props[ects + " si >="] = (
                '=IF(ISERROR(PERCENTILE({}, {})), NA(), PERCENTILE({}, {}))'.format(
                    self.get_column_range('Note admis'),
                    get_address_of_cell(percentile_cell),
                    self.get_column_range('Note admis'),
                    get_address_of_cell(percentile_cell)
                )
            )

        lower_right, percentiles_theo = write_key_value_props(
            current_cell,
            "Percentiles théoriques",
            props
        )
        current_cell = lower_right.below(2).left(1)

        # Percentiles effectifs
        props = {}
        self.wb.active = self.gradesheet
        for i, ects in enumerate('ABCD'):
            props[ects + " si >="] = (
                "=" + get_address_of_cell(percentiles_theo[ects + " si >="])
            )

        lower_right, percentiles_used = write_key_value_props(
            current_cell,
            "Percentiles utilisés",
            props
        )
        current_cell = lower_right.below(2).left(1)

        # On écrit les proportions de notes ECTS en fonction de la
        # colonne `Admis`
        ref_cell = self.gradesheet.cell(row=1, column=4)
        props = {}
        for ects in 'ABCDEF':
            props["Nombre de " + ects] = (
                '=COUNTIF({}, "{}")'
            ).format(
                self.get_column_range('Note ECTS'),
                ects
            )
        props["Nombre d'admis"] = (
            '=SUMIF({}, "<>#N/A")'.format(
                self.get_column_range('Admis')
            )
        )
        props["Effectif total"] = (
            "=COUNTA({})".format(
                self.get_column_range('Admis')
            )
        )
        props["Ratio"] = (
            '=AVERAGEIF({}, "<>#N/A")'.format(
                self.get_column_range('Admis')
            )
        )
        lower_right, statistiques = write_key_value_props(
            ref_cell,
            "Statistiques",
            props
        )

        # On écrit la note des admis uniquement pour le calcul des
        # percentiles
        self.wb.active = self.ws_data
        for i, (index, record) in enumerate(self.df.iterrows()):
            record['Note admis'].value = (
                '=IFERROR(IF({}=1, {}, ""), "")'.format(
                    get_address_of_cell(record['Admis']),
                    get_address_of_cell(record['Note agrégée'])
                )
            )

        # On écrit la colonnes des admis/refusés
        self.wb.active = self.ws_data
        for i, (index, record) in enumerate(self.df.iterrows()):
            record["Admis"].value = (
                '=IFERROR(IF(OR({}), NA(), IF(AND({}), 1, 0)), NA())'.format(
                    ", ".join(
                        "ISBLANK({})".format(
                            get_address_of_cell(record[name])
                        )
                        for name, _ in self.config['columns'].items()
                    ),
                    ", ".join(
                        "IF(ISNUMBER({}), {}>={}, 1)".format(
                            get_address_of_cell(record[name]),
                            get_address_of_cell(record[name]),
                            get_address_of_cell(
                                grades_options[name]['note_éliminatoire']
                            )
                        )
                        for name, _ in self.config['columns'].items()
                    )
                )
            )

        # On écrit la note ECTS en fonction des autres notes
        self.wb.active = self.ws_data
        for i, (index, record) in enumerate(self.df.iterrows()):
            record['Note ECTS'].value = (
                '=IF(ISNA({}), NA(), IF({}="RESERVE", "RESERVE", IF({}="ABS", "ABS", IF({}=0, "F", IF({}>={}, "A", IF({}>={}, "B", IF({}>={}, "C", IF({}>={}, "D", "E"))))))))'.format(
                    get_address_of_cell(record['Admis']),
                    get_address_of_cell(record['Note agrégée']),
                    get_address_of_cell(record['Note agrégée']),
                    get_address_of_cell(record['Admis']),
                    get_address_of_cell(record['Note agrégée']),
                    get_address_of_cell(percentiles_used["A si >="]),
                    get_address_of_cell(record['Note agrégée']),
                    get_address_of_cell(percentiles_used["B si >="]),
                    get_address_of_cell(record['Note agrégée']),
                    get_address_of_cell(percentiles_used["C si >="]),
                    get_address_of_cell(record['Note agrégée']),
                    get_address_of_cell(percentiles_used["D si >="])
                )
            )
        for cell in self.df["Note ECTS"]:
            cell.alignment = Alignment(horizontal='center')

        range = self.get_column_range('Note ECTS')
        for ects, color in zip('ABCDEF', [
                '00FF00',
                'C2FF00',
                'F7FF00',
                'FFC100',
                'FF6900',
                'FF0000'
        ]):
            fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            rule = CellIsRule(operator='equal', formula=[f'"{ects}"'], fill=fill)
            self.ws_data.conditional_formatting.add(range, rule)

        # Formattage conditionnel pour les notes éliminatoires
        self.wb.active = self.ws_data
        redFill = PatternFill(start_color='EE1111',
                              end_color='EE1111',
                              fill_type='solid')

        for name, opts in grades_options.items():
            threshold_cell = opts["note_éliminatoire"]
            threshold_addr = get_address_of_cell(
                threshold_cell,
                compat=True
            )
            self.ws_data.conditional_formatting.add(
                self.get_column_range(name),
                CellIsRule(operator='lessThan',
                           formula=[threshold_addr],
                           fill=redFill))

        # On offre la possibilité de trier
        self.wb.active = self.ws_data
        max_column = self.ws_data.max_column
        max_row = self.ws_data.max_row
        self.ws_data.auto_filter.ref = 'A1:{}{}'.format(
            utils.get_column_letter(max_column),
            max_row)
        range = self.get_column_range('Note ECTS')
        self.ws_data.auto_filter.add_sort_condition(range)

        # On sauve le classeur
        self.wb.save(self.output_file)


class GradeSheetGroupStruct(GradeSheetExamWriter):
    """Feuille de notes avec grille, par groupes avec modifications par étudiants"""

    name = 'group-struct'

    def __init__(self, args):
        super().__init__(args)
        self.group = args.group

    def get_columns(self, **kwargs):
        return {
            'Nom': 'raw',
            'Prénom': 'raw',
            'Courriel': 'raw',
            kwargs['group']: 'raw',
            self.name: 'cell'
        }

    @classmethod
    def get_parser(cls):
        parser = super(GradeSheetAssignmentWriter, GradeSheetAssignmentWriter).get_parser()
        parser.add_argument('-g', '--group', required=True, dest='group')
        return parser

    def write(self, ref=None):
        # Write new gradesheet
        self.gradesheet = self.wb.create_sheet(title=self.name)
        self.wb.active = self.gradesheet

        if ref is None:
            ref = (4, 1)
        row, col = self.write_structure(upper_left=ref)

        # First, last and total grade for column "all"
        first_all = self.gradesheet.cell(row=ref[0], column=col+1)
        total_all = self.gradesheet.cell(row=row+1, column=col+1)
        last_all = total_all.above()

        for i, (key, group) in enumerate(self.df.groupby(self.group)):
            # Name of column common to all members of group
            first_all.above(2).value = "Group"
            self.gradesheet.merge_cells2(first_all.above(2).set_border(), first_all.above(1)).center()

            # Write all names
            for j, (index, record) in enumerate(group.iterrows()):
                first_all.above(2).right(j+1).value = record['Nom']
                first_all.above().right(j+1).value = record['Prénom']
                fit_cells_at_col(first_all.above(2).right(j+1), first_all.above().right(j+1))
                frame_range(first_all.above(2).right(j+1), first_all.above().right(j+1))

            # Name of current group
            first_all.above(3).value = key
            self.gradesheet.merge_cells2(first_all.above(3).set_border(), first_all.above(3).right(len(group))).center()

            for j, (index, record) in enumerate(group.iterrows()):
                first_stu_grade = first_all.right(j+1)
                last_stu_grade = last_all.right(j+1)
                total_stu_grade = total_all.right(j+1)

                formula = "=" + "+".join(
                    "IF(ISBLANK(%s),%s,%s)" % (
                        get_address_of_cell(stu_g),
                        get_address_of_cell(all_g, absolute=True),
                        get_address_of_cell(stu_g)
                    )
                    for all_g, stu_g in zip(
                            get_segment(first_all, last_all),
                            get_segment(first_stu_grade, last_stu_grade)
                    )
                )

                total_stu_grade.value = formula

                record[self.name].value = "=" + get_address_of_cell(
                    total_stu_grade,
                    add_worksheet_name=True
                )

            frame_range(first_all, last_all.right(len(group)))
            total_all = total_all.right(len(group)+1)
            first_all = first_all.right(len(group)+1)
            last_all = total_all.above()

        self.wb.save(self.output_file)


WRITERS = [
    GradeSheetSimpleWriter,
    GradeSheetExamWriter,
    GradeSheetExamMultipleWriter,
    GradeSheetAssignmentWriter,
    GradeSheetJuryWriter,
    GradeSheetSimpleGroup,
    GradeSheetGroupStruct
]


def run(argv=sys.argv[1:], prog=os.path.basename(__file__), description=None):
    parser = argparse.ArgumentParser(prog=prog, description=description)
    subparsers = parser.add_subparsers(dest="sub_command", required=True)
    name_to_writer = {}
    for type in WRITERS:
        name_to_writer[type.name] = type
        base_parser = type.get_parser()
        subparsers.add_parser(type.name, help=type.__doc__, parents=[base_parser])
    args = parser.parse_args(argv)

    writer_class = name_to_writer[args.sub_command]
    writer = writer_class(args)
    writer.write()


if __name__ == '__main__':
    # run("--type assignment --group TPE -s median.yaml --name bar -d ../generated --uv SY02 --planning P2018".split())
    # run("--type simple --name bar -d ../generated --uv SY02 --planning P2018".split())
    # run("--type jury --name bar -c config.yaml -d ../generated --uv SY02 --planning P2018".split())
    # run()
    pass
