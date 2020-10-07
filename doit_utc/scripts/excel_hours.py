__doc__ = """
Create Excel files
"""

import openpyxl
from ..openpyxl_patched import fixit
fixit(openpyxl)

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

from ..openpyxl_utils import row_and_col

import numpy as np


def create_excel_file(filename, df):
    wb = Workbook()
    ws_remp = wb.active
    ws_remp.title = "remplacements"

    cell_ref = ws_remp["A1"]
    bottom_right, C_total = create_double_entry_table(
        ws_remp, cell_ref, df, "a remplacé x séances de Cours"
    )

    cell_ref = row_and_col(bottom_right, cell_ref).below(2)
    bottom_right, TD_total = create_double_entry_table(
        ws_remp, cell_ref, df, "a remplacé x séances de TD"
    )

    cell_ref = row_and_col(bottom_right, cell_ref).below(2)
    bottom_right, TP_total = create_double_entry_table(
        ws_remp, cell_ref, df, "a remplacé x séances de TP"
    )

    ws_sum = wb.create_sheet(title="heures")
    create_summary_table(ws_sum, "A1", df, ws_remp, C_total, TD_total, TP_total)
    wb.save(filename)


def create_double_entry_table(ws, cell_ref, df, title):
    cell_ref.value = title
    instructors = df["Intervenants"].to_list()
    instructors.append("/dev/null")

    # Diagonale grisée
    solid = PatternFill(fill_type="solid", bgColor="00000000")
    for i in range(len(instructors)):
        cell_ref.below(i + 1).right(i + 1).fill = solid

    # En-tête ligne et colonne
    ws.column_dimensions[get_column_letter(cell_ref.col_idx)].width = 20

    for i, instructor in enumerate(instructors):
        ws.column_dimensions[get_column_letter(cell_ref.col_idx + 1 + i)].width = 20
        cell_ref.below(i + 1).value = instructor
        cell_ref.right(i + 1).value = instructor

    # Instructors vs sum of cells in a row
    inst_to_cellsp = {
        inst: "+".join(
            cell_ref.right(i + 1).below(k + 1).coordinate
            for i in range(len(instructors))
            if k != i
        )
        for k, inst in enumerate(instructors)
        for i in range(len(instructors))
    }

    # Instructors vs sum of cells in a column
    inst_to_cellsm = {
        inst: "+".join(
            cell_ref.right(k + 1).below(i + 1).coordinate
            for i in range(len(instructors))
            if k != i
        )
        for k, inst in enumerate(instructors)
        for i in range(len(instructors))
    }

    # Balance des séances
    total_cell = cell_ref.right(len(instructors) + 1)
    total_cell.value = "Balance"

    inst_to_total = {}
    for i, inst in enumerate(instructors):
        cell = total_cell.below(i+1)
        cell.value = "={}-({})".format(inst_to_cellsp[inst], inst_to_cellsm[inst])
        inst_to_total[inst] = cell

    # Total des échanges en absolu
    abs_cell = cell_ref.right(len(instructors) + 2)
    abs_cell.value = "Échange"

    for i, inst in enumerate(instructors):
        cell = abs_cell.below(i+1)
        cell.value = "={}+{}".format(inst_to_cellsp[inst], inst_to_cellsm[inst])

    bottom_right = abs_cell.below(len(instructors))
    return bottom_right, inst_to_total


def get_value(cell, elt):
    if callable(elt):
        return elt(cell)
    return elt


def fill_row(refcell, *elements):
    for i, elt in enumerate(elements):
        cell = refcell.right(i)
        cell.value = get_value(cell, elt)


def create_summary_table(ws, reference, df, ws_remp, C_total, TD_total, TP_total):
    refcell = ws[reference]

    fill_row(
        refcell,
        "Intervenants",
        "Statut",
        "Cours",
        "TD",
        "TP",
        "Heures Cours prév",
        "Heures TD prév",
        "Heures TP prév",
        "UTP prév",
        "Remp. Cours",
        "Remp. TD",
        "Remp. TP",
        "Heures Cours finales",
        "Heures TD finales",
        "Heures TP finales",
        "UTP finales",
    )

    statut_mult = {
        "MCF": 1.5,
        "PR": 1.5,
        "PRAG": 1.5,
        "PRCE": 1.5,
        "PAST": 1.5,
        "ECC": 1.5,
        "Doct": 1.5,
        "ATER": 1,
        "Vacataire": 1,
        np.nan: 1,
    }

    for i, row in df.iterrows():
        cell = refcell.below(i + 1)
        inst = row["Intervenants"]
        statut = row["Statut"]
        elts = [
            inst,
            statut_mult[statut],
            row["Cours"],
            row["TD"],
            row["TP"],
            lambda cell: "=2*16*{}".format(cell.left(3).coordinate),
            lambda cell: "=2*16*{}".format(cell.left(3).coordinate),
            lambda cell: "=2*16*{}".format(cell.left(3).coordinate),
            lambda cell: "=2*16*2.25*{}+2*16*1.5*{}+2*16*{}*{}".format(
                cell.left(6).coordinate,
                cell.left(5).coordinate,
                cell.left(4).coordinate,
                cell.left(7).coordinate,
            ),
            "={}!{}".format(ws_remp.title, C_total[inst].coordinate),
            "={}!{}".format(ws_remp.title, TD_total[inst].coordinate),
            "={}!{}".format(ws_remp.title, TP_total[inst].coordinate),
            lambda cell: "={}+2*{}".format(
                cell.left(7).coordinate, cell.left(3).coordinate
            ),
            lambda cell: "={}+2*{}".format(
                cell.left(7).coordinate, cell.left(3).coordinate
            ),
            lambda cell: "={}+2*{}".format(
                cell.left(7).coordinate, cell.left(3).coordinate
            ),
            lambda cell: "=2.25*{} + 1.5*{} + {}*{}".format(
                cell.left(3).coordinate,
                cell.left(2).coordinate,
                cell.left(1).coordinate,
                cell.left(14).coordinate,
            ),
        ]
        fill_row(cell, *elts)

    cell = refcell.below(len(df.index) + 1)

    solid = PatternFill(fill_type="solid", bgColor="00000000")
    fill_row(
        cell,
        "Total",
        lambda cell: setattr(cell, "fill", solid),
        lambda cell: "=SUM({}:{})".format(
            cell.above(len(df.index)).coordinate, cell.above().coordinate
        ),
        lambda cell: "=SUM({}:{})".format(
            cell.above(len(df.index)).coordinate, cell.above().coordinate
        ),
        lambda cell: "=SUM({}:{})".format(
            cell.above(len(df.index)).coordinate, cell.above().coordinate
        ),
    )

    fill_row(cell.below(), "Prévision", lambda cell: setattr(cell, "fill", solid))

    fill_row(
        cell.below().below(),
        "Manque",
        lambda cell: setattr(cell, "fill", solid),
        lambda cell: "={}-{}".format(cell.above().coordinate, cell.above(2).coordinate),
        lambda cell: "={}-{}".format(cell.above().coordinate, cell.above(2).coordinate),
        lambda cell: "={}-{}".format(cell.above().coordinate, cell.above(2).coordinate),
    )

    fill_row(
        cell.below(3),
        "Manque (UTP)",
        lambda cell: setattr(cell, "fill", solid),
        lambda cell: "=2*2.25*16*{}".format(cell.above().coordinate),
        lambda cell: "=2*1.5*16*{}".format(cell.above().coordinate),
        lambda cell: "=2*1.5*16*{}".format(cell.above().coordinate),
    )

    fill_row(
        cell.below(4),
        "Manque (UTP)",
        lambda cell: setattr(cell, "fill", solid),
        lambda cell: setattr(cell, "fill", solid),
        lambda cell: setattr(cell, "fill", solid),
        lambda cell: "=2*16*{}".format(cell.above(2).coordinate),
    )
