"""
Fichier qui regroupe des tâches de création d'un fichier Excel sur
l'effectif d'une UV.
"""

import os
import re
import math
import random
import textwrap
import numpy as np
import pandas as pd
from unidecode import unidecode

from . import openpyxl_patched as openpyx
from openpyxl import Workbook
from openpyxl import utils
from openpyxl.utils.dataframe import dataframe_to_rows

from .utils_config import Output
from .utils import (
    sort_values,
    argument,
    check_columns,
    rel_to_dir,
    slugrot,
    slugrot_string,
)
from .tasks import UVTask, CliArgsMixin


class CsvInscrits(UVTask):
    """Construit un fichier CSV à partir des données brutes de la promo fournies par l'UTC"""

    target_name = "inscrits.csv"
    target_dir = "generated"
    unique_uv = False

    def __init__(self, planning, uv, info):
        super().__init__(planning, uv, info)
        utc_listing_fn = self.settings.AFFECTATION_LISTING
        self.utc_listing = os.path.join(
            self.settings.SEMESTER_DIR, self.uv, utc_listing_fn
        )
        self.target = self.build_target()
        self.file_dep = [self.utc_listing]

    def parse_UTC_listing(self):
        """Parse FILENAME into DataFrame"""

        if "RX_STU" in self.settings:
            RX_STU = re.compile(self.settings.RX_STU)
        else:
            # 042   NOM PRENOM            GI02
            RX_STU = re.compile(
                r"^\s*"
                r"\d{3}"
                r"\s{3}"
                r"(?P<name>.{23})"
                r"\s{3}"
                r"(?P<branche>[A-Z]{2})"
                r"(?P<semestre>[0-9]{2})"
                r"$"
            )

        if "RX_UV" in self.settings:
            RX_UV = re.compile(self.settings.RX_UV)
        else:
            # SY19       C 1   ,PL.MAX= 73 ,LIBRES=  0 ,INSCRITS= 73  H=MERCREDI 08:00-10:00,F1,S=
            RX_UV = re.compile(
                r"^\s*"
                r"(?P<uv>\w+)"
                r"\s+"
                r"(?P<course>[CTD])"
                r"\s*"
                r"(?P<number>[0-9]+)"
                r"\s*"
                r"(?P<week>[AB])?"
            )

        with open(self.utc_listing, "r") as fd:
            course_name = course_type = None
            rows = []
            for line in fd:
                m = RX_UV.match(line)
                if m:
                    number = m.group("number") or ""
                    week = m.group("week") or ""
                    course = m.group("course") or ""
                    course_name = course + number + week
                    course_type = {"C": "Cours", "D": "TD", "T": "TP"}[course]
                else:
                    m = RX_STU.match(line)
                    if m:
                        name = m.group("name").strip()
                        spe = m.group("branche")
                        sem = int(m.group("semestre"))
                        if spe == "HU":
                            spe = "HuTech"
                        elif spe == "MT":
                            spe = "ISC"
                        rows.append(
                            {
                                "Name": name,
                                "course_type": course_type,
                                "course_name": course_name,
                                "Branche": spe,
                                "Semestre": sem,
                            }
                        )

        df = pd.DataFrame(rows)
        df = pd.pivot_table(
            df,
            columns=["course_type"],
            index=["Name", "Branche", "Semestre"],
            values="course_name",
            aggfunc="first",
        )
        df = df.reset_index()

        # Il peut arriver qu'un créneau A/B ne soit pas marqué comme tel
        # car il n'a pas de pendant pour l'autre semaine. On le fixe donc
        # manuellement à A ou B.
        if "TP" in df.columns:
            semAB = [i for i in df.TP.unique() if re.match("T[0-9]{,2}[AB]", i)]
            if semAB:
                gr = [i for i in df.TP.unique() if re.match("^T[0-9]{,2}$", i)]
                rep = {}
                for g in gr:
                    while True:
                        try:
                            choice = input(f"Semaine pour le créneau {g} (A ou B) ? ")
                            if choice.upper() in ["A", "B"]:
                                rep[g] = g + choice.upper()
                            else:
                                raise ValueError
                        except ValueError:
                            continue
                        else:
                            break

                df = df.replace({"TP": rep})
        return df

    def run(self):
        if not os.path.exists(self.utc_listing):
            raise Exception("Le fichier '{0}' n'existe pas".format(
                rel_to_dir(self.utc_listing, self.settings.SEMESTER_DIR)
            ))
        df = self.parse_UTC_listing()
        with Output(self.target) as target:
            df.to_csv(target(), index=False)


class XlsStudentData(UVTask):
    """Construit le fichier Excel des données étudiants fournies par l'UTC

    Les données utilisées sont le fichier disponible sur l'ENT de
    l'effectif officiel de l'UV repéré par la variable ENT_LISTING
    dans le fichier config.py de l'UV, le fichier des affectations aux
    créneaux de Cours/TD/TP repéré par la variable AFFECTATION_LISTING
    dans le fichier config.py de l'UV et le fichier Moodle des
    inscrits à l'UV (si disponible) repéré par la variable
    MOODLE_LISTING dans le fichier config.py de l'UV.
    """

    target_dir = "generated"
    target_name = "student_data.xlsx"
    unique_uv = False

    def __init__(self, planning, uv, info):
        super().__init__(planning, uv, info)
        self.extraction_ENT = os.path.join(
            self.settings.SEMESTER_DIR, self.uv, self.settings.ENT_LISTING
        )
        self.csv_UTC = CsvInscrits.target_from(**self.info)
        self.target = self.build_target()
        self.file_dep = [self.extraction_ENT, self.csv_UTC]

        if "MOODLE_LISTING" in self.settings and self.settings.MOODLE_LISTING:
            self.csv_moodle = os.path.join(
                self.settings.SEMESTER_DIR, self.uv, self.settings.MOODLE_LISTING
            )
            self.file_dep += [self.csv_moodle]
        else:
            self.csv_moodle = None

    def run(self):
        if not os.path.exists(self.extraction_ENT):
            raise Exception("Le fichier '{}' n'existe pas".format(self.extraction_ENT))

        print("Chargement de données issues de l'ENT")
        df = pd.read_csv(self.extraction_ENT, sep="\t", encoding='ISO_8859_1')

        # Split information in 2 columns
        df[["Branche", "Semestre"]] = df.pop('Spécialité 1').str.extract(
            '(?P<Branche>[a-zA-Z]+) *(?P<Semestre>[0-9]+)',
            expand=True
        )
        df["Semestre"] = pd.to_numeric(df['Semestre'])

        # Drop unrelevant columns
        df = df.drop(['Inscription', 'Spécialité 2', 'Résultat ECTS', 'UTC', 'Réussite', 'Statut'], axis=1)

        # Drop unamed columns
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]

        print("Ajout des affectations aux Cours/TD/TP")
        df = self.add_UTC_data(df, self.csv_UTC)

        if self.csv_moodle is not None:
            print("Ajout des données issues de Moodle")
            df = self.add_moodle_data(df, self.csv_moodle)

        dff = sort_values(df, ["Nom", "Prénom"])

        with Output(self.target) as target:
            dff.to_excel(target(), index=False)

    def add_UTC_data(self, df, fn):
        "Incorpore les données Cours/TD/TP des inscrits UTC"

        print("Chargement du fichier de répartition des étudiants dans les créneaux")

        # Données issues du fichier des affectations au Cours/TD/TP
        dfu = pd.read_csv(fn)

        if "Nom" in df.columns and "Prénom" in df.columns:
            fullnames = df["Nom"] + " " + df["Prénom"]

            def slug(e):
                return unidecode(e.upper()[:23].strip())

            df["fullname_slug"] = fullnames.apply(slug)

            dfr = pd.merge(
                df,
                dfu,
                suffixes=("", "_utc"),
                how="outer",
                left_on=["fullname_slug", "Branche", "Semestre"],
                right_on=["Name", "Branche", "Semestre"],
                indicator=True,
            )

        else:
            raise Exception("")

        dfr_clean = dfr.loc[dfr["_merge"] == "both"]

        lo = dfr.loc[dfr["_merge"] == "left_only"]
        for index, row in lo.iterrows():
            key = row["fullname_slug"]
            branch = row["Branche"]
            semester = row["Semestre"]
            print(f"(`{key}`, `{branch}`, `{semester}`) not in UTC data")

        ro = dfr.loc[dfr["_merge"] == "right_only"]
        for index, row in ro.iterrows():
            key = row["Name"]
            branch = row["Branche"]
            semester = row["Semestre"]
            print(f"(`{key}`, `{branch}`, `{semester}`) only in UTC data")

        # Trying to merge manually lo and ro
        for index, row in lo.iterrows():
            fullname = row["Nom"] + " " + row["Prénom"]
            print(f"Trying to find a match for {fullname}")
            for i, (index_ro, row_ro) in enumerate(ro.iterrows()):
                fullname_ro = row_ro["Name"]
                print(f"({i}) {fullname_ro}")
            while True:
                try:
                    choice = input("Your choice? (enter if no match) ")
                    if choice and int(choice) not in range(len(ro.index)):
                        raise ValueError
                except ValueError:
                    print("Value error")
                    continue
                else:
                    break

            if choice:
                row_merge = lo.loc[index, :].combine_first(ro.iloc[int(choice), :])
                row_merge["_merge"] = "both"
                dfr_clean = dfr_clean.append(row_merge)

        dfr_clean = dfr_clean.drop(["_merge", "fullname_slug"], axis=1)

        return dfr_clean

    def add_moodle_data(self, df, fn):
        """Incorpore les données du fichier extrait de Moodle"""

        print("Chargement de données issues de Moodle")
        if fn.endswith(".csv"):
            dfm = pd.read_csv(fn)
        elif fn.endswith(".xlsx") or fn.endswith(".xls"):
            dfm = pd.read_excel(fn)

        # On laisse tomber les colonnes inintéressantes
        dfm = dfm.drop(
            ["Institution", "Département", "Dernier téléchargement depuis ce cours"],
            axis=1,
        )

        # On réalise la jointure à l'aide de l'adresse email s'il
        # elle est présente

        if "Courriel" in df.columns:
            dfr = pd.merge(
                df,
                dfm,
                suffixes=("", "_moodle"),
                how="outer",
                left_on="Courriel",
                right_on="Adresse de courriel",
                indicator=True,
            )

            dfr_clean = dfr.loc[dfr["_merge"] == "both"]

            # On affiche la différence symétrique
            lo = dfr.loc[dfr["_merge"] == "left_only"]
            for index, row in lo.iterrows():
                fullname = row["Nom"] + " " + row["Prénom"]
                print(f"add_moodle_data: {fullname} not in Moodle data")

            ro = dfr.loc[dfr["_merge"] == "right_only"]
            for index, row in ro.iterrows():
                fullname = row["Nom_moodle"] + " " + row["Prénom_moodle"]
                print(f"add_moodle_data: {fullname} only in Moodle data")

            dfr = dfr.drop("_merge", axis=1)
            dfr = dfr.loc[~pd.isnull(dfr.Nom)]

        elif "Name" in df.columns:
            fullnames = dfm["Nom"] + " " + dfm["Prénom"]

            def slug(e):
                return unidecode(e.upper()[:23].strip())

            fullnames = fullnames.apply(slug)
            dfm["fullname_slug"] = fullnames

            dfr = pd.merge(
                df,
                dfm,
                how="outer",
                left_on="Name",
                right_on="fullname_slug",
                indicator=True,
            )

            lo = dfr.loc[dfr["_merge"] == "left_only"]
            for index, row in lo.iterrows():
                fullname = row["Name"]
                print(f"add_moodle_data: {fullname} not in Moodle data")

            ro = dfr.loc[dfr["_merge"] == "right_only"]
            for index, row in ro.iterrows():
                fullname = row["Nom"] + " " + row["Prénom"]
                print(f"add_moodle_data: {fullname} only in Moodle data")

        else:
            raise Exception("Pas de colonne Courriel ou Nom, Prénom")

        # On demande à l'utilisateur de réaliser les correspondances
        for index, row in lo.iterrows():
            fullname = row["Nom"] + " " + row["Prénom"]
            print(f"Trying to find a match for {fullname}")
            for i, (index_ro, row_ro) in enumerate(ro.iterrows()):
                fullname_ro = row_ro["Nom_moodle"] + " " + row_ro["Prénom_moodle"]
                print(f"({i}) {fullname_ro}")
            while True:
                try:
                    choice = input("Your choice? (enter if no match) ")
                    if choice and int(choice) not in range(len(ro.index)):
                        raise ValueError
                except ValueError:
                    print("Value error")
                    continue
                else:
                    break

            if choice:
                row_merge = lo.loc[index, :].combine_first(ro.iloc[int(choice), :])
                row_merge["_merge"] = "both"
                dfr_clean = dfr_clean.append(row_merge)

        dfr_clean = dfr_clean.drop("_merge", axis=1)

        return dfr_clean


class XlsStudentDataMerge(UVTask):
    """Ajoute toutes les autres informations étudiants

    Ajoute les informations de changement de TD/TP, les tiers-temps et
    des informations par étudiants. Ajoute également les informations
    spécifiées dans AGGREGATE_DOCUMENTS.
    """

    target_name = "student_data_merge.xlsx"
    target_dir = "generated"
    unique_uv = False

    def __init__(self, planning, uv, info):
        super().__init__(planning, uv, info)
        self.student_data = XlsStudentData.target_from(**self.info)
        self.target = self.build_target()

        # Documents to aggregate
        self.docs = []

        if "TIERS_TEMPS" in self.settings and self.settings.TIERS_TEMPS:
            tiers_temps = self.build_dep(self.settings.TIERS_TEMPS)
            if os.path.exists(tiers_temps):
                self.docs.append((tiers_temps, self.add_tiers_temps))

        if "CHANGEMENT_TD" in self.settings and self.settings.CHANGEMENT_TD:
            TD_switches = self.build_dep(self.settings.CHANGEMENT_TD)
            if os.path.exists(TD_switches):
                self.docs.append((TD_switches, self.add_switches("TD")))

        if "CHANGEMENT_TP" in self.settings and self.settings.CHANGEMENT_TP:
            TP_switches = self.build_dep(self.settings.CHANGEMENT_TP)
            if os.path.exists(TP_switches):
                self.docs.append((TP_switches, self.add_switches("TP")))

        if "INFO_ETUDIANT" in self.settings and self.settings.INFO_ETUDIANT:
            info_etu = self.build_dep(self.settings.INFO_ETUDIANT)
            if os.path.exists(info_etu):
                self.docs.append((info_etu, self.add_student_info))

        agg_docs = (
            [
                [self.build_dep(k) if k is not None else None, v]
                for k, v in self.settings.AGGREGATE_DOCUMENTS
            ]
            if "AGGREGATE_DOCUMENTS" in self.settings
            else []
        )

        self.docs = self.docs + agg_docs
        deps = [path for path, _ in self.docs if path is not None]
        self.file_dep = deps + [self.student_data] + self.settings.config_files

    def run(self):
        df = pd.read_excel(self.student_data)

        for path, aggregater in self.docs:
            if path is None:
                print("File is None, aggregating without file")
                df = aggregater(df, None)
            elif os.path.exists(path):
                print("Aggregating '%s'" % rel_to_dir(self.build_dep(path), self.settings.SEMESTER_DIR))
                df = aggregater(df, path)
            else:
                print("WARNING: File '%s' not found!" % rel_to_dir(self.build_dep(path), self.settings.SEMESTER_DIR))

        dff = sort_values(df, ["Nom", "Prénom"])

        wb = Workbook()
        ws = wb.active

        for r in dataframe_to_rows(dff, index=False, header=True):
            ws.append(r)

        for cell in ws[1]:
            cell.style = 'Pandas'

        max_column = ws.max_column
        max_row = ws.max_row
        ws.auto_filter.ref = 'A1:{}{}'.format(
            utils.get_column_letter(max_column),
            max_row)

        # On fige la première ligne
        ws.freeze_panes = "A2"

        with Output(self.target) as target0:
            wb.save(target0())

        target = os.path.splitext(self.target)[0] + ".csv"
        with Output(target) as target:
            dff.to_csv(target(), index=False)

    def add_tiers_temps(self, df, fn):
        # Aucun tiers-temps
        df['Tiers-temps'] = False

        # Add column that acts as a primary key
        tf_df = slugrot("Nom", "Prénom")
        df["fullname_slug"] = tf_df(df)

        with open(fn, 'r') as fd:
            for line in fd:
                # Saute commentaire ou ligne vide
                line = line.strip()
                if line.startswith('#'):
                    continue
                if not line:
                    continue

                slugname = slugrot_string(line)

                res = df.loc[df.fullname_slug == slugname]
                if len(res) == 0:
                    raise Exception('Pas de correspondance pour `{:s}`'.format(line))
                elif len(res) > 1:
                    raise Exception('Plusieurs correspondances pour `{:s}`'.format(line))
                df.loc[res.index[0], 'Tiers-temps'] = True

        df = df.drop('fullname_slug', axis=1)
        return df

    def add_switches(self, ctype):
        def add_switches_ctype(df, fn):
            def swap_record(df, idx1, idx2, col):
                tmp = df.loc[idx1, col]
                df.loc[idx1, col] = df.loc[idx2, col]
                df.loc[idx2, col] = tmp

            # Add column that acts as a primary key
            tf_df = slugrot("Nom", "Prénom")
            df["fullname_slug"] = tf_df(df)
            df[f'{ctype}_orig'] = df[ctype]
            names = df[ctype].unique()

            with open(fn, 'r') as fd:
                for line in fd:
                    if line.strip().startswith('#'):
                        continue
                    if not line.strip():
                        continue

                    stu1, stu2, *t = [e.strip() for e in line.split('---')]
                    assert len(t) == 0

                    if '@etu' in stu1:
                        stu1row = df.loc[df['Courriel'] == stu1]
                        if len(stu1row) != 1:
                            raise Exception('Nombre d\'enregistrement != 1', len(stu1row), stu1)
                        stu1idx = stu1row.index[0]
                    else:
                        stu1row = df.loc[df.fullname_slug == slugrot_string(stu1)]
                        if len(stu1row) != 1:
                            raise Exception('Nombre d\'enregistrement != 1', len(stu1row), stu1)
                        stu1idx = stu1row.index[0]

                    if stu2 in names:
                        df.loc[stu1idx, ctype] = stu2
                    elif '@etu' in stu2:
                        stu2row = df.loc[df['Courriel'] == stu2]
                        if len(stu2row) != 1:
                            raise Exception('Nombre d\'enregistrement != 1', len(stu2row), stu2)
                        stu2idx = stu2row.index[0]
                        swap_record(df, stu1idx, stu2idx, ctype)
                    else:
                        stu2row = df.loc[df.fullname_slug == slugrot_string(stu2)]
                        if len(stu2row) != 1:
                            raise Exception('Nombre d\'enregistrement != 1', len(stu2row), stu2)
                        stu2idx = stu2row.index[0]
                        swap_record(df, stu1idx, stu2idx, ctype)

            df = df.drop('fullname_slug', axis=1)
            return df

        return add_switches_ctype

    def add_student_info(self, df, fn):
        df['Info'] = ""

        # Add column that acts as a primary key
        tf_df = slugrot("Nom", "Prénom")
        df["fullname_slug"] = tf_df(df)

        infos = open(fn, 'r').read()
        if infos:
            for chunk in re.split("^\\* *", infos, flags=re.MULTILINE):
                if not chunk:
                    continue

                etu, *text = chunk.split("\n", maxsplit=1)
                text = "\n".join(text).strip("\n")
                text = textwrap.dedent(text)
                slugname = slugrot_string(etu)

                res = df.loc[df.fullname_slug == slugname]
                if len(res) == 0:
                    raise Exception('Pas de correspondance pour `{:s}`'.format(etu))
                elif len(res) > 1:
                    raise Exception('Plusieurs correspondances pour `{:s}`'.format(etu))

                df.loc[res.index[0], 'Info'] = text

        df = df.drop('fullname_slug', axis=1)
        return df


class CsvExamGroups(UVTask, CliArgsMixin):
    """Fichier csv des demi-groupe de TP pour le passage des examens de TP"""

    target_dir = "generated"
    target_name = "exam_groups.csv"
    cli_args = (
        argument(
            "-t",
            "--tiers-temps",
            required=False,
            default="Tiers-temps",
            help="Nom de la colonne des tiers-temps",
        ),
        argument(
            "-p",
            "--tp",
            required=False,
            default="TP",
            help="Nom de la colonne pour réaliser un groupement",
        ),
    )

    def __init__(self, planning, uv, info):
        super().__init__(planning, uv, info)
        self.xls_merge = XlsStudentDataMerge.target_from(**self.info)
        self.target = self.build_target()
        target_name = os.path.splitext(self.target_name)[0] + '_moodle.csv'
        self.target_moodle = self.build_target(target_name=target_name)
        self.file_dep = [self.xls_merge]

    def run(self):
        df = pd.read_excel(self.xls_merge)

        def exam_split(df):
            if self.tiers_temps_col in df.columns:
                dff = df.sort_values(self.tiers_temps, ascending=False)
            else:
                dff = df
            n = len(df.index)
            m = math.ceil(n / 2)
            sg1 = dff.iloc[:m, :][self.tp] + "i"
            sg2 = dff.iloc[m:, :][self.tp] + "ii"
            dff["TPE"] = pd.concat([sg1, sg2])
            return dff

        check_columns(df, [self.tp, self.tiers_temps], file=self.xls_merge, base_dir=self.settings.SEMESTER_DIR)

        dff = df.groupby(self.tp, group_keys=False).apply(exam_split)
        if 'Adresse de courriel' in dff.columns:
            dff = dff[["Adresse de courriel", "TPE"]]  # Prefer moodle email
        else:
            dff = dff[["Courriel", "TPE"]]

        with Output(self.target) as target0:
            dff.to_csv(target0(), index=False)

        with Output(self.target_moodle) as target:
            dff.to_csv(target(), index=False, header=False)


class CsvGroups(UVTask, CliArgsMixin):
    """Fichiers csv des groupes de Cours/TD/TP/singleton pour Moodle

    Crée des fichiers csv pour chaque UV sélectionnées
    """

    target_dir = "generated"
    target_name = "{ctype}_group_moodle.csv"

    cli_args = (
        argument(
            "-g",
            "--groups",
            nargs="*",
            default=["Cours", "TD", "TP", "singleton"],
            help="Liste des groupements à considérer",
        ),
    )

    def __init__(self, planning, uv, info):
        super().__init__(planning, uv, info)
        self.xls_merge = XlsStudentDataMerge.target_from(**self.info)
        self.targets = [
            self.build_target(ctype=ctype)
            for ctype in self.groups
        ]
        self.file_dep = [self.xls_merge]

    def run(self):
        df = pd.read_excel(self.xls_merge)

        for ctype, target in zip(self.groups, self.targets):
            if ctype == "singleton":
                dff = df[["Courriel", "Login"]]

                with Output(target) as target:
                    dff.to_csv(target(), index=False, header=False)
            else:
                check_columns(df, ctype, file=self.xls_merge, base_dir=self.settings.SEMESTER_DIR)
                dff = df[["Courriel", ctype]]

                with Output(target) as target:
                    dff.to_csv(target(), index=False, header=False)


class CsvMoodleGroups(CliArgsMixin, UVTask):
    """Fichier csv de sous-groupes (binômes ou trinômes) aléatoires"""

    target_dir = "generated"
    target_name = "{course}_{project}_binomes.csv"
    cli_args = (
        argument(
            "-c",
            "--course",
            required=True,
            help="Nom de colonne pour réaliser un groupement",
        ),
        argument(
            "-p", "--project", required=True, help="Nom du groupement que sera réalisé"
        ),
        argument(
            "-g",
            "--group-names",
            required=False,
            default=None,
            help="Fichier de noms de groupes",
        ),
        argument(
            "-o",
            "--other-group",
            required=False,
            default=None,
            help="Nom de colonne d'un autre groupement",
        ),
    )

    def __init__(self, planning, uv, info):
        super().__init__(planning, uv, info)
        self.xls_merge = XlsStudentDataMerge.target_from(**self.info)
        self.target_csv = self.build_target(**self.info)
        self.target_moodle = os.path.splitext(self.target_csv)[0] + '_moodle.csv'

        # target_moodle only to avoid circular dep
        self.target = self.target_moodle
        self.file_dep = [self.xls_merge]

    def run(self):
        df = pd.read_excel(self.xls_merge)
        check_columns(
            df, self.course, file=self.xls_merge, base_dir=self.settings.SEMESTER_DIR
        )

        if self.other_group is not None:
            if not isinstance(self.other_group, list):
                other_groups = [self.other_group]

            diff = set(self.other_group) - set(df.columns.values)
            if diff:
                s = "s" if len(diff) > 1 else ""
                raise Exception(
                    f"Colonne{s} inconnue{s} : `{', '.join(diff)}'; les colonnes sont : {', '.join(df.columns)}"
                )

        # Build list of group names
        if self.group_names is not None and os.path.exists(self.group_names):
            with open(self.group_names, "r") as fd:
                group_names = [l.strip() for l in fd.readlines()]
                shuffle = True
                group_names = [f"%(gn)s_{e}" for e in group_names]
        else:
            shuffle = False
            self.group_names = [
                f"%(gn)s_{self.project}_{e}" for e in list("ABCDEFGHIJKLMNOPQRSTUVWXYZ")
            ]

        def binome_match(df_group, idx1, idx2, other_groups=None, foreign=True):
            "Renvoie vrai si le binôme est bon"

            etu1 = df_group.loc[idx1]
            etu2 = df_group.loc[idx2]

            if foreign:
                e = [
                    "DIPLOME ETAB ETRANGER SECONDAIRE",
                    "DIPLOME ETAB ETRANGER SUPERIEUR",
                    "AUTRE DIPLOME UNIVERSITAIRE DE 1ER CYCLE HORS DUT",
                ]

                # 2 étrangers == catastrophe
                if (
                    etu1["Dernier diplôme obtenu"] in e
                    and etu2["Dernier diplôme obtenu"] in e
                ):
                    return False

                # 2 GB == catastrophe
                if etu1["Branche"] == "GB" and etu2["Branche"] == "GB":
                    return False

            # Binomes précédents
            if other_groups is not None:
                for gp in other_groups:
                    if etu1[gp] == etu2[gp]:
                        return False

            return True

        def trinome_match(df_group, idx1, idx2, idx3, other_groups=None, foreign=True):
            a = binome_match(
                df_group, idx1, idx2, other_groups=other_groups, foreign=foreign
            )
            b = binome_match(
                df_group, idx1, idx3, other_groups=other_groups, foreign=foreign
            )
            c = binome_match(
                df_group, idx2, idx3, other_groups=other_groups, foreign=foreign
            )
            return a or b or c

        class Ooops(Exception):
            pass

        def add_group(df_group, other_groups=None, foreign=True, group_names=None):
            gpn = df_group.name
            index = list(df_group.index)
            if shuffle:
                random.shuffle(group_names)

            def make_random_groups(index, num=3):
                random.shuffle(index)
                num_groups = math.ceil(len(index) / num)
                return np.array_split(index, num_groups)

            def valid_groups(df_group, groups, other_groups=None, foreign=True):
                def valid_group(group):
                    if len(group) == 3:
                        return trinome_match(
                            df_group, *group, other_groups=other_groups, foreign=foreign
                        )
                    elif len(group) == 2:
                        return binome_match(
                            df_group, *group, other_groups=other_groups, foreign=foreign
                        )
                    elif len(group) == 1:
                        return True
                    else:
                        raise Exception("Size of group not supported", len(group))

                for group in groups:
                    if not valid_group(group):
                        return False
                return True

            while True:
                try:
                    groups = make_random_groups(index, num=3)

                    # On vérifie que tous les groupes sont bons
                    if not valid_groups(df_group, groups):
                        raise Ooops

                    if len(groups) > len(group_names):
                        raise Exception("Not enough group names")

                    groups_names = group_names[: len(groups)]
                    final_groups = {
                        etu: (gn % dict(gn=gpn))
                        for gn, group in zip(groups_names, groups)
                        for etu in group
                    }
                except Ooops:
                    continue
                break

            final_df_groups = pd.Series(final_groups, name="group")
            df_group = df_group[["Courriel"]]
            gb = pd.concat((df_group, final_df_groups), axis=1)
            return gb

        gdf = df.groupby(self.course)
        df = gdf.apply(
            add_group, other_groups=self.other_group, group_names=self.group_names
        )
        df = df.sort_values("group")

        with Output(self.target_csv) as target:
            df.to_csv(target(), index=False)

        with Output(self.target_moodle) as target:
            df.to_csv(target(), index=False, header=False)


class CsvGroupsGroupings(UVTask, CliArgsMixin):
    """Fichier csv de groupes et groupements à charger sur Moodle pour les créer

    Il faut spécifier le nombre de groupes dans chaque groupement avec
    l'argument `ngroups` et le nombre de groupements dans
    `ngroupings`.

    Le nom des groupements est controlé par un modèle spécifié par
    l'argument -F (par défault "D##_P1"). Les remplacements
    disponibles sont :
    - ## : remplacé par des nombres
    - @@ : remplacé par des lettres

    Le nom des groupes est controlé par un modèle spécifié par
    l'argument -f (par défault "D##_P1_@"). Les remplacements
    disponibles sont :
    - # : remplacé par des nombres
    - @ : remplacé par des lettres
    """

    target_dir = "generated"
    target_name = "groups_groupings.csv"
    cli_args = (
        argument(
            "-g",
            type=int,
            dest="ngroups",
            required=True,
            help="Nombre de groupes dans chaque groupement",
        ),
        argument(
            "-f",
            dest="ngroupsf",
            default="D##_P1_@",
            help="Format du nom de groupe (defaut: %(default)s)",
        ),
        argument(
            "-G",
            dest="ngroupings",
            type=int,
            required=True,
            help="Nombre de groupements",
        ),
        argument(
            "-F",
            dest="ngroupingsf",
            default="D##_P1",
            help="Format du nom de groupement (defaut: %(default)s)",
        ),
    )

    def __init__(self, planning, uv, info):
        super().__init__(planning, uv, info)
        self.target = self.build_target(**self.info)

    def run(self):
        letters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        ngroupings = min(26, self.ngroupings)
        ngroups = min(26, self.ngroups)

        groups = []
        groupings = []
        for G in range(ngroupings):
            grouping_letter = letters[G]
            grouping_number = str(G + 1)
            grouping = (self.ngroupingsf
                        .replace("@@", grouping_letter)
                        .replace("##", grouping_number))
            for g in range(ngroups):
                group_letter = letters[g]
                group_number = str(g + 1)
                group = (self.ngroupsf
                         .replace("@@", grouping_letter)
                         .replace("##", grouping_number)
                         .replace("@", group_letter)
                         .replace("#", group_number))

                groups.append(group)
                groupings.append(grouping)

        df_groups = pd.DataFrame({"groupname": groups, 'groupingname': self.groupings})
        with Output(self.target, protected=True) as target:
            df_groups.to_csv(target(), index=False)
