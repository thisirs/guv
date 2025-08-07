import openpyxl


def fixit(openpyxl):
    from openpyxl.cell import Cell as _Cell
    from openpyxl.cell import MergedCell as _MergedCell
    from openpyxl.styles import Alignment as _Alignment
    from openpyxl.styles import Border as _Border
    from openpyxl.styles import Side as _Side
    from openpyxl.worksheet.worksheet import Worksheet as _Worksheet

    # Custom navigation functions
    def _left(self, step=1):
        return self.offset(0, -step)


    def _right(self, step=1):
        return self.offset(0, step)


    def _above(self, step=1):
        return self.offset(-step, 0)


    def _below(self, step=1):
        return self.offset(step, 0)


    def _top(self):
        return self.parent.cell(column=self.column, row=1)


    def _text(self, value):
        self.value = value
        return self


    def _center(self):
        self.alignment = _Alignment(horizontal='center', vertical='center')
        return self


    def _set_border(self):
        thin = _Side(border_style="thin", color="000000")
        border = _Border(top=thin, left=thin, right=thin, bottom=thin)
        self.border += border
        return self

    # Add offset function to MergedCell
    def _offset(self, row=0, column=0):
        offset_column = self.column + column
        offset_row = self.row + row
        return self.parent.cell(column=offset_column, row=offset_row)


    def _merge(self, cell):
        assert self.parent == cell.parent

        self.parent.merge_cells(
            start_row=self.row,
            start_column=self.column,
            end_row=cell.row,
            end_column=cell.column
        )
        return self


    _MergedCell.offset = _offset

    _Cell.left = _left
    _Cell.right = _right
    _Cell.above = _above
    _Cell.below = _below
    _Cell.top = _top
    _Cell.text = _text
    _Cell.center = _center
    _Cell.set_border = _set_border
    _Cell.merge = _merge
    _MergedCell.left = _left
    _MergedCell.right = _right
    _MergedCell.above = _above
    _MergedCell.below = _below
    _MergedCell.top = _top
    _MergedCell.text = _text
    _MergedCell.center = _center
    _MergedCell.set_border = _set_border


    def _merge_cells2(self, cell1, cell2):
        """Merge rectangle defined by upper left and lower right cells"""

        self.merge_cells(
            start_row=cell1.row,
            start_column=cell1.col_idx,
            end_row=cell2.row,
            end_column=cell2.col_idx
        )

        return self.cell(row=cell1.row, column=cell1.col_idx)


    _Worksheet.merge_cells2 = _merge_cells2


fixit(openpyxl)

__all__ = ['openpyxl']
