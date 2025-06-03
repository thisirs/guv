import math
from itertools import groupby
from operator import attrgetter
from types import SimpleNamespace

from openpyxl import utils
from openpyxl.styles import Border, Side
from openpyxl.utils.cell import absolute_coordinate


def frame_range(cell1, cell2):
    """Draw a frame around range delimited by `cell1` and `cell2`."""

    cell_range = cell1.coordinate + ":" + cell2.coordinate
    thin = Side(border_style="thin", color="000000")
    top = Border(top=thin)
    left = Border(left=thin)
    right = Border(right=thin)
    bottom = Border(bottom=thin)

    rows = cell1.parent[cell_range]
    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right


def get_address_of_cell(cell, absolute=False, add_worksheet_name=None, compat=False):
    """Renvoie l'adresse d'un objet Cell sous forme "A1" en prenant en
    compte la feuille courante si la cellule se trouve sur une
    autre feuille.
    """

    cell_worksheet = cell.parent
    workbook = cell.parent.parent
    active_worksheet = workbook.active

    if absolute:
        coordinate = absolute_coordinate(cell.coordinate)
    else:
        coordinate = cell.coordinate

    if add_worksheet_name == False or (add_worksheet_name is None and cell_worksheet == active_worksheet):
        return coordinate
    else:
        if compat:          # GoogleSheet compatibility
            return "INDIRECT(\"'{}'!{}\")".format(cell_worksheet.title, coordinate)
        else:
            return "'{}'!{}".format(cell_worksheet.title, coordinate)


def get_range_from_cells(cell1, cell2, absolute=False, add_worksheet_name=None, compat=False):
    cell_worksheet = cell1.parent
    workbook = cell1.parent.parent
    active_worksheet = workbook.active

    # Upper-left and lower right cells
    cell1_row, cell2_row = (cell1.row, cell2.row) if cell1.row < cell2.row else (cell2.row, cell1.row)
    cell1_col, cell2_col = (cell1.column, cell2.column) if cell1.column < cell2.column else (cell2.column, cell1.column)

    # Excel-like coordinates
    cell1_addr = utils.get_column_letter(cell1_col) + str(cell1_row)
    cell2_addr = utils.get_column_letter(cell2_col) + str(cell2_row)

    if absolute:
        range = absolute_coordinate(cell1_addr) + ":" + absolute_coordinate(cell2_addr)
    else:
        range = cell1_addr + ":" + cell2_addr

    if add_worksheet_name == False or (add_worksheet_name is None and cell_worksheet == active_worksheet):
        return range
    else:
        if compat:          # GoogleSheet compatibility
            return "INDIRECT(\"'{}'!{}\")".format(cell_worksheet.title, range)
        else:
            return "'{}'!{}".format(cell_worksheet.title, range)


def get_segment(cell1, cell2):
    """Return a generator giving cells from cell1 to cell2"""

    if cell1.row == cell2.row:
        if cell1.column > cell2.column:
            cell1, cell2 = cell2, cell1

        for i in range(cell1.column, cell2.column+1):
            yield cell1.parent.cell(row=cell1.row, column=i)
    elif cell1.column == cell2.column:
        if cell1.row > cell2.row:
            cell1, cell2 = cell2, cell1

        for i in range(cell1.row, cell2.row+1):
            yield cell1.parent.cell(column=cell1.column, row=i)
    else:
        raise RuntimeError(f'Must have same row or column: {cell1.coordinate} -- {cell2.coordinate}')


def row_and_col(cell1, cell2):
    assert cell1.parent == cell2.parent
    parent = cell1.parent
    return parent.cell(row=cell1.row, column=cell2.column)


def col_and_row(cell1, cell2):
    return row_and_col(cell2, cell1)


def if_empty_formula(formula, blank_value=""):
    if formula.startswith("="):
        formula = formula[1:]

    return "=IF(ISBLANK(%s),\"%s\",%s)" % (
        formula,
        blank_value,
        formula
    )


def fit_columns_dimension(*cells):
    """Make dimensions of columns containing cells the same."""

    # Same worksheet for all cells
    worksheet = cells[0].parent
    assert all(c.parent == worksheet for c in cells)

    # Get maximum length by column
    column_lengths = {}
    cells = list(cells)
    cells.sort(key=attrgetter("column"))
    for k, g in groupby(cells, key=attrgetter("column")):
        column_lengths[k] = [len(l) for c in g if c.value is not None for l in str(c.value).splitlines()]

    # Quantile to avoid problematic cells
    lengths = sorted([max(v) if v else 0 for k, v in column_lengths.items()])
    idx = math.ceil(len(lengths) * 0.9) - 1
    final_length = lengths[idx]

    # Set all columns to final length
    for k, v in column_lengths.items():
        worksheet.column_dimensions[utils.get_column_letter(k)].width = (
            1.3 * final_length
        )


def fit_cells_at_col(*cells):
    """Adjust dimension of each column independently."""

    # Same worksheet for all cells
    worksheet = cells[0].parent
    assert all(c.parent == worksheet for c in cells)

    cells = list(cells)
    cells.sort(key=attrgetter("column"))
    for k, g in groupby(cells, key=attrgetter("column")):
        lengths = [len(l) for c in g if c.value is not None for l in str(c.value).splitlines()]
        if lengths:
            max_len = max(lengths)
            worksheet.column_dimensions[utils.get_column_letter(k)].width = 1.3*max_len


def generate_ranges(start_cell, end_cell, nranges=None):
    if start_cell.row == end_cell.row:
        if start_cell.column > end_cell.column:
            start_cell, end_cell = end_cell, start_cell

        for i in range(nranges):
            yield list(get_segment(start_cell.below(i), end_cell.below(i)))
    elif start_cell.column == end_cell.column:
        if start_cell.row > end_cell.row:
            start_cell, end_cell = end_cell, start_cell

        for i in range(nranges):
            yield list(get_segment(start_cell.right(i), end_cell.right(i)))
    else:
        raise RuntimeError(f'Must have same row or column: {start_cell.coordinate} -- {end_cell.coordinate}')


def get_row_cells(ref_cell, i, *keywords):
    """Return a dictionary of `keywords` vs cells at row `i` wrt to `ref_cell`."""

    return {
        kw: ref_cell.below(i).right(j) for j, kw in enumerate(keywords)
    }

def get_range_cells(ref_cell, i, *keywords):
    return {
        kw: get_range_from_cells(ref_cell.right(j), ref_cell.below(i).right(j))
        for j, kw in enumerate(keywords)
    }

def fill_row(row_cells, **elts):
    """Fill cells in dictionary `row_cells` with elements from dictionary `elts`.

    `row_cells` is a dictionary of keywords vs cells and `elts` a
    dictionary of keywords vs values. If values are callables, they are
    called with `row_cells`.

    """

    for column_name, value in elts.items():
        if column_name not in row_cells:
            raise ValueError(f"Keyword `{column_name}` in `elts` but not in `row_cells`")

        cell = row_cells[column_name]
        if callable(value):
            value = value(row_cells)

        cell.value = value


class Block():
    def __init__(self, upper_left_cell=None, lower_right_cell=None, **kwargs):
        self.upper_left_cell = upper_left_cell
        self.lower_right_cell = lower_right_cell
        self._content = kwargs

    @property
    def content(self):
        return SimpleNamespace(**self._content)

    def set_content(self, **kwargs):
        self._content = kwargs

    @property
    def lower_left(self):
        return row_and_col(self.lower_right_cell, self.upper_left_cell)

    @property
    def lower_right(self):
        return self.lower_right_cell

    @property
    def upper_left(self):
        return self.upper_left_cell
