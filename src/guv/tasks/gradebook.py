import json
import os

import guv
import jsonschema
import openpyxl
import yaml

from ..openpyxl_patched import fixit

fixit(openpyxl)

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

from ..openpyxl_utils import (fit_columns_dimension, frame_range, generate_ranges,
                              get_address_of_cell, get_range_from_cells,
                              get_segment, row_and_col)
from ..translations import _, TaskDocstring, _file
from ..utils import sort_values, normalize_string, generate_groupby
from ..utils_ask import checkboxlist_prompt, prompt_number
from ..utils_config import rel_to_dir, ask_choice
from . import base
from . import base_gradebook as baseg


__all__ = ["XlsGradeBookGroup", "XlsGradeBookJury", "XlsGradeBookNoGroup"]


def generate_tree_blocks_coordinates(tree):
    """Generate coordinates of rectangles according to a tree"""

    def compute_depth(tree):
        if isinstance(tree, dict):
            return 1 + max(compute_depth(child) for child in tree.values())
        else:
            return 1

    def compute_n_leaves(tree):
        if isinstance(tree, dict):
            return sum(compute_n_leaves(child) for child in tree.values())
        else:
            return 1

    depth = compute_depth(tree) - 1

    def generate_tree_blocks_coordinates_aux(name, tree, i, j):
        remaining_width = depth - j
        tree_depth = compute_depth(tree)
        width = int(remaining_width / tree_depth)

        if not isinstance(tree, dict):
            yield name, i, j, 1, width
        else:
            if name != "root":
                height = compute_n_leaves(tree)
                yield name, i, j, height, width
            else:
                width = 0

            acc_height = 0
            for child_name, child in tree.items():
                yield from generate_tree_blocks_coordinates_aux(child_name, child, i + acc_height, j + width)
                acc_height += compute_n_leaves(child)

    yield from generate_tree_blocks_coordinates_aux("root", tree, 0, 0)


def get_values(tree, prop, default=1):
    "Return list of terminal values of property"

    if isinstance(tree, dict):
        return [pts for node in tree.values() for pts in get_values(node, prop, default=default)]
    else:
        # Should be a list of dict, concatenating
        a = {}
        for s in tree:
            a.update(s)
        return [a.get(prop, default)]


class MarkingScheme:
    def __init__(self, name, tree):
        self.name = name
        self.tree = tree
        self.points_cells = None
        self.scale_cells = None
        self.bottom_right = None

    @property
    def scales(self):
        return get_values(self.tree, "scale")

    @property
    def points(self):
        """Return depth-first list of points"""
        return get_values(self.tree, "points")

    @property
    def n_grades(self):
        return len(self.points)

    def write(self, ref):
        self.top_left = ref

        # Write marking scheme
        bottom_right = self.write_marking_scheme(ref)

        ref_points = row_and_col(ref, bottom_right).right()
        self.write_points_column(ref_points)

        ref_scales = row_and_col(ref, bottom_right).right(2)
        self.write_scales_column(ref_scales)

        self.bottom_right = bottom_right.right(2)
        return self.bottom_right

    def write_marking_scheme(self, ref):
        """Write marking scheme, return lower-right cell"""

        row = ref.row
        col = ref.col_idx
        maxi = maxj = -1
        worksheet = ref.parent

        for name, i, j, di, dj in generate_tree_blocks_coordinates(self.tree):
            maxi = max(maxi, i + di)
            maxj = max(maxj, j + dj)
            worksheet.merge_cells(
                start_row=row + i,
                start_column=col + j,
                end_row=row + i + di - 1,
                end_column=col + j + dj - 1,
            )
            worksheet.cell(row=row + i, column=col + j).value = name
            al = Alignment(horizontal="center", vertical="center")
            worksheet.cell(row=row + i, column=col + j).alignment = al

        return worksheet.cell(row=row + maxi - 1, column=col + maxj - 1)

    def write_points_column(self, ref):
        # Column of points
        ref.above().text(_("Points"))

        for i, points in enumerate(self.points):
            ref.below(i).text(points)
        self.points_cells = [ref.below(i) for i in range(len(self.points))]

        ref_points_last = ref.below(len(self.points) - 1)

        range_cell = get_range_from_cells(ref, ref_points_last)
        self.global_total = ref_points_last.below(2).text(f"=SUM({range_cell})")
        self.global_total_rescale = ref_points_last.below(3).text(20)

        ref_points_last.below(2).left().text(_("Grade"))
        ref_points_last.below(3).left().text(_("Grade /20"))

    def write_scales_column(self, ref):
        # Column of scales
        ref.above().text(_("Scale"))

        for i, scale in enumerate(self.scales):
            ref.below(i).text(scale)
        self.scale_cells = [ref.below(i) for i in range(len(self.scales))]

    def get_formula(self, cells):
        subformula = " + ".join(
            "{points} * {grade} / {scale}".format(
                points=get_address_of_cell(points_cell, absolute=True),
                grade=get_address_of_cell(grade_cell),
                scale=get_address_of_cell(scale_cell, absolute=True)
            )
            for grade_cell, points_cell, scale_cell in zip(cells, self.points_cells, self.scale_cells)
        )
        return subformula


class XlsGradeBookNoGroup(baseg.AbstractGradeBook, base.MultipleConfigOpt):
    __doc__ = TaskDocstring()

    config_argname = "--marking-scheme"
    config_help = _("Files containing the detailed grading scales")
    config_required = False
    config_number = _("How many grading scales? ")
    config_num = _("Grading scale {i}")

    def __init__(self, planning, uv, info):
        super().__init__(planning, uv, info)

    def ask_one_config(self):
        def ask_subitem(item_name=None):
            subitems = {}
            if item_name is not None:
                prompt = _("Name of a sub-part/question of `{item_name}` (nothing if question)? ").format(item_name=item_name)
                name = input(prompt)
            else:
                prompt = _("Name of a part/question? ")
                name = None
                while not name:
                    name = input(prompt)

            if name:
                if item_name is not None:
                    prompt = _("Name of another sub-part/question of `{item_name}`? ").format(item_name=item_name)
                else:
                    prompt = _("Name of another part/question? ")

                subitems[name] = ask_subitem(item_name=name)
                while True:
                    name = input(prompt)
                    if name:
                        subitems[name] = ask_subitem(item_name=name)
                    else:
                        break
            else:
                points = prompt_number(_("Points for the question `{item_name}`: ").format(item_name=item_name), default="1")
                scale = prompt_number(_("Scale for the question `{item_name}`: ").format(item_name=item_name), default=str(points))
                subitems = [{"points": points, "scale": scale}]

            return subitems

        while True:
            config = ask_subitem()
            print(_("Constructed YAML file:"))
            print(yaml.dump(config))
            result = ask_choice(_("Validate? (y/n) "), {"y": True, "n": False})
            if result:
                break

        return self.validate_config(config)

    def validate_config(self, config):
        """Validation du fichier de configuration"""

        tmpl_dir = os.path.join(guv.__path__[0], "schemas")
        schema_file = os.path.join(tmpl_dir, "gradebook_marking_schema.json")
        schema = json.load(open(schema_file, "rb"))

        jsonschema.validate(config, schema)

        def set_default(elt):
            if isinstance(elt, dict):
                for key, value in elt.items():
                    if value is None:
                        elt[key] = [{
                            "points": 1,
                            "scale": 1
                        }]
                    else:
                        set_default(value)
            elif isinstance(elt, list):
                if all("points" not in d for d in elt):
                    elt.append({"points": 1})
                if all("scale" not in d for d in elt):
                    elt.append({"scale": 1})
            else:
                raise TypeError

        set_default(config)

        return config

    def get_columns(self):
        columns = super().get_columns()

        # Add order_by column if specified in command line and don't include it
        # in worksheet
        if self.order_by is not None:
            columns.append((self.order_by, "hide", 5))

        if self.extra_cols is not None:
            for colname in self.extra_cols:
                columns.append((colname, "raw", 6))

        if self.group_by is not None:
            columns.append((self.group_by, "raw", 6))

        for i, ms in enumerate(self.marking_schemes):
            columns.append((ms.name + " " + _("raw"), "cell", 100+i))
            columns.append((ms.name, "cell", 200+i))

        # Say which columns are to be aggregated
        self.agg_colname = [ms.name for ms in self.marking_schemes]
        if len(self.marking_schemes) > 1:
            columns.append((_("final grade"), "cell", 1000))
            self.agg_colname += [_("final grade")]

        return columns

    def add_arguments(self):
        super().add_arguments()

        self.add_argument(
            "-o",
            "--order-by",
            metavar="colname",
            required=False,
            help=_("Column used to order the names in each sheet")
        )

        self.add_argument(
            "-w",
            "--worksheets",
            required=False,
            metavar="colname",
            dest="group_by",
            help=_("Column used to group into multiple sheets")
        )

        self.add_argument(
            "-e",
            "--extra-cols",
            metavar="COL,[COL,...]",
            type=lambda t: [s.strip() for s in t.split(",")],
            help=_("Additional columns to include in the grade sheet")
        )

    @property
    def marking_schemes(self):
        ms = [
            MarkingScheme(_("part {i}").format(i=i+1), conf) for i, conf in enumerate(self.config)
        ]
        if len(ms) == 1:
            ms[0].name = _("grade")
        return ms

    def create_first_worksheet(self):
        # Check columns first before maybe asking interactively for a marking
        # scheme
        if self.group_by is not None:
            self.check_if_present(self.data_df, self.group_by)

        if self.order_by is not None:
            self.check_if_present(self.data_df, self.order_by)

        if self.extra_cols is not None:
            self.check_if_present(self.data_df, self.extra_cols)

        super().create_first_worksheet()

    def create_other_worksheets(self):
        """Create one or more worksheets based on `worksheets` and `marking_scheme` arguments."""

        order_by = self.order_by if self.order_by is not None else self.settings.LASTNAME_COLUMN

        if self.group_by is not None:
            gen_group = list(generate_groupby(self.first_df, self.group_by))
        else:
            gen_group = [("", self.first_df)]

        for ms in self.marking_schemes:
            for name, group in gen_group:
                group = sort_values(group, [order_by])
                self.create_worksheet(name, ms, group)

    def create_worksheet(self, name, ms, group):
        """Create one worksheet for group with marking scheme and name."""

        worksheet_name = name if name else ms.name
        worksheet_name = normalize_string(worksheet_name, type="excel")
        gradesheet = self.workbook.create_sheet(title=worksheet_name)

        # Make current worksheet the default one, useful for get_address_of_cell
        self.workbook.active = gradesheet

        # Leave room from statistics
        ref_stats = gradesheet.cell(4, 2)
        formula = '=IF(ISERROR(QUARTILE({{marks_range}}, {num})), "", QUARTILE({{marks_range}}, {num}))'
        stats = [
            (name, formula.format(num=num))
            for name, num in ((_("Min"), 0), (_("Q1"), 1), (_("median"), 2), (_("Q3"), 3), (_("Max"), 4))
        ]
        ref_marking_scheme = ref_stats.right(len(stats))

        bottom_right = ms.write(ref_marking_scheme)
        ref = row_and_col(ref_marking_scheme, bottom_right).right()

        # Freeze the structure
        gradesheet.freeze_panes = ref.top()

        def insert_record(ref_cell, i, record):
            """Insert a column in worksheet"""

            # Header, last name and first name
            index = ref_cell.text(_("Student {i}").format(i=i))
            last_name = index.below().text(record[self.settings.LASTNAME_COLUMN])
            first_name = last_name.below().text(record[self.settings.NAME_COLUMN])

            # Other important cells
            first_grade = first_name.below()
            last_grade = row_and_col(ms.bottom_right, first_grade)
            total = last_grade.below(2)
            total_rescaled = total.below()

            # Formula to compute grade with points/scale
            cells = get_segment(first_grade, last_grade)
            subformula = ms.get_formula(cells)

            # Use COUNTBLANK to display grade once every points is available
            marks_range = get_range_from_cells(first_grade, last_grade)
            formula = '=IF(COUNTBLANK({marks_range}) > 0, "", {subformula})'.format(
                marks_range=marks_range,
                subformula=subformula
            )
            total.value = formula

            total_rescaled.text(
                '=IF(ISTEXT({stu_total}),"",{stu_total}/{global_total}*{rescaling})'.format(
                    stu_total=get_address_of_cell(total),
                    global_total=get_address_of_cell(ms.global_total, absolute=True),
                    rescaling=get_address_of_cell(ms.global_total_rescale, absolute=True)
                )
            )

            # Link total_rescaled to cell in first worksheet
            cell = record[ms.name]
            cell.value = "=" + get_address_of_cell(
                total_rescaled, add_worksheet_name=True, absolute=True
            )

            # Link total to cell in first worksheet
            cell = record[ms.name + " " + _("raw")]
            cell.value = "=" + get_address_of_cell(
                total, add_worksheet_name=True, absolute=True
            )

            return total_rescaled

        ref_cells = []
        for j, (index, record) in enumerate(group.iterrows()):
            ref_cell = ref.right(j).above(3)
            ref_cells.append(ref_cell)
            last_cell = insert_record(ref_cell, j + 1, record)

        # Add statistics of grades
        for i, (first, last) in enumerate(zip(
            get_segment(ref, row_and_col(last_cell, ref)),
            get_segment(row_and_col(ref, last_cell), last_cell),
        )):
            marks_range = get_range_from_cells(first, last)
            for j, (name, formula) in enumerate(stats):
                if i == 0:
                    ref_stats.right(j).below(-1).text(name).center()
                ref_stats.right(j).below(i).text(formula.format(marks_range=marks_range))

        # Set column dimensions
        fit_columns_dimension(*[c for ref in ref_cells for c in [ref.below(1), ref.below(2)]])

        # Around grades
        frame_range(ref.above(3), last_cell.above(3))

        # Around totals
        frame_range(row_and_col(last_cell, ref).above(), last_cell)


class XlsGradeBookGroup(XlsGradeBookNoGroup):
    __doc__ = TaskDocstring()

    config_argname = "--marking-scheme"

    def get_columns(self):
        columns = super().get_columns()
        columns.append((self.subgroup_by, "raw", 5))
        return columns

    def add_arguments(self):
        super().add_arguments()

        self.add_argument(
            "-g",
            "--group-by",
            dest="subgroup_by",
            metavar="colname",
            required=True,
            help=_("Group column used to grade groups of students")
        )

    def create_first_worksheet(self):
        # Check columns first before maybe asking interactively for a marking
        # scheme
        if self.subgroup_by is not None:
            self.check_if_present(self.data_df, self.subgroup_by)

        super().create_first_worksheet()

    def create_worksheet(self, name, ms, group):
        if name:
            worksheet_name = ms.name + " " + name
        else:
            worksheet_name = ms.name

        worksheet_name = normalize_string(worksheet_name, type="excel")
        gradesheet = self.workbook.create_sheet(title=worksheet_name)

        # Make current worksheet the default one, useful for get_address_of_cell
        self.workbook.active = gradesheet

        # Leave room from statistics
        ref_marking_scheme = gradesheet.cell(4, 2)

        # Write marking scheme and points/scale
        bottom_right = ms.write(ref_marking_scheme)

        # Write group block
        ref = row_and_col(ref_marking_scheme, bottom_right).right().above(3)

        # Freeze the structure
        gradesheet.freeze_panes = ref.top()

        blocks = []
        for i, (subname, subgroup) in enumerate(group.groupby(self.subgroup_by)):
            if self.order_by is not None:
                group = group.sort_values(self.order_by)

            header = _("Group {i}").format(i=i+1)

            block = GroupBlock(
                marking_scheme=ms,
                group_name=subname,
                group=subgroup,
                header=header,
                settings=self.settings
            )
            blocks.append(block)
            block.write(ref)

            ref = row_and_col(ref, block.bottom_right.right())


class GroupBlock:
    def __init__(self, marking_scheme, group_name, group, header, settings):
        self.marking_scheme = marking_scheme
        self.header = header
        self.group_name = group_name
        self.group = group
        self.settings = settings

        self.bottom_right = None

    @property
    def n_grades(self):
        return self.marking_scheme.n_grades

    @property
    def total_height(self):
        return 3 + self.n_grades + 1 + 2

    def grade_cells(self, i):
        cell1 = self.first_student.below(3+i)
        cell2 = cell1.right(self.N - 1)
        return list(get_segment(cell1, cell2))

    def write(self, ref_cell):
        self.N = len(self.group)  # Number of students

        self.first_student = ref_cell.right()
        self.last_student = ref_cell.right(self.N)

        # First column is group column
        group_range = list(get_segment(ref_cell, ref_cell.below(self.total_height-1)))

        # Important cells
        group_idx = group_range[0]
        group_name = group_range[1]
        group_first_grade = group_range[3]
        group_last_grade = group_range[-4]
        group_total = group_range[-2]
        group_total_rescaled = group_range[-1]

        # Header
        group_idx.text(self.header).merge(self.last_student).center()

        # Group name
        group_name.text(self.group_name).merge(group_name.below()).center()

        group_grade_cells = list(get_segment(group_first_grade, group_last_grade))
        subformula = self.marking_scheme.get_formula(group_grade_cells)

        # Use COUNTBLANK to display total grades once every grade is available
        marks_range = get_range_from_cells(group_first_grade, group_last_grade)
        formula = '=IF(COUNTBLANK({marks_range}) > 0, "", {subformula})'.format(
            marks_range=marks_range,
            subformula=subformula
        )
        group_total.value = formula

        group_total_rescaled.text(
            '=IF(ISTEXT({group_total}),"",{group_total}/{global_total}*{rescaling})'.format(
                group_total=get_address_of_cell(group_total),
                global_total=get_address_of_cell(self.marking_scheme.global_total, absolute=True),
                rescaling=get_address_of_cell(self.marking_scheme.global_total_rescale, absolute=True)
            )
        )

        # Next columns are per-student columns
        gen = generate_ranges(self.first_student, self.first_student.below(self.total_height-1), nranges=self.N)

        # Group student dataframe record and corresponding column range
        for stu_range, (index, record) in zip(gen, self.group.iterrows()):
            # Important cells
            idx_cell = stu_range[0]
            lastname_cell = stu_range[1]
            name_cell = stu_range[2]
            stu_first_grade = stu_range[3]
            stu_last_grade = stu_range[-4]
            stu_total = stu_range[-2]
            stu_total_rescaled = stu_range[-1]

            lastname_cell.value = record[self.settings.LASTNAME_COLUMN]
            name_cell.value = record[self.settings.NAME_COLUMN]

            # Mirror group grade
            stu_grade_cells = list(get_segment(stu_first_grade, stu_last_grade))
            for group_cell, stu_cell in zip(group_grade_cells, stu_grade_cells):
                stu_cell.value = '=IF(ISBLANK({addr}),"",{addr})'.format(
                    addr=get_address_of_cell(group_cell)
                )

            # Total of student
            subformula = self.marking_scheme.get_formula(stu_grade_cells)

            # Use COUNTBLANK to display grade once every points is available
            marks_range = get_range_from_cells(stu_first_grade, stu_last_grade)
            formula = '=IF(COUNTBLANK({marks_range}) > 0, "", {subformula})'.format(
                marks_range=marks_range,
                subformula=subformula
            )
            stu_total.value = formula

            # Total of student rescaled
            stu_total_rescaled.text(
                '=IF(ISTEXT({stu_total}),"",{stu_total}/{global_total}*{rescaling})'.format(
                    stu_total=get_address_of_cell(stu_total),
                    global_total=get_address_of_cell(self.marking_scheme.global_total, absolute=True),
                    rescaling=get_address_of_cell(self.marking_scheme.global_total_rescale, absolute=True)
                )
            )

            record[self.marking_scheme.name].value = "=" + get_address_of_cell(
                stu_total_rescaled, add_worksheet_name=True
            )
            record[self.marking_scheme.name + " " + _("raw")].value = "=" + get_address_of_cell(
                stu_total, add_worksheet_name=True
            )

        bottom_right = ref_cell.below(self.total_height - 1).right(self.N)
        self.bottom_right = bottom_right

        # Around grades
        frame_range(ref_cell, stu_last_grade)

        # Around totals
        frame_range(group_total, stu_total_rescaled)


class XlsGradeBookJury(baseg.AbstractGradeBook, base.ConfigOpt):
    __doc__ = TaskDocstring()

    config_help = _("Configuration file specifying the grades to use")
    config_required = False
    name_required = False
    name_default = _("jury")

    def __init__(self, planning, uv, info):
        super().__init__(planning, uv, info)

    def message(self, target):
        return _file("XlsGradeBookJury_message").format(
            filename=rel_to_dir(target, self.settings.UV_DIR)
        )

    def get_columns(self, **kwargs):
        # Les colonnes classiques
        columns = [(self.settings.LASTNAME_COLUMN, "raw", 0), (self.settings.NAME_COLUMN, "raw", 0), (self.settings.EMAIL_COLUMN, "raw", 0)]

        # Les colonnes nécessaires pour les différents calculs
        columns.append((_("Passed"), "cell", 2))
        columns.append((_("Passing grade"), "cell", 2))

        # Les colonnes spécifiées dans le fichier de configuration
        for grade in self.config["grades"]:
            name = grade["name"]
            columns.append((name, "cell", 3))

        for other in self.config["others"]:
            columns.append((other, "raw", 1))

        # La colonne de la note finale : A, B, C, D, E, F
        columns.append((_("ECTS grade"), "cell", 4))

        self.agg_colname = [_("Aggregated grade"), _("ECTS grade")]

        return columns

    def create_other_worksheets(self):
        self.create_second_worksheet()
        self.update_first_worksheet()

    def ask_config(self):
        cols = self.data_df.columns.values.tolist()

        # Remove some columns, keep it ordered
        grade_cols = cols.copy()
        for c in [
                self.settings.NAME_COLUMN,
                self.settings.LASTNAME_COLUMN,
                self.settings.EMAIL_COLUMN,
                self.settings.LOGIN_COLUMN
        ]:
            if c in grade_cols:
                grade_cols.remove(c)

        grades = checkboxlist_prompt(
            _("Indicate the columns containing grades to use directly for the final grade (SPACE: select, ENTER: validate):"),
            [(c, c) for c in grade_cols]
        )
        result = ask_choice(
            _("Add intermediate grades to use directly for the final grade and not present in the file ``effectif.xlsx``? (y/n) "),
            {"y": True, "n": False},
        )
        if result:
            while True:
                choice = input(_("Name of the intermediate grade? (leave blank to finish) "))
                if not choice:
                    break
                if choice not in grades:
                    grades.append(choice)
                else:
                    print(_("Column already present"))

        grades_props = []
        for grade in grades:
            props = {"name": grade}
            props["coefficient"] = prompt_number(_("Coefficient for {grade}: ").format(grade=grade), default="1")
            props["passing grade"] = prompt_number(_("Passing grade for {grade}: ").format(grade=grade), default="-1")
            grades_props.append(props)

        other_cols = cols.copy()
        for c in [self.settings.LASTNAME_COLUMN, self.settings.NAME_COLUMN, self.settings.EMAIL_COLUMN,]:
            if c in other_cols:
                other_cols.remove(c)
        for c in grades:
            if c in other_cols:
                other_cols.remove(c)

        others = checkboxlist_prompt(
            _("Indicate other columns for information (project group, tutorial group,...):"),
            [(c, c) for c in other_cols]
        )

        return self.validate_config({
            "grades": grades_props,
            "others": others
        })

    def validate_config(self, config):
        """Validation du fichier de configuration

        grades:
          - name: grade1
            passing grade: 8
            coefficient: 2
          - name: grade2
          - name: grade3
        others:
          - info
        """

        tmpl_dir = os.path.join(guv.__path__[0], "schemas")
        schema_file = os.path.join(tmpl_dir, "gradebook_jury_schema.json")
        schema = json.load(open(schema_file, "rb"))

        jsonschema.validate(config, schema)

        # Add default value in grades
        if _("Percentile A") not in config:
            config[_("Percentile grade A")] = .9
        if _("Percentile grade B") not in config:
            config[_("Percentile grade B")] = .65
        if _("Percentile grade C") not in config:
            config[_("Percentile grade C")] = .35
        if _("Percentile grade D") not in config:
            config[_("Percentile grade D")] = .1

        for grade in config["grades"]:
            if "coefficient" not in grade:
                grade["coefficient"] = 1
            if "maximum grade" not in grade:
                grade["maximum grade"] = 20
            if "passing grade" not in grade:
                grade["passing grade"] = -1

        config["grades"].append({
            "name": _("Aggregated grade"),
            "passing grade": -1
        })

        if "others" not in config:
            config["others"] = []

        return config

    @property
    def grade_columns(self):
        """Colonnes associées à des notes"""

        for props in self.config["grades"]:
            props2 = props.copy()
            name = props2["name"]
            del props2["name"]
            yield name, props2

    def write_key_value_props(self, ref_cell, title, props):
        """Helper function to write key-value table.

        Written at `ref_cell` with header `title` and key-value in `props`.
        Return lower-right cell and dictionary of property name to actual cell.

        """

        keytocell = {}
        ref_cell.text(title).merge(ref_cell.right()).style = "Pandas"
        current_cell = ref_cell
        for key, value in props.items():
            current_cell = current_cell.below()
            current_cell.value = key
            current_cell.right().value = value
            keytocell[key] = current_cell.right()

        frame_range(ref_cell, current_cell.right())
        return current_cell.right(), keytocell

    def create_second_worksheet(self):
        # Write new gradesheet
        self.gradesheet = self.workbook.create_sheet(title=_("Parameters"))
        self.workbook.active = self.gradesheet
        current_cell = self.gradesheet.cell(row=1, column=1)

        # Write option blocks for grade columns
        self.grades_options = {}
        for name, props in self.grade_columns:
            # Add some stats on marks
            for stat, q in ((_("Min"), 0), (_("Q1"), 1), (_("Median"), 2), (_("Q3"), 3), (_("Max"), 4)):
                props[stat] = '=IF(ISERROR(QUARTILE({0}, 0)), NA(), QUARTILE({0}, {1}))'.format(
                    self.get_column_range(name),
                    q
                )

            # Write key-value table
            keys = ["name", _("passing grade"), "coefficient", _("maximum grade")]
            ordered_props = {k: props[k] for k in keys if k in props}
            rest = {k: props[k] for k in props if k not in keys}
            lower_right, keytocell = self.write_key_value_props(
                current_cell, name, ordered_props | rest
            )
            self.grades_options[name] = keytocell
            current_cell = current_cell.right(3)

        # Maximum number of options
        max_height = max(len(v) for k, v in self.grades_options.items())

        # Add default global options block
        current_cell = self.gradesheet.cell(row=max_height + 4, column=1)
        options = self.config.copy()
        del options["grades"]
        if "others" in options:
            del options["others"]
        lower_right, self.global_options = self.write_key_value_props(
            current_cell, _("Global options"), options
        )
        current_cell = current_cell.right(3)

        # On écrit les seuils des notes correspondants aux percentiles choisis
        props = {}
        for i, ects in enumerate("ABCD"):
            percentile_cell = self.global_options[_("Percentile grade") + " " + ects]
            props[
                _("{ects} if >=").format(ects=ects)
            ] = "=IF(ISERROR(PERCENTILE({a}, {b})), NA(), PERCENTILE({a}, {b}))".format(
                a=self.get_column_range(_("Passing grade")),
                b=get_address_of_cell(percentile_cell),
            )

        lower_right, percentiles_theo = self.write_key_value_props(
            current_cell, _("Theoretical percentiles"), props
        )
        current_cell = current_cell.right(3)

        # Percentiles effectifs
        props = {}
        for i, ects in enumerate("ABCD"):
            props[_("{ects} if >=").format(ects=ects)] = "=" + get_address_of_cell(
                percentiles_theo[_("{ects} if >=").format(ects=ects)]
            )

        lower_right, self.percentiles_used = self.write_key_value_props(
            current_cell, _("Used percentiles"), props
        )
        current_cell = current_cell.right(3)

        # On écrit les proportions de notes ECTS en fonction de la
        # colonne `Admis`
        props = {}
        for ects in "ABCDEF":
            props[_("Number of {ects}").format(ects=ects)] = ('=COUNTIF({}, "{}")').format(
                self.get_column_range(_("ECTS grade")), ects
            )
        props[_("Number of passed")] = '=SUMIF({}, "<>#N/A")'.format(
            self.get_column_range(_("Passed"))
        )
        props[_("Total number")] = "=COUNTA({})".format(self.get_column_range(_("Passed")))
        props[_("Ratio")] = '=IF(ISERROR(AVERAGEIF({0}, "<>#N/A")), NA(), AVERAGEIF({0}, "<>#N/A"))'.format(
            self.get_column_range(_("Passed"))
        )
        lower_right, statistiques = self.write_key_value_props(
            current_cell, _("Statistics"), props
        )

    def update_first_worksheet(self):
        # Pour que get_address_of_cell marche correctement
        self.workbook.active = self.first_ws

        # On écrit la note agrégée basée sur la note maximum et le coefficient
        # de chaque note
        coef_sum = "+".join(
            get_address_of_cell(
                self.grades_options[name]["coefficient"],
                absolute=True,
            )
            for name, props in self.grade_columns
            if name != _("Aggregated grade")
        )

        for i, (index, record) in enumerate(self.first_df.iterrows()):
            formula = "=(" + "+".join(
                "{coef}*{grade}/{grade_max}*20".format(
                    coef=get_address_of_cell(
                        self.grades_options[name]["coefficient"],
                        absolute=True,
                    ),
                    grade=get_address_of_cell(record[name]),
                    grade_max=get_address_of_cell(
                        self.grades_options[name]["maximum grade"],
                        absolute=True,
                    ),
                )
                for name, props in self.grade_columns
                if name != _("Aggregated grade")
            ) + f")/({coef_sum})"

            record[_("Aggregated grade")].value = formula

        # On écrit la colonne "Admis" des admis/refusés basée sur les
        # barres
        for i, (index, record) in enumerate(self.first_df.iterrows()):
            # Teste si une des notes est vide
            any_blank_cell = "OR({})".format(
                ", ".join(
                    "ISBLANK({})".format(get_address_of_cell(record[name]))
                    for name, props in self.grade_columns
                )
            )

            # Teste si toutes les barres sont atteintes
            above_all_threshold = "AND({})".format(
                ", ".join(
                    "IF(ISNUMBER({0}), {0}>={1}, 1)".format(
                        get_address_of_cell(record[name]),
                        get_address_of_cell(
                            self.grades_options[name]["passing grade"],
                            absolute=True,
                        ),
                    )
                    for name, props in self.grade_columns
                )
            )

            # NA() s'il y a un problème, 0 si recalé, 1 si reçu
            formula = f"=IFERROR(IF({any_blank_cell}, NA(), IF({above_all_threshold}, 1, 0)), NA())"
            record[_("Passed")].value = formula

        # On écrit la note agrégée des admis dans la colonne "Note
        # admis" pour faciliter le calcul des percentiles sur les
        # admis
        for i, (index, record) in enumerate(self.first_df.iterrows()):
            record[_("Passing grade")].value = '=IFERROR(IF({}=1, {}, ""), "")'.format(
                get_address_of_cell(record[_("Passed")]),
                get_address_of_cell(record[_("Aggregated grade")]),
            )

        # On écrit la note ECTS en fonction de la note agrégée et des
        # percentiles utilisés
        perc_A = get_address_of_cell(self.percentiles_used[_("A if >=")], absolute=True)
        perc_B = get_address_of_cell(self.percentiles_used[_("B if >=")], absolute=True)
        perc_C = get_address_of_cell(self.percentiles_used[_("C if >=")], absolute=True)
        perc_D = get_address_of_cell(self.percentiles_used[_("D if >=")], absolute=True)
        percs = dict(perc_A=perc_A, perc_B=perc_B, perc_C=perc_C, perc_D=perc_D)

        for i, (index, record) in enumerate(self.first_df.iterrows()):
            note_admis = get_address_of_cell(record[_("Passed")])
            note_agregee = get_address_of_cell(record[_("Aggregated grade")])
            opts = dict(note_admis=note_admis, note_agregee=note_agregee)
            opts.update(percs)

            # Fonction pour la construction des IF imbriqués
            def switch(ifs, default):
                if ifs:
                    cond, then = ifs[0]
                    return "IF({cond}, {then}, {else_})".format(
                        cond=cond.format(**opts),
                        then=then,
                        else_=switch(ifs[1:], default),
                    )
                else:
                    return default

            # Formule de type switch/case
            formula = switch(
                (
                    ("ISNA({note_admis})", "NA()"),
                    ('{note_agregee}="RESERVE"', '"RESERVE"'),
                    ('{note_agregee}="ABS"', '"ABS"'),
                    ("{note_admis}=0", '"F"'),
                    ("{note_agregee}>={perc_A}", '"A"'),
                    ("{note_agregee}>={perc_B}", '"B"'),
                    ("{note_agregee}>={perc_C}", '"C"'),
                    ("{note_agregee}>={perc_D}", '"D"'),
                ),
                '"E"',
            )

            record[_("ECTS grade")].value = "=" + formula

        # On centre les notes ECTS
        for cell in self.first_df[_("ECTS grade")]:
            cell.alignment = Alignment(horizontal="center")

        # On colore la cellule en fonction de la note ECTS
        range = self.get_column_range(_("ECTS grade"))
        for ects, color in zip(
            "ABCDEF", ["00FF00", "C2FF00", "F7FF00", "FFC100", "FF6900", "FF0000"]
        ):
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            rule = CellIsRule(operator="equal", formula=[f'"{ects}"'], fill=fill)
            self.first_ws.conditional_formatting.add(range, rule)

        # Formatage conditionnel pour les notes éliminatoires
        red_fill = PatternFill(
            start_color="EE1111", end_color="EE1111", fill_type="solid"
        )

        for name, opts in self.grades_options.items():
            threshold_cell = opts["passing grade"]
            threshold_addr = get_address_of_cell(threshold_cell, compat=True)
            self.first_ws.conditional_formatting.add(
                self.get_column_range(name),
                CellIsRule(
                    operator="lessThan", formula=[threshold_addr], fill=red_fill
                ),
            )

        # On offre la possibilité de trier
        max_column = self.first_ws.max_column
        max_row = self.first_ws.max_row
        self.first_ws.auto_filter.ref = "A1:{}{}".format(
            get_column_letter(max_column), max_row
        )
        range = self.get_column_range(_("ECTS grade"))
        self.first_ws.auto_filter.add_sort_condition(range)
