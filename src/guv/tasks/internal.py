import logging
import os

import pandas as pd
from doit.exceptions import TaskFailed
from doit.tools import check_timestamp_unchanged, config_changed
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from .. import openpyxl_patched
from ..config import settings
from ..exceptions import ImproperlyConfigured
from ..logger import logger
from ..translations import TaskDocstring, _
from ..utils import pformat
from ..utils_config import Output, selected_uv
from .base import UVTask


__all__ = ["XlsStudentData"]


def split_list_by_token_inclusive(lst):
    """Split a list of objects at locations where cache attribute is True"""

    result = []
    current = []
    for item in lst:
        current.append(item)
        if item.cache:
            result.append(current)
            current = []
    if current:
        result.append(current)
    return result


class check_file_and_config_unchanged:
    """Custom doit check

    Return True (up-to-date) if file hasn't changed and it it has, check that
    value has.

    """

    def __init__(self, filename, value):
        self.check_timestamp_unchanged = check_timestamp_unchanged(filename)
        self.config_changed = config_changed(value)

    def configure_task(self, task):
        task.value_savers.append(lambda: {'_config_changed': self.config_changed.config_digest})

    def __call__(self, task, values):
        try:
            res = self.check_timestamp_unchanged(task, values)
        except FileNotFoundError:
            return False
        return res or self.config_changed(task, values)


class Documents:
    """Class recording operations done to central file"""

    target_dir = "generated"
    target_name = "student_data_{step}.csv"

    def __init__(self):
        self.uv = None
        self._actions = []

    @classmethod
    def target_from(cls, **kwargs):
        target = os.path.join(
            settings.SEMESTER_DIR,
            kwargs["uv"],
            cls.target_dir,
            cls.target_name
        )
        return pformat(target, step=kwargs["step"])

    def setup(self, settings, info):
        for action in self.actions:
            action.setup(settings=settings, info=info)
        self.uv = info["uv"]

    def generate_doit_tasks(self):
        steps = split_list_by_token_inclusive(self.actions)
        for i, lst in enumerate(steps):
            step = i if i < len(steps) - 1 else "final"
            target = self.target_from(step=step, uv=self.uv)
            cache_file = self.target_from(step=i-1, uv=self.uv) if i > 0 else None

            other_deps = [d for a in lst for d in a.deps]
            deps = other_deps if cache_file is None else [cache_file] + other_deps

            def build_action(lst, cache_file, target):
                def func():
                    df = pd.read_csv(cache_file) if cache_file is not None else None
                    for a in lst:
                        logger.info(a.message())
                        try:
                            df = a.apply(df)
                        except Exception as e:
                            if settings.DEBUG <= logging.DEBUG:
                                raise e from e
                            return TaskFailed(_("The step `{name}` failed: {e}").format(name=a.name(), e=str(e)))

                    df.to_csv(target, index=False)
                return func

            value = "-".join(op.hash() for op in lst)
            config_file = os.path.join(settings.SEMESTER_DIR, self.uv, "config.py")

            doit_task = {
                "basename": f"DOCS_{i}",
                "actions": [build_action(lst, cache_file, target)],
                "file_dep": deps,
                "targets": [target],
                "uptodate": [check_file_and_config_unchanged(config_file, value)],
                "verbosity": 2
            }

            def format_task(doit_task):
                return "\n".join(f"{key}: {value}" for key, value in doit_task.items()
                                 if key not in ["doc"])

            logger.debug("Task properties are:")
            logger.debug(format_task(doit_task))

            yield doit_task

    def add_action(self, action):
        self._actions.append(action)

    @property
    def actions(self):
        return self._actions


class XlsStudentData(UVTask):
    __doc__ = TaskDocstring()

    target_name = "effectif.xlsx"
    target_dir = "."
    unique_uv = False

    @classmethod
    def create_doit_tasks_aux(cls):
        """Overriding UVTask to also generate tasks from DOCS."""
        tasks = []
        generators = []
        for planning, uv, info in selected_uv():
            instance = cls(planning, uv, info)

            if "DOCS" not in instance.settings:
                continue

            docs = instance.settings.DOCS
            if not isinstance(docs, Documents):
                raise ImproperlyConfigured(_("The DOCS variable must be of type `Documents`: `DOCS = Documents()`"))

            docs.setup(settings=instance.settings, info=info)

            task = instance.to_doit_task(
                name=f"{instance.planning}_{instance.uv}",
                uv=instance.uv
            )

            # Remove dependency to student_data_final.csv
            if len(docs.actions) == 0:
                task["file_dep"] = []

            tasks.append(task)
            generators.append(docs.generate_doit_tasks())

        return (item for gen in [*generators, tasks] for item in gen)

    def setup(self):
        super().setup()
        self.student_data = Documents.target_from(uv=self.info["uv"], step="final")
        self.file_dep = [self.student_data]
        self.target = self.build_target()

    def get_column_dimensions(self):
        if not os.path.exists(self.target):
            return {}

        def column_dimensions(ws):
            max_column = ws.max_column
            for i in range(1, max_column+1):
                colname = ws.cell(row=1, column=i).value
                width = ws.column_dimensions[get_column_letter(i)].width
                yield colname, width

        wb = load_workbook(self.target)
        ws = wb.active
        return {colname: width for colname, width in column_dimensions(ws)}

    def run(self):
        if "DOCS" not in self.settings:
            raise ImproperlyConfigured(_("The `config.py` file must contain a `DOCS` variable"))

        if not isinstance(self.settings.DOCS, Documents):
            raise ImproperlyConfigured(_("The `DOCS` variable must be of type `Documents`"))

        if len(self.settings.DOCS.actions) == 0:
            logger.warning(_("`DOCS` does not contain any operation"))
            return

        df = pd.read_csv(self.student_data)

        # Write set of columns for completion
        fp = os.path.join(self.settings.SEMESTER_DIR, self.uv, "generated", ".columns.list")
        with open(fp, "w") as file:
            file.write("\n".join(f"{e}" for e in df.columns.values))

        # Keep dataframe ordered the same as original effectif.xlsx
        # if os.path.exists(self.target):
        #     df_ordered = XlsStudentData.read_target(self.target)
        #     if len(df_ordered.index) == len(df.index):
        #         for colname in ["Login", "Courriel", "Adresse de courriel"]:
        #             if colname in df.columns and colname in df_ordered.columns:
        #                 if set(df[colname]) == set(df_ordered[colname]):
        #                     df = df.set_index("Login", drop=False).loc[df_ordered["Login"]].reset_index(drop=True)
        #                 break

        # Get column dimensions of original effectif.xlsx
        column_dimensions = self.get_column_dimensions()

        wb = Workbook()
        ws = wb.active

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        for cell in ws[1]:
            cell.style = 'Pandas'

        max_column = ws.max_column
        max_row = ws.max_row
        ws.auto_filter.ref = 'A1:{}{}'.format(
            get_column_letter(max_column),
            max_row)

        # Freeze header row and first 1–2 columns only if they contain NAME /
        # LASTNAME columns
        offset = 0
        if df.columns[0] in [self.settings.NAME_COLUMN, self.settings.LASTNAME_COLUMN]:
            offset += 1
        if df.columns[1] in [self.settings.NAME_COLUMN, self.settings.LASTNAME_COLUMN]:
            offset += 1

        if offset > 0:
            ws.freeze_panes = f"{get_column_letter(offset + 1)}2"

        # On redimensionne les colonnes d'après la taille précédente
        # ou la taille de l'en-tête
        for cell in ws[1]:
            width = None
            header_value = str(cell.value)

            if header_value in column_dimensions:
                width = column_dimensions[header_value]
            elif header_value == self.settings.NAME_COLUMN:
                width = 1.3 * 16
            elif header_value == self.settings.LASTNAME_COLUMN:
                width = 1.3 * 16
            elif header_value:
                width = 1.3 * max(len(header_value), 4)

            if width is not None:
                ws.column_dimensions[cell.column_letter].width = width

        with Output(self.target) as out:
            wb.save(out.target)

        target = os.path.splitext(self.target)[0] + ".csv"
        with Output(target) as out:
            df.to_csv(out.target, index=False)

    @staticmethod
    def read_target(student_data):
        return pd.read_excel(student_data, engine="openpyxl")
