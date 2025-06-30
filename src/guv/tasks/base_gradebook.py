import shlex
import sys

import openpyxl
import pandas as pd

from ..logger import logger
from ..openpyxl_patched import fixit
from ..utils import normalize_string, smart_cast
from ..utils_config import Output, rel_to_dir
from ..translations import _, _file
from .base import CliArgsInheritMixin, UVTask
from .internal import XlsStudentData

fixit(openpyxl)

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ..openpyxl_utils import fit_cells_at_col, get_range_from_cells


class AbstractGradeBook(UVTask, CliArgsInheritMixin):
    """Abstract UVTask that factor out all common gradebook logic"""

    uptodate = False
    target_dir = "generated"
    target_name = "{name}_gradebook.xlsx"

    # Flag to identify subclasses of AbstractGradeBook: module loading
    # issue if using issubclass in Sphinx's conf.py.
    doc_flag = True

    def get_columns(self):
        """Renvoie les colonnes utilisées pour créer la feuille de calcul.

        Il faut renvoyer une liste d'éléments de la forme (NOM, TYPE,
        PRIORITY). NOM est le nom d'une colonne qui sera utilisée. Il
        y a 4 types de colonnes :
        - `raw` : la colonne est copiée du fichier `effectif.xlsx` si
          elle existe ou alors juste créée et laissée vide.
        - `cell` : même chose que `raw`. Les cellules de la colonnes
          sont stockées de manière à pouvoir être réutilisée dans des
          formules.
        - `grade`: même chose que `cell`. La colonne est en plus
          enregistrée comme contenant ou pouvant contenir une note
          (utilisé dans la classe `XlsGradeBookJury`).
        - `hide`: même chose que `raw`. La colonne est disponible pour
          construire la ou les feuilles de calcul mais ne figurera pas
          dans la feuille de classeur finale.

        """

        # Default columns
        columns = [
            (self.settings.NAME_COLUMN, "raw", 0),
            (self.settings.LASTNAME_COLUMN, "raw", 0),
            (self.settings.EMAIL_COLUMN, "raw", 0),
        ]

        return columns

    def get_column_range(self, colname):
        "Renvoie la plage de cellule de la colonne COLNAME sans l'en-tête."

        if colname not in self.first_df.columns:
            raise ValueError("Unknown column name: {}".format(colname))

        cells = self.first_df[colname]
        first, last = cells.iloc[0], cells.iloc[-1]

        return get_range_from_cells(first, last)

    def setup(self):
        super().setup()
        self.xls_merge = XlsStudentData.target_from(**self.info)
        self.file_dep = [self.xls_merge]
        self.parse_args()

        # No targets to avoid circular deps in doit as we probably
        # want to aggregate target in effectif.xlsx
        self.targets = []

    def run(self):
        self.data_df = XlsStudentData.read_target(self.xls_merge)
        self.create_first_worksheet()
        self.create_other_worksheets()
        target = self.build_target(name=normalize_string(self.name, type="file"))
        with Output(target, protected=True) as out:
            self.workbook.save(out.target)

        logger.info(self.message(target))

    def message(self, target):
        if isinstance(self.agg_colname, list):
            columns = "[" + ", ".join(f'"{c}"' for c in self.agg_colname) + "]"
        else:
            columns = f'"{self.agg_colname}"'

        return _file("XlsGradeBook_message").format(
            filename=rel_to_dir(target, self.settings.UV_DIR),
            columns=columns,
            email=self.settings.EMAIL_COLUMN,
            command_line="guv " + " ".join(map(shlex.quote, sys.argv[1:]))
        )

    def get_sorted_columns(self):
        """Return list of columns sorted by priority."""

        columns = self.get_columns()

        # Sort columns according to priority and index in columns
        columns = [(i, *rest) for i, rest in enumerate(columns)]

        def sort_func(elt):
            i, _, _, priority = elt
            return priority, i

        columns = sorted(columns, key=sort_func)
        columns = [(name, type) for i, name, type, priority in columns]

        return columns

    def create_first_worksheet(self):
        # Create workbook and first worksheet named "data"
        self.workbook = Workbook()
        self.first_ws = self.workbook.active
        self.first_ws.title = "data"

        # Pandas dataframe that mirrors the first worksheet
        self.first_df = pd.DataFrame()

        # Number of students
        N = len(self.data_df.index)

        # Get columns to be eventually copied from source of
        # information DATA_DF to first worksheet
        columns = self.get_sorted_columns()

        # Write `first_ws` and `first_df` attributes column by column from
        # data in `data_df`
        idx = 1
        for name, type in columns:
            # Update first_ws
            if type != "hide":
                # Write header of column with Pandas style
                self.first_ws.cell(1, idx).value = name
                self.first_ws.cell(1, idx).style = "Pandas"

                # Copy data from `data_df` if existing into first
                # worksheet
                if name in self.data_df.columns:
                    for i, value in enumerate(self.data_df[name]):
                        # Try to convert to number from string to avoid leading
                        # quote in Excel/LibreOffice
                        value = smart_cast(value)
                        self.first_ws.cell(i + 2, idx).value = value

                # Get cells
                cells = self.first_ws[get_column_letter(idx)][1 : (N + 1)]

                # Fit width of column to actual content
                fit_cells_at_col(self.first_ws.cell(1, idx), *cells)

                # Next column index
                idx += 1

            else:
                # Non-existent column
                cells = None

            # Update first_df
            if name in self.data_df.columns:  # If column exists
                if type in ["hide", "raw"]:
                    self.first_df[name] = self.data_df[name]
                elif type in ["grade", "cell"]:
                    if cells is None:
                        raise RuntimeError("Logical error")
                    self.first_df[name] = cells
                else:
                    raise ValueError("Unknown type of column ", type)
            else:
                if type in ["grade", "cell"]:
                    self.first_df[name] = cells
                elif type in ["hide", "raw"]:
                    logger.warning(_("The column `%s` does not exist in the central file, it will be empty in the created file"), name)
                    self.first_df[name] = pd.Series(dtype=str)
                else:
                    raise ValueError("Unknown type of column ", type)

        # Sort all columns
        self.first_ws.auto_filter.ref = "A1:{}{}".format(
            get_column_letter(idx - 1), N + 1
        )

        # Freeze header row and first two columns
        self.first_ws.freeze_panes = "C2"

    def create_other_worksheets(self):
        raise NotImplementedError
