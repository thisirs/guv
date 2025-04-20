"""
Ce module rassemble les tâches de gestion des intervenants au sein
d'une UV : fichier Excel du nombre d'heures théoriques, décompte des
heures remplacées.
"""

import os

import numpy as np
import openpyxl
import pandas as pd
from pandas.api.types import CategoricalDtype

from ..openpyxl_patched import fixit

fixit(openpyxl)

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from ..logger import logger
from ..openpyxl_utils import (Block, fill_row, frame_range,
                              get_range_from_cells, get_row_cells, get_segment,
                              row_and_col)
from ..utils import convert_to_time, score_codenames
from ..utils_config import Output, rel_to_dir
from .base import SemesterTask, UVTask
from .internal import PlanningSlots, WeekSlots


class XlsInstructors(SemesterTask):
    """Fichier de détails global des intervenants toutes UV confondues.

    Il sert à la tâche :class:`~guv.tasks.moodle.HtmlInst` pour
    générer un descriptif des intervenants d'une UV.
    """

    hidden = True
    target_dir = "documents"
    target_name = "intervenants.xlsx"

    def setup(self):
        super().setup()
        self.target = self.build_target()

    def run(self):
        if not os.path.exists(self.target):
            logger.warning(
                "Le fichier `%s` n'existe pas, création d'un fichier vide",
                rel_to_dir(self.target),
            )
            columns = ["Intervenants", "Statut", "Email", "Website"]
            with Output(self.target) as out:
                pd.DataFrame(columns=columns).to_excel(out.target, index=False)

    @staticmethod
    def read_target(insts):
        return pd.read_excel(insts, engine="openpyxl")


class WeekSlotsDetails(UVTask):
    """Fichier Excel des intervenants par UV avec détails.

    Les détails sont pris dans le fichier de détails global. Les
    affectations sont prises pour chaque UV.
    """

    hidden = True
    target_dir = "generated"
    target_name = "intervenants_details.xlsx"

    def setup(self):
        super().setup()
        self.target = self.build_target()

        self.insts = XlsInstructors.target_from()
        self.week_slots = WeekSlots.target_from(**self.info)
        self.file_dep = [self.week_slots, self.insts]

    def run(self):
        week_slots = WeekSlots.read_target(self.week_slots)

        if week_slots["Intervenants"].isnull().any():
            logger.warning("Certains créneaux n'ont pas d'intervenant renseigné")

        insts = XlsInstructors.read_target(self.insts)

        # Add details from inst_details
        df_outer = week_slots.merge(
            insts,
            how="outer",
            on="Intervenants",
            indicator=True,
        )

        df_left = df_outer[df_outer["_merge"].isin(["left_only", "both"])]
        df_left = df_left.drop(["_merge"], axis=1)

        df_missing = df_outer[df_outer["_merge"].isin(["left_only"])]
        inst_missing = df_missing["Intervenants"].dropna().unique()
        for inst in inst_missing:
            logger.warning("Pas d'informations détaillées sur l'intervenant: %s", inst)

        with Output(self.target) as out:
            df_left.to_excel(out.target, index=False)

    @staticmethod
    def read_target(week_slots_details):
        sts = ["MCF", "PR", "PRAG", "PRCE", "PAST", "ECC", "Doct", "ATER", "Vacataire"]
        status_type = CategoricalDtype(categories=sts, ordered=True)

        df = pd.read_excel(week_slots_details, engine="openpyxl")

        df["Statut"] = df["Statut"].astype(status_type)
        df["Email"] = df["Email"].fillna("").astype("string")
        df["Locaux"] = df["Locaux"].fillna("")
        df["Heure début"] = df["Heure début"].apply(convert_to_time)
        df["Heure fin"] = df["Heure fin"].apply(convert_to_time)

        return df


class XlsRemplacements(UVTask):
    """Crée un fichier Excel pour enregistrer les remplacements.

    La première feuille garde la trace des séances qui ont été remplacées. La
    deuxième feuille permet calculer le total des UTP.

    """

    target_dir = "documents"
    target_name = "remplacements.xlsx"
    uptodate = False

    def setup(self):
        super().setup()

        self.target = self.build_target()
        self.planning_slots = PlanningSlots.target_from(**self.info)
        self.week_slots_details = WeekSlotsDetails.target_from(**self.info)
        self.file_dep = [self.week_slots_details, self.planning_slots]

    def run(self):
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "remplacements"

        self.df_week_slots = WeekSlotsDetails.read_target(self.week_slots_details)
        if self.df_week_slots["Intervenants"].isnull().any():
            week_slots = WeekSlots.target_from(**self.info)
            fn = rel_to_dir(week_slots)
            raise Exception(
                f"La colonne `Intervenants` du fichier `{fn}` n'est pas "
                "totalement remplie."
            )

        ref = self.worksheet.cell(1, 1)
        context = self.write_slots(ref)

        ref = self.worksheet.cell(1, 10)
        self.write_table(ref, context)

        target = self.build_target()
        with Output(target, protected=True) as out:
            self.workbook.save(out.target)

    def write_slots(self, ref):
        activities = self.df_week_slots["Activité"].unique()

        inst_dict = {a: [] for a in activities}
        rep_dict = {a: [] for a in activities}

        df_planning_slots = PlanningSlots.read_target(self.planning_slots)

        for i, (activity, group) in enumerate(df_planning_slots.groupby("Activité")):
            if i == 0:
                block = self.write_activity_header(activity, ref)
            else:
                block = self.write_activity_header(activity, block.lower_left.below(3))

            for num, subgroup in group.groupby("num"):
                sorted_subgroup = subgroup.sort_values("Lib. créneau")
                block = self.write_week_header(num, block.lower_left.below())
                block = self.write_week(sorted_subgroup, block.lower_left.below())
                inst_dict[activity].append(block.content.inst)
                rep_dict[activity].append(block.content.rep)

        return inst_dict, rep_dict

    def write_activity_header(self, activity, ref):
        ref.text(activity).merge(ref.right(5)).style = "Pandas"
        return Block(ref, ref.right(5))

    def write_week_header(self, num, ref):
        weekAB = "A" if num % 2 == 1 else "B"
        numAB = (num + 1) // 2
        week_name = f"Semaine {weekAB}{numAB}"
        ref.text(week_name).merge(ref.right(5)).style = "Pandas"
        return Block(ref, ref.right(5))

    def write_week(self, group, ref):
        current = ref
        for index, row in group.iterrows():
            instructor = (
                current.text(row["Lib. créneau"])
                .right()
                .text(row["Jour"])
                .right()
                .text(row["Heure début"])
                .right()
                .text(row["Heure fin"])
                .right()
                .text(row["Abbrev"])
            )
            current = current.below()

        grey = PatternFill(fill_type="solid", start_color="AAAAAA", end_color="AAAAAA")
        for cell in get_segment(row_and_col(ref, instructor.right()), instructor.right()):
            cell.fill = grey

        block = Block(
            ref,
            instructor.right(),
            inst=get_range_from_cells(
                row_and_col(ref, instructor), instructor
            ),
            rep=get_range_from_cells(
                row_and_col(ref, instructor.right()), instructor.right()
            ),
        )
        frame_range(block.upper_left, block.lower_right)
        return block

    def write_table(self, ref, context):
        inst_dict, rep_dict = context

        df_ws = self.df_week_slots
        semAB = not df_ws.loc[df_ws["Activité"] == "TP", "Semaine"].isna().all()

        header = [
            "Intervenant",
            "Abbrev",
            "Statut",
            "Cours",
            "TD",
            "TP",
            "Balance Cours",
            "Balance TD",
            "Balance TP",
            "UTP prévi",
            "UTP final",
        ]

        if "TP" not in inst_dict:
            del header[header.index("Statut")]
            del header[header.index("TP")]
            del header[header.index("Balance TP")]

        # Write header
        row_cells = get_row_cells(ref, 0, *header)
        fill_row(row_cells, **{e: e for e in header})
        for cell in row_cells.values():
            cell.style = "Pandas"

        # Compute dataframe with columns "Intervenants", "Abbrev",
        # "Cours", "TD", ("TP")
        index_cols = ["Intervenants", "Abbrev"]
        df1 = df_ws.groupby(index_cols).agg(
            Order=("Activité", score_codenames),
            Statut=("Statut", "first")
        )
        df2 = (
            df_ws.groupby(index_cols + ["Activité"])
            .agg(Nombre=("Jour", "count"))
            .reset_index()
            .pivot(
                index=index_cols,
                columns="Activité",
                values="Nombre",
            )
        )
        df = pd.concat((df1, df2), axis=1).reset_index()

        def get_formula(tmpl, activity, inst_dict, rep_dict, inst_cell):
            return "=" + "+".join(
                tmpl.format(
                    rep_range=rep_range,
                    inst_range=inst_range,
                    inst_id_cell=inst_cell.coordinate,
                )
                for inst_range, rep_range in zip(
                        inst_dict[activity], rep_dict[activity]
                )
            )

        tmpl = 'COUNTIF({rep_range}, {inst_id_cell}) - COUNTIFS({inst_range}, {inst_id_cell}, {rep_range}, "<>")'

        statut_mult = {
            "MCF": 1.5,
            "PR": 1.5,
            "PRAG": 1.5,
            "PRCE": 1.5,
            "PAST": 1.5,
            "ECC": 1.5,
            "Doct": 1.5,
            "ATER": 1,
            "Vacataire": 1,
            np.nan: 1,
        }

        for i, (index, data) in enumerate(df.sort_values("Order", ascending=False).iterrows()):
            row_cells = get_row_cells(ref, i+1, *header)

            elts = {
                "Intervenant": data["Intervenants"],
                "Abbrev": data["Abbrev"],
                "Cours": data["Cours"],
                "TD": data["TD"],
                "Balance Cours": lambda row: get_formula(tmpl, "Cours", inst_dict, rep_dict, row["Abbrev"]),
                "Balance TD": lambda row: get_formula(tmpl, "TD", inst_dict, rep_dict, row["Abbrev"]),
            }

            if "TP" in inst_dict:
                elts["Statut"] = statut_mult[data["Statut"]]
                elts["TP"] = data["TP"]
                elts["Balance TP"] = lambda row: get_formula(tmpl, "TP", inst_dict, rep_dict, row["Abbrev"])
                elts["UTP prévi"] = lambda row: "=2*16*2.25*{cours_cell}+2*16*1.5*{TD_cell}+2*{num_week}*{TP_cell}*{status_cell}".format(
                    cours_cell=row["Cours"].coordinate,
                    TD_cell=row["TD"].coordinate,
                    num_week="8" if semAB else "16",
                    TP_cell=row["TP"].coordinate,
                    status_cell=row["Statut"].coordinate,
                )
                elts["UTP final"] = lambda row: "=2*2.25*(16*{cours_cell}+{cours_bal})+2*1.5*(16*{TD_cell}+{TD_bal})+2*({num_week}*{TP_cell}+{TP_bal})*{status_cell}".format(
                    cours_cell=row["Cours"].coordinate,
                    cours_bal=row["Balance Cours"].coordinate,
                    TD_cell=row["TD"].coordinate,
                    TD_bal=row["Balance TD"].coordinate,
                    num_week="8" if semAB else "16",
                    TP_cell=row["TP"].coordinate,
                    TP_bal=row["Balance TP"].coordinate,
                    status_cell=row["Statut"].coordinate,
                )
            else:
                elts["UTP prévi"] = lambda row: "=2*16*2.25*{cours_cell}+2*16*1.5*{TD_cell}".format(
                    cours_cell=row["Cours"].coordinate,
                    TD_cell=row["TD"].coordinate,
                )
                elts["UTP final"] = lambda row: "=2*2.25*(16*{cours_cell}+{cours_bal})+2*1.5*(16*{TD_cell}+{TD_bal})".format(
                    cours_cell=row["Cours"].coordinate,
                    cours_bal=row["Balance Cours"].coordinate,
                    TD_cell=row["TD"].coordinate,
                    TD_bal=row["Balance TD"].coordinate,
                )

            fill_row(row_cells, **elts)

        frame_range(ref, row_cells[header[-1]])


class XlsUTP(SemesterTask):
    """Crée un fichier Excel de prévisions des UTP globales."""

    target_dir = "."
    target_name = "UTP.xlsx"

    def setup(self):
        super().setup()
        self.target = self.build_target()

    def write_UV_block(self, ref_cell):
        """Write table at `ref_cell`."""

        header = [
            "Libellé",
            "Type (CM/TD/TP)",
            "Nombre de créneaux de 2h par semaine",
            "Heures",
            "UTP",
            "Heures éq. TD",
            "",
            "CM",
            "TD",
            "TP",
            "",
            "Heures éq. TD CM",
            "Heures éq. TD TD",
            "Heures éq. TD TP",
        ]

        def switch(ifs, default):
            """Nested if conditionals with a default value."""

            if ifs:
                cond, then = ifs[0]
                return "IF({cond}, {then}, {else_})".format(
                    cond=cond,
                    then=then,
                    else_=switch(ifs[1:], default),
                )
            else:
                return default

        formula = switch(
            (
                ("{type_cell} <> {type}", '""'),
                ("NOT(ISBLANK({utp}))", "{utp_result}"),
                ("NOT(ISBLANK({h_eq_tp}))", "{h_eq_tp_result}"),
                ("NOT(ISBLANK({heures}))", "{mult} * {heures}"),
                ("NOT(ISBLANK({slots_2h}))", "2*16*{slots_2h}*{mult}"),
            ),
            ""
        )

        def get_formula(formula, metric, ctype, mult):
            def inner(row):
                return "=" + formula.format(
                    slots_2h=row["Nombre de créneaux de 2h par semaine"].coordinate,
                    heures=row["Heures"].coordinate,
                    utp=row["UTP"].coordinate,
                    h_eq_tp=row["Heures éq. TD"].coordinate,
                    utp_result=row["UTP"].coordinate if metric == "UTP" else "2/3*{}".format(row["UTP"].coordinate),
                    h_eq_tp_result=row["Heures éq. TD"].coordinate if metric == "eqTD" else "3/2*{}".format(row["Heures éq. TD"].coordinate),
                    type=f'"{ctype}"',
                    type_cell=row["Type (CM/TD/TP)"].coordinate,
                    mult=mult
                )
            return inner

        elts = {
            colname: get_formula(formula, metric, ctype, mult)
            for metric, colname, ctype, mult, in (
                    ("UTP", "CM", "CM", 2.25),
                    ("UTP", "TD", "TD", 1.5),
                    ("UTP", "TP", "TP", 1.5),
                    ("eqTD", "Heures éq. TD CM", "CM", 1.5),
                    ("eqTD", "Heures éq. TD TD", "TD", 1),
                    ("eqTD", "Heures éq. TD TP", "TP", 1),
            )
        }

        # Write header at `ref_cell`
        row_cells = get_row_cells(ref_cell, 0, *header)
        fill_row(row_cells, **{e: e for e in header})

        # Write header above
        row_cells2 = get_row_cells(ref_cell, -1, *header)

        row_cells2["Nombre de créneaux de 2h par semaine"].merge(row_cells2["Heures éq. TD"]).center().value = "Charge effectuée"
        frame_range(row_cells2["Nombre de créneaux de 2h par semaine"], row_cells["Heures éq. TD"])

        row_cells2["CM"].merge(row_cells2["TP"]).center().value = "UTP"
        frame_range(row_cells2["CM"], row_cells["TP"])

        row_cells2["Heures éq. TD CM"].merge(row_cells2["Heures éq. TD TP"]).center().value = "Heures"
        frame_range(row_cells2["Heures éq. TD CM"], row_cells["Heures éq. TD TP"])

        # Write formulas in rows
        n_row = 30
        for i in range(n_row):
            row_cells = get_row_cells(ref_cell, i + 1, *header)
            fill_row(row_cells, **elts)

        # Columns on which to make a total
        kw_totals = [
            "CM",
            "TD",
            "TP",
            "Heures éq. TD CM",
            "Heures éq. TD TD",
            "Heures éq. TD TP",
        ]
        first_row = get_row_cells(ref_cell, 1, *header)
        last_row = get_row_cells(ref_cell, n_row, *header)

        # Subtotals
        for kw in kw_totals:
            range_cells = get_range_from_cells(first_row[kw], last_row[kw])
            total_cell = last_row[kw].below()
            total_cell.value = "=SUM({})".format(range_cells)

        # Overall totals
        range_cells = get_range_from_cells(last_row["CM"].below(), last_row["TP"].below())
        last_row["TP"].below().right().value = "=SUM({})".format(range_cells)

        range_cells = get_range_from_cells(last_row["Heures éq. TD CM"].below(), last_row["Heures éq. TD TP"].below())
        last_row["Heures éq. TD TP"].below().right().value = "=SUM({})".format(range_cells)

    def run(self):
        wb = Workbook()
        ws = wb.active

        ref_cell = ws.cell(3, 2)
        self.write_UV_block(ref_cell)

        with Output(self.target, protected=True) as out:
            wb.save(out.target)
