"""
Ce module rassemble les tâches de création des créneaux officiels de
Cours/TD/TP.
"""

import os
import re

import numpy as np
import openpyxl
import pandas as pd
from PyPDF2 import PdfReader
from tabula import read_pdf

from ..openpyxl_patched import fixit

fixit(openpyxl)

from openpyxl import Workbook, load_workbook

from ..config import settings
from ..exceptions import ImproperlyConfigured
from ..logger import logger
from ..openpyxl_utils import fill_row, get_row_cells, get_range_from_cells, get_segment, frame_range, row_and_col, Block, get_range_cells
from ..utils import convert_author, score_codenames, convert_to_time, plural, ps, px
from ..utils_config import (Output, ask_choice, generate_row, rel_to_dir, selected_uv)
from .base import TaskBase, UVTask


class UtcUvListToCsv(TaskBase):
    """Crée un fichier CSV des créneaux de toutes les UVs à partir du PDF"""

    hidden = True
    target_dir = "documents"
    target_name = "UTC_UV_list.csv"

    def setup(self):
        super().setup()

        self.target = self.build_target()

        if "CRENEAU_UV" in self.settings:
            self.uv_list_filename = os.path.join(
                self.settings.SEMESTER_DIR,
                self.settings.CRENEAU_UV
            )
        else:
            self.uv_list_filename = None

        if self.uv_list_filename is not None:
            self.file_dep = [self.uv_list_filename]
        else:
            self.file_dep = []

    def read_pdf(self):
        pdf = PdfReader(open(self.uv_list_filename, 'rb'))
        npages = len(pdf.pages)

        possible_cols = ['Code enseig.', 'Activité', 'Jour', 'Heure début',
                         'Heure fin', 'Semaine', 'Locaux', 'Type créneau',
                         'Lib. créneau', 'Responsable enseig.']

        tables = []
        pdo = {"header": None}
        for i in range(npages):
            logger.info("Processing page (%d/%d)", i+1, npages)
            page = i + 1
            tabula_args = {'pages': page}
            # Use pdo.copy(): pdo is changed by read_pdf
            df = read_pdf(self.uv_list_filename, **tabula_args, pandas_options=pdo.copy())[0]

            if page == 1:
                # Detect header (might be a two-line header)
                header_height = ((re.match('[A-Z]{,3}[0-9]+', str(df.iloc[0, 0])) is None) +
                                 (re.match('[A-Z]{,3}[0-9]+', str(df.iloc[1, 0])) is None))
                if header_height == 0:
                    raise Exception("No header detected")
                logger.info("Detected header has %d lines", header_height)

                # Compute single line/multiline header
                header = df.iloc[:header_height].fillna('').agg(['sum']).iloc[0]
                logger.info("Header is: %s", ' '.join(header))

                # Extract real data
                df = df.iloc[header_height:]

                # Set name of columns from header
                df = df.rename(columns=header)

                # Rename incorrectly parsed headers
                df = df.rename(columns={
                    'Activit': 'Activité',
                    'Type cr neau': 'Type créneau',
                    'Lib. cr': 'Lib. créneau',
                    'Lib.créneau': 'Lib. créneau',
                    'Lib.': 'Lib. créneau',
                    'Lib.\rcréneau': 'Lib. créneau',
                    'Heuredébut': 'Heure début',
                    'Heure d': 'Heure début',
                    'Heurefin': 'Heure fin',
                    'Locaux hybrides': "Locaux"
                })

                unknown_cols = list(set(df.columns) - set(possible_cols))
                if unknown_cols:
                    raise Exception("Colonnes inconnues détectées:", " ".join(unknown_cols))

                # Get list of detected columns
                cols = df.columns.to_list()

                # 'Semaine' is the only column that might not be
                # detected in next pages because it can be empty.
                # Store its index to insert a blank column if needed
                if "Semaine" in cols:
                    week_idx = cols.index('Semaine')
                else:
                    week_idx = cols.index('Heure fin') + 1

                logger.info("%d columns found", len(df.columns))
                logger.info(" ".join(df.columns))
            else:
                logger.info("%d columns found", len(df.columns))

                # Semaine column might be empty and not detected
                if len(df.columns) == len(cols):
                    pass
                elif len(df.columns) == len(cols) - 1:
                    df.insert(week_idx, 'Semaine', np.nan)
                else:
                    raise Exception("Mauvais nombre de colonnes détectées")
                df.columns = cols

                # Detect possible multiline header
                header_height = ((re.match('[A-Z]{,3}[0-9]+', str(df.iloc[0, 0])) is None) +
                                 (re.match('[A-Z]{,3}[0-9]+', str(df.iloc[1, 0])) is None))
                logger.info("Header has %d lines", header_height)
                df = df.iloc[header_height:]

            tables.append(df)

        df = pd.concat(tables)
        useful_cols = [
            "Code enseig.",
            "Activité",
            "Jour",
            "Heure début",
            "Heure fin",
            "Semaine",
            "Locaux",
            "Lib. créneau",
        ]
        df = df[useful_cols]

        return df

    def run(self):
        if self.uv_list_filename is None:
            raise Exception("La variable `CRENEAU_UV` doit être définie")

        # Lire tous les créneaux par semaine de toutes les UVs
        if not os.path.exists(self.uv_list_filename):
            uv_fn = rel_to_dir(self.uv_list_filename, self.settings.SEMESTER_DIR)
            raise Exception(f"Le fichier n'existe pas: {uv_fn}")

        df = self.read_pdf()

        # Remove duplicate indexes from concat
        df.reset_index(drop=True, inplace=True)

        # T1 instead of T 1
        df['Lib. créneau'].replace(' +', '', regex=True, inplace=True)

        # A ou B au lieu de semaine A et semaine B
        df['Semaine'].replace("^semaine ([AB])$", "\\1", regex=True, inplace=True)

        # Semaine ni A ni B pour les TP: demander
        uvs = self.settings.UVS

        def fix_semaineAB(group):
            if group.name[1] == 'TP' and len(group.index) > 1:
                nans = group.loc[(pd.isnull(group['Semaine']))]
                if 0 < len(nans.index) or len(nans.index) < len(group.index):
                    if group.name[0] in uvs:
                        for index, row in nans.iterrows():
                            choice = ask_choice(
                                f'Semaine pour le créneau {row["Lib. créneau"]} de TP de {group.name[0]} (A ou B) ? ',
                                choices={"A": "A", "a": "A", "B": "B", "b": "B"},
                            )
                            group.loc[index, "Semaine"] = choice
                    else:
                        group.loc[nans.index, 'Semaine'] = 'A'
                return group
            else:
                return group

        df = df.groupby(['Code enseig.', 'Activité'], group_keys=False).apply(fix_semaineAB)

        with Output(self.target) as out:
            df.to_csv(out.target, index=False)


class WeekSlotsAll(TaskBase):
    """Rassemble les fichiers ``planning_hebdomadaire.xlsx`` de chaque UV/UE.

    Les colonnes sont :

    - Planning
    - Code enseig.
    - Jour
    - Heure début
    - Heure fin
    - Locaux
    - Semaine
    - Lib. créneau
    - Intervenants
    - Abbrev
    - Responsable

    """

    hidden = True
    target_dir = "generated"
    target_name = "planning_hebdomadaire.xlsx"

    def setup(self):
        super().setup()
        self.target = self.build_target()
        self.affectations = [
            (planning, uv, WeekSlots.target_from(**info))
            for planning, uv, info in selected_uv()
        ]
        self.file_dep = [f for _, _, f in self.affectations]

    def run(self):
        def func(planning, uv, xls_aff):
            df = WeekSlots.read_target(xls_aff)
            df.insert(0, "Code enseig.", uv)
            df.insert(0, "Planning", planning)
            return df

        df_affs = [func(planning, uv, xls_aff) for planning, uv, xls_aff in self.affectations]
        df_aff = pd.concat(df_affs, ignore_index=True)
        df_aff.Semaine = df_aff.Semaine.astype(object)

        with Output(self.target) as out:
            df_aff.to_excel(out.target, index=False)

    @staticmethod
    def read_target(week_slots_all):
        df = pd.read_excel(week_slots_all, engine="openpyxl")

        df["Heure début"] = df["Heure début"].apply(convert_to_time)
        df["Heure fin"] = df["Heure fin"].apply(convert_to_time)

        return df


class PlanningSlotsAll(TaskBase):
    """Rassemble les fichiers `plannings.xlsx` de chaque UE/UV."""

    hidden = True
    unique_uv = False
    target_dir = "generated"
    target_name = "planning_all.xlsx"

    def setup(self):
        super().setup()
        self.target = self.build_target()
        self.planning_slots_files = [
            (planning, uv, PlanningSlots.target_from(**info))
            for planning, uv, info in selected_uv()
        ]
        self.file_dep = [f for _, _, f in self.planning_slots_files]

    def run(self):
        def func(planning, uv, xls_aff):
            df = PlanningSlots.read_target(xls_aff)
            df.insert(0, "Code enseig.", uv)
            df.insert(0, "Planning", planning)
            return df

        dfs = [
            func(planning, uv, xls_aff)
            for planning, uv, xls_aff in self.planning_slots_files
        ]

        df = pd.concat(dfs, ignore_index=True)
        df.Semaine = df.Semaine.astype(object)

        with Output(self.target) as out:
            df.to_excel(out.target, index=False)

    @staticmethod
    def read_target(planning_slots_all):
        df = pd.read_excel(planning_slots_all, engine="openpyxl")

        df["date"] = pd.to_datetime(df["date"]).dt.date
        df["Heure début"] = df["Heure début"].apply(convert_to_time)
        df["Heure fin"] = df["Heure fin"].apply(convert_to_time)

        return df


class PlanningSlots(UVTask):
    """Fichier Excel des créneaux sur le planning entier.

    Les colonnes du fichier sont :

    - Code enseig.: SY02
    - Activité: TP
    - Jour: Lundi
    - Heure début: 14:15
    - Heure fin: 16:15
    - Semaine: B
    - Locaux: BF B 113
    - Lib. créneau: T1
    - Planning: P2021
    - Responsable:
    - Intervenants: Fisher
    - Responsable:
    - date: 2021-03-08
    - num: 1
    - numAB: 1
    - nweek: 4

"""

    hidden = True
    unique_uv = False
    target_name = "planning.xlsx"
    target_dir = "generated"

    def setup(self):
        super().setup()
        self.target = self.build_target()

        self.week_slots = WeekSlots.target_from(**self.info)
        self.file_dep = [self.week_slots]

        # Set uptodate value without raising Exception or displaying
        # warnings
        self.uptodate = {}
        try:
            props = self.settings.PLANNINGS[self.planning]
        except ImproperlyConfigured:
            props = {}

        for name in ["PL_BEG", "PL_END", "TURN", "SKIP_DAYS_C", "SKIP_DAYS_D", "SKIP_DAYS_T"]:
            try:
                if name in props:
                    value = props[name]
                else:
                    value = self.settings[name]
                self.uptodate[self.uv + "_" + name.lower()] = value
            except ImproperlyConfigured:
                self.uptodate[self.uv + "_" + name.lower()] = None

    def set_vars(self):
        props = self.settings.PLANNINGS[self.planning]

        for name in ["PL_BEG", "PL_END", "TURN", "SKIP_DAYS_C", "SKIP_DAYS_D", "SKIP_DAYS_T"]:
            if name not in props:
                logger.warning(
                    f"La clé `{name}` est absente du planning `{self.planning}` dans la "
                    f"variable `PLANNINGS`, utilisation de la variable globale `{name}`."
                )
                value = self.settings[name]
            else:
                value = props[name]

            setattr(self, name.lower(), value)

    def run(self):
        self.set_vars()

        df = WeekSlots.read_target(self.week_slots)

        mask_C = df["Activité"] == "Cours"
        mask_D = df["Activité"] == "TD"
        mask_T = df["Activité"] == "TP"

        df_C = df.loc[mask_C]
        df_D = df.loc[mask_D]
        df_T = df.loc[mask_T]

        df_other = df.loc[~(mask_C | mask_D | mask_T)]

        if len(df_other) > 0:
            for name, group in df_other.groupby("Activité"):
                logger.warning(
                    "%d créneau%s %s étiqueté%s `%s`, le%s compter comme `Cours`, `TD` ou `TP` ?",
                    len(df_other),
                    px(len(df_other)),
                    plural(len(df_other), "sont", "est"),
                    ps(len(df_other)),
                    name,
                    ps(len(df_other)),
                )
                result = ask_choice("Choix ? ", {"Cours": "Cours", "TD": "TD", "TP": "TP"})
                if result == "Cours":
                    df_C = pd.concat((df_C, df_other))
                elif result == "TD":
                    df_D = pd.concat((df_D, df_other))
                elif result == "TP":
                    df_T = pd.concat((df_T, df_other))
                else:
                    raise RuntimeError("Logical error")

        # DataFrame of days in planning
        planning_C = pd.DataFrame(
            generate_row(self.pl_beg, self.pl_end, self.skip_days_c, self.turn),
            columns=["date", "Jour", "num", "Semaine", "numAB", "nweek"],
        )
        planning_D = pd.DataFrame(
            generate_row(self.pl_beg, self.pl_end, self.skip_days_d, self.turn),
            columns=["date", "Jour", "num", "Semaine", "numAB", "nweek"],
        )
        planning_T = pd.DataFrame(
            generate_row(self.pl_beg, self.pl_end, self.skip_days_t, self.turn),
            columns=["date", "Jour", "num", "Semaine", "numAB", "nweek"],
        )

        for planning, text, number in (
                (planning_C, "cours", 14),
                (planning_D, "TD", 13),
                (planning_T, "TP", 14),
        ):
            counts = planning["Jour"].value_counts()
            unique = counts.unique()
            if len(unique) != 1:
                serie = ", ".join(f"{index} : {value}" for index, value in counts.items())
                logger.warning("Le nombre de créneaux de %s n'est pas le même pour tous les jours : %s", text, serie)
            elif unique.item() != number:
                logger.warning("Le nombre de créneaux de %s est différent de %d : %d", text, number, unique.item())

        planning_C = planning_C.drop("Semaine", axis=1)
        df_Cp = pd.merge(df_C, planning_C, how="left", on="Jour")

        planning_D = planning_D.drop("Semaine", axis=1)
        df_Dp = pd.merge(df_D, planning_D, how="left", on="Jour")

        if df_T["Semaine"].hasnans:
            planning_T = planning_T.drop("Semaine", axis=1)
            df_Tp = pd.merge(df_T, planning_T, how="left", on="Jour")
        else:
            df_Tp = pd.merge(
                df_T,
                planning_T,
                how="left",
                on=["Jour", "Semaine"],
            )

        dfp = pd.concat([df_Cp, df_Dp, df_Tp], ignore_index=True)

        with Output(self.target) as out:
            dfp.to_excel(out.target, index=False)

    @staticmethod
    def read_target(planning_slots):
        df = pd.read_excel(planning_slots, engine="openpyxl")

        df["date"] = pd.to_datetime(df["date"]).dt.date
        df["Heure début"] = df["Heure début"].apply(convert_to_time)
        df["Heure fin"] = df["Heure fin"].apply(convert_to_time)

        return df


class WeekSlots(UVTask):
    """Fichier Excel des créneaux hebdomadaires d'une UV.

    Crée un fichier "planning_hebdomadaire.xlsx" dans chaque dossier
    d'UV/UE. Le fichier est prérempli d'après le fichier pdf des
    créneaux de toutes les UVs. Si l'UV/UE n'est pas trouvée, un
    fichier avec en-tête mais sans créneau est créé.

    """

    hidden = True
    unique_uv = False
    target_name = "planning_hebdomadaire.xlsx"
    target_dir = "documents"

    def setup(self):
        super().setup()
        self.target = self.build_target()
        self.uvlist_csv = UtcUvListToCsv.target_from()
        self.file_dep = [self.uvlist_csv]

    def run(self):
        df = pd.read_csv(self.uvlist_csv)
        self.df_uv = df.loc[df["Code enseig."] == self.uv, :]

        output_obj = self.create_excel_file()
        if output_obj.action not in ["abort", "keep"]:
            workbook = load_workbook(filename=self.target)
            self.add_second_worksheet(workbook)

    @staticmethod
    def read_target(week_slots):
        df = pd.read_excel(week_slots, engine="openpyxl")
        if len(df.index) == 0:
            fn = rel_to_dir(week_slots, settings.CWD)
            logger.warning(f"Le fichier `{fn}` est vide")

        if df["Activité"].isnull().any():
            fn = rel_to_dir(week_slots, settings.CWD)
            raise Exception(f"La colonne `Activité` du fichier `{fn}` ne doit pas contenir d'élément vide.")

        rest = set(df["Activité"]) - set(["Cours", "TD", "TP"])
        if rest:
            rest_msg = ", ".join(f"`{e}`" for e in rest)
            fn = rel_to_dir(week_slots, settings.CWD)
            logger.warning("La colonne `Activité` du fichier `%s` contient des libellés non standards (`Cours`, `TD` et `TP`) : %s", fn, rest_msg)

        if df["Jour"].isnull().any():
            fn = rel_to_dir(week_slots, settings.CWD)
            raise Exception(f"La colonne `Jour` du fichier `{fn}` ne doit pas contenir d'élément vide.")

        rest = set(df["Jour"]) - set(["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"])
        if rest:
            rest_msg = ", ".join(f"`{e}`" for e in rest)
            fn = rel_to_dir(week_slots, settings.CWD)
            raise Exception(f"La colonne `Jour` du fichier `{fn}` ne doit contenir que des jours de la semaine, elle contient {rest_msg}")

        if df["Lib. créneau"].isnull().any():
            fn = rel_to_dir(week_slots, settings.CWD)
            raise Exception(f"La colonne `Lib. créneau` du fichier `{fn}` ne doit pas contenir d'élément vide.")

        df["Heure début"] = df["Heure début"].apply(convert_to_time)
        df["Heure fin"] = df["Heure fin"].apply(convert_to_time)

        # Populate "Abbrev" column where possible
        empty = df["Abbrev"].isna() & ~df["Intervenants"].isna()
        df.loc[empty, "Abbrev"] = df.loc[empty, "Intervenants"].apply(convert_author)

        # Warn if abbrev clashes
        for index, group in df.groupby("Abbrev"):
            insts = group["Intervenants"].dropna().unique()
            if len(insts) > 1:
                insts_list = ", ".join(insts)
                fn = rel_to_dir(week_slots, settings.CWD)
                logger.warning(f"Les intervenants suivants ont les mêmes initiales : {insts_list}. "
                               f"Modifier la colonne `Abbrev` dans le fichier `{fn}`.")

        return df

    def create_excel_file(self):
        df = pd.read_csv(self.uvlist_csv)
        self.df_uv = df.loc[df["Code enseig."] == self.uv, :]

        # Test if UV/UE is in listing from UtcUvListToCsv
        if len(self.df_uv) == 0:
            creneau_uv = rel_to_dir(self.settings.CRENEAU_UV, self.settings.CWD)
            logger.warning(
                "L'UV/UE `%s` n'existe pas dans le fichier `%s`, "
                "un fichier Excel sans créneau est créé.",
                self.uv,
                creneau_uv
            )
            columns = [
                "Activité",
                "Jour",
                "Heure début",
                "Heure fin",
                "Locaux",
                "Semaine",
                "Lib. créneau",
                "Intervenants",
                "Abbrev",
                "Responsable",
            ]
            self.df_uv = pd.DataFrame(columns=columns)
        else:
            self.df_uv = self.df_uv.sort_values(["Lib. créneau", "Semaine"])
            self.df_uv = self.df_uv.drop(["Code enseig."], axis=1)
            self.df_uv["Intervenants"] = np.nan
            self.df_uv["Abbrev"] = np.nan
            self.df_uv["Responsable"] = np.nan

        # Write to disk
        with Output(self.target, protected=True) as out:
            self.df_uv.to_excel(out.target, sheet_name="Intervenants", index=False)

        # Return decision in Output
        return out

    def add_second_worksheet(self, workbook):
        worksheet = workbook.create_sheet("Décompte des heures")

        # Make current worksheet the default one, useful for get_address_of_cell
        workbook.active = worksheet

        ref = worksheet.cell(1, 1)
        num_record = len(self.df_uv)
        if num_record == 0:
            num_record = 10

        semAB = not self.df_uv.loc[self.df_uv["Activité"] == "TP", "Semaine"].isna().all()
        has_TP = "TP" in self.df_uv["Activité"].values

        keywords = [
            "Intervenants",
            "Statut",
            "Cours",
            "TD",
            "TP",
            "Heures Cours prév",
            "Heures TD prév",
            "Heures TP prév",
            "UTP",
            "Heure équivalent TD",
            "Heure brute"
        ]

        if not has_TP:
            del keywords[keywords.index("Statut")]
            del keywords[keywords.index("TP")]
            del keywords[keywords.index("Heures TP prév")]

        # Write header
        row_cells = get_row_cells(ref, 0, *keywords)
        headers = {e: e for e in keywords}
        if has_TP and semAB:
            headers["TP"] = "TP A/B"
        fill_row(row_cells, **headers)
        for cell in row_cells.values():
            cell.style = "Pandas"

        # Write rows
        for i in range(num_record):
            row_cells = get_row_cells(ref, i+1, *keywords)
            elts = {
                "Heures Cours prév": lambda row: "=2*16*{}".format(row["Cours"].coordinate),
                "Heures TD prév": lambda row: "=2*16*{}".format(row["TD"].coordinate),
                "Heure équivalent TD": lambda row: "=2/3*{}".format(row["UTP"].coordinate),
            }

            if has_TP:
                elts["Heures TP prév"] = lambda row: "=2*{num_week}*{TP_cell}".format(
                    TP_cell=row["TP"].coordinate,
                    num_week="8" if semAB else "16"
                )
                elts["UTP"] = lambda row: "=2*16*2.25*{cours_cell}+2*16*1.5*{TD_cell}+2*{num_week}*{TP_cell}*{status_cell}".format(
                    cours_cell=row["Cours"].coordinate,
                    TD_cell=row["TD"].coordinate,
                    TP_cell=row["TP"].coordinate,
                    status_cell=row["Statut"].coordinate,
                    num_week="8" if semAB else "16"
                )
                elts["Heure brute"] = lambda row: "=2*16*{cours_cell}+2*16*{TD_cell}+2*{num_week}*{TP_cell}".format(
                    cours_cell=row["Cours"].coordinate,
                    TD_cell=row["TD"].coordinate,
                    TP_cell=row["TP"].coordinate,
                    num_week="8" if semAB else "16"
                )
                elts["Statut"] = 1.5
            else:
                elts["UTP"] = lambda row: "=2*16*2.25*{cours_cell}+2*16*1.5*{TD_cell}".format(
                    cours_cell=row["Cours"].coordinate,
                    TD_cell=row["TD"].coordinate,
                )
                elts["Heure brute"] = lambda row: "=2*16*{cours_cell}+2*16*{TD_cell}".format(
                    cours_cell=row["Cours"].coordinate,
                    TD_cell=row["TD"].coordinate,
                )

            fill_row(row_cells, **elts)

        frame_range(ref, row_cells[keywords[-1]])

        # Write real
        row_cells = get_row_cells(ref, num_record+1, *keywords)
        range_cells = get_range_cells(ref.below(), num_record-1, *keywords)
        row_cells["Cours"].value = "=SUM({})".format(range_cells["Cours"])
        row_cells["TD"].value = "=SUM({})".format(range_cells["TD"])
        if has_TP:
            row_cells["TP"].value = "=SUM({})".format(range_cells["TP"])
        row_cells["Cours"].left().text("Total")

        # Write expected
        n_cours = len(self.df_uv.loc[self.df_uv["Activité"] == "Cours"])
        n_TD = len(self.df_uv.loc[self.df_uv["Activité"] == "TD"])

        row_cells = get_row_cells(ref, num_record+2, *keywords)
        row_cells["Cours"].value = n_cours
        row_cells["TD"].value = n_TD
        if has_TP:
            n_TP = len(self.df_uv.loc[self.df_uv["Activité"] == "TP"])
            row_cells["TP"].value = n_TP
        row_cells["Cours"].left().text("Attendu")

        workbook.save(self.target)


class XlsUTP(TaskBase):
    """Crée un fichier Excel de prévisions des UTP globales."""

    target_dir = "."
    target_name = "UTP.xlsx"

    def setup(self):
        super().setup()
        self.target = self.build_target()

    def write_UV_block(self, ref_cell):
        """Write table at `ref_cell`."""

        header = [
            "Libellé",
            "Type (CM/TD/TP)",
            "Nombre de créneaux de 2h par semaine",
            "Heures",
            "UTP",
            "Heures éq. TD",
            "",
            "CM",
            "TD",
            "TP",
            "",
            "Heures éq. TD CM",
            "Heures éq. TD TD",
            "Heures éq. TD TP",
        ]

        def switch(ifs, default):
            """Nested if conditionals with a default value."""

            if ifs:
                cond, then = ifs[0]
                return "IF({cond}, {then}, {else_})".format(
                    cond=cond,
                    then=then,
                    else_=switch(ifs[1:], default),
                )
            else:
                return default

        formula = switch(
            (
                ("{type_cell} <> {type}", '""'),
                ("NOT(ISBLANK({utp}))", "{utp_result}"),
                ("NOT(ISBLANK({h_eq_tp}))", "{h_eq_tp_result}"),
                ("NOT(ISBLANK({heures}))", "{mult} * {heures}"),
                ("NOT(ISBLANK({slots_2h}))", "2*16*{slots_2h}*{mult}"),
            ),
            ""
        )

        def get_formula(formula, metric, ctype, mult):
            def inner(row):
                return "=" + formula.format(
                    slots_2h=row["Nombre de créneaux de 2h par semaine"].coordinate,
                    heures=row["Heures"].coordinate,
                    utp=row["UTP"].coordinate,
                    h_eq_tp=row["Heures éq. TD"].coordinate,
                    utp_result=row["UTP"].coordinate if metric == "UTP" else "2/3*{}".format(row["UTP"].coordinate),
                    h_eq_tp_result=row["Heures éq. TD"].coordinate if metric == "eqTD" else "3/2*{}".format(row["Heures éq. TD"].coordinate),
                    type=f'"{ctype}"',
                    type_cell=row["Type (CM/TD/TP)"].coordinate,
                    mult=mult
                )
            return inner

        elts = {
            colname: get_formula(formula, metric, ctype, mult)
            for metric, colname, ctype, mult, in (
                    ("UTP", "CM", "CM", 2.25),
                    ("UTP", "TD", "TD", 1.5),
                    ("UTP", "TP", "TP", 1.5),
                    ("eqTD", "Heures éq. TD CM", "CM", 1.5),
                    ("eqTD", "Heures éq. TD TD", "TD", 1),
                    ("eqTD", "Heures éq. TD TP", "TP", 1),
            )
        }

        # Write header at `ref_cell`
        row_cells = get_row_cells(ref_cell, 0, *header)
        fill_row(row_cells, **{e: e for e in header})

        # Write header above
        row_cells2 = get_row_cells(ref_cell, -1, *header)

        row_cells2["Nombre de créneaux de 2h par semaine"].merge(row_cells2["Heures éq. TD"]).center().value = "Charge effectuée"
        frame_range(row_cells2["Nombre de créneaux de 2h par semaine"], row_cells["Heures éq. TD"])

        row_cells2["CM"].merge(row_cells2["TP"]).center().value = "UTP"
        frame_range(row_cells2["CM"], row_cells["TP"])

        row_cells2["Heures éq. TD CM"].merge(row_cells2["Heures éq. TD TP"]).center().value = "Heures"
        frame_range(row_cells2["Heures éq. TD CM"], row_cells["Heures éq. TD TP"])

        # Write formulas in rows
        n_row = 30
        for i in range(n_row):
            row_cells = get_row_cells(ref_cell, i + 1, *header)
            fill_row(row_cells, **elts)

        # Columns on which to make a total
        kw_totals = [
            "CM",
            "TD",
            "TP",
            "Heures éq. TD CM",
            "Heures éq. TD TD",
            "Heures éq. TD TP",
        ]
        first_row = get_row_cells(ref_cell, 1, *header)
        last_row = get_row_cells(ref_cell, n_row, *header)

        # Subtotals
        for kw in kw_totals:
            range_cells = get_range_from_cells(first_row[kw], last_row[kw])
            total_cell = last_row[kw].below()
            total_cell.value = "=SUM({})".format(range_cells)

        # Overall totals
        range_cells = get_range_from_cells(last_row["CM"].below(), last_row["TP"].below())
        last_row["TP"].below().right().value = "=SUM({})".format(range_cells)

        range_cells = get_range_from_cells(last_row["Heures éq. TD CM"].below(), last_row["Heures éq. TD TP"].below())
        last_row["Heures éq. TD TP"].below().right().value = "=SUM({})".format(range_cells)

    def run(self):
        wb = Workbook()
        ws = wb.active

        ref_cell = ws.cell(3, 2)
        self.write_UV_block(ref_cell)

        with Output(self.target, protected=True) as out:
            wb.save(out.target)
